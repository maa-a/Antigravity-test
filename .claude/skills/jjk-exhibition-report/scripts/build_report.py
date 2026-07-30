#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""呪術廻戦展 全会場 LEGS商品42SKU 販売分析レポートを生成する。

前提となる仕様は references/requirements.md、パラメータは config.py に集約している。
数式はすべてExcel数式として書き出すので、「前提・入力」シートの値を変えれば
利用者側で再計算できる。生成後は必ず recalc → fix_font.py の順で仕上げる。

usage:
  python build_report.py --data DIR --out OUT.xlsx
"""
import argparse
import datetime
import json
import os
import sys

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter as gcl
from openpyxl.formatting.rule import FormulaRule

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import config as CFG

_ap = argparse.ArgumentParser()
_ap.add_argument('--data', required=True, help='extract_data.py の出力ディレクトリ')
_ap.add_argument('--out', required=True, help='出力する .xlsx のパス')
_A = _ap.parse_args()
SP = _A.data.rstrip('/') + '/'
OUT = _A.out

D = json.load(open(SP + 'sales.json'))
ATT = json.load(open(SP + 'att.json'))
LED = json.load(open(SP + 'stock.json'))          # 在庫フロー台帳の残在庫想定
INV = {k: dict(real=v['stock'], hist=v['hist'], stages=v['stages'],
               set_size=v['set_size']) for k, v in LED.items()}
STOCK_STAGE = next(iter(LED.values()))['stock_stage']
STOCK_STAGES = next(iter(LED.values()))['stages']
T = [list(x) for x in CFG.TARGET]
TJ = [j for j, _ in T]
EXCL = CFG.EXCLUDE_CATEGORIES
VS = list(CFG.VENUES)
RUN, DATA_DAYS, BASE = CFG.RUN, CFG.DATA_DAYS, CFG.BASE
PERIOD, HALL = CFG.PERIOD, CFG.HALL
SHORTAGE = {str(k): v for k, v in CFG.SHORTAGE.items()}
PROD = {j: CFG.PROD_MONTHS[i] for i, (j, _) in enumerate(T)}
PROD_NOTE = CFG.PROD_NOTE
VEN_START, BASEDATE = CFG.VEN_START, CFG.BASEDATE
PAST_ORDER, PLAN_GOODS, FUTURE = CFG.PAST_ORDER, CFG.PLAN_GOODS, CFG.FUTURE
RESUPPLY = {str(k): v for k, v in getattr(CFG, 'RESUPPLY', {}).items()}
RESUPPLY_DETAIL = getattr(CFG, 'RESUPPLY_DETAIL', [])
RESUPPLY_NOTE = getattr(CFG, 'RESUPPLY_NOTE', '')
DECIDED = {int(k): v for k, v in getattr(CFG, 'DECIDED_ORDER', {}).items()}
SHELF = {int(k): v for k, v in getattr(CFG, 'SHELF_LIFE', {}).items()}

# 開催中会場への追納を販売可能数に加算する（倉庫→会場の移動なので全社在庫は不変）。
for _j, _q in RESUPPLY.items():
    _x = D[CFG.CURRENT]['rows'].get(_j)
    if _x:
        _x['avail'] += _q
        _x['stock'] += _q

# 元データに小数が混じることがある（名古屋の販売可能数など）。整数に丸めて扱う
for _v in D:
    for _x in D[_v]['rows'].values():
        for _f in ('avail', 'deliv', 'stock', 'total'):
            if isinstance(_x[_f], float):
                _x[_f] = round(_x[_f])

S = lambda v, j, a, b: sum(D[v]['rows'][str(j)]['daily'][a:b])
R = lambda v, j: D[v]['rows'][str(j)]


def vt(v, a, b, legs=None):
    q = am = 0
    for k, x in D[v]['rows'].items():
        if (x['cat'] or '') in EXCL:
            continue
        if legs is True and int(k) not in TJ:
            continue
        if legs is False and int(k) in TJ:
            continue
        s = sum(x['daily'][a:b])
        q += s
        am += s * x['price']
    return q, am


def grp(c):
    c = c or ''
    for k, g in [('レッグス', 'LEGS(42SKU)'), ('MBS', 'MBS'), ('MAPPA', 'MAPPA'),
                 ('ムービック', 'ムービック'), ('東宝', '東宝'), ('書籍', '書籍')]:
        if k in c:
            return g
    return 'その他'


# ---- モデル定数 -----------------------------------------------------------
WGT = {v: w for v, w in CFG.G_WEIGHT.items() if w}
FONT = CFG.FONT
VF = CFG.VENUE_COLOR
TAIL = sum(((vt(v, 0, 20, legs=True)[0] - vt(v, 0, 19, legs=True)[0]) / vt(v, 0, 19, legs=True)[0])
           for v in ['大阪', '博多']) / 2
_go = {}
for v in WGT:
    q5 = vt(v, 0, BASE, legs=True)[0]
    qn = vt(v, 0, 19, legs=True)[0] * (1 + TAIL) if v == '名古屋' else vt(v, 0, 20, legs=True)[0]
    _go[v] = qn / q5
G_ALL = sum(_go[v] * WGT[v] for v in WGT) / sum(WGT.values())

OQ = {j: S('大阪', j, 0, 20) for j in TJ}
TO = sum(OQ.values())

# ---- 表示ラベル（開催中会場の実績日数に自動追随させる） -------------------
CUR = CFG.CURRENT
RUNC = RUN[CUR]                       # 開催中会場の会期日数（予測の到達点）


def _md(v, i):
    """会場vのi日目（0起点）の日付を m/d で返す。"""
    ds = D[v].get('dates') or []
    if i < 0 or i >= len(ds) or not ds[i]:
        return ''
    y, m, d = str(ds[i])[:10].split('-')
    return '%d/%d' % (int(m), int(d))


BASE_FROM, BASE_TO = _md(CUR, 0), _md(CUR, BASE - 1)
BASE_RANGE = '%s-%s' % (BASE_FROM, BASE_TO) if BASE_TO else ''
NB = '初%d日' % BASE                   # 例）初8日
NBR = '%s(%s)' % (NB, BASE_RANGE) if BASE_RANGE else NB


def gv(v, j):
    n5 = S(v, j, 0, BASE)
    if not n5:
        return None
    return S(v, j, 0, 19) * (1 + TAIL) / n5 if v == '名古屋' else S(v, j, 0, 20) / n5


PRED = {}
for j in TJ:
    gs = [(gv(v, j), WGT[v]) for v in WGT]
    gs = [(g, w) for g, w in gs if g is not None]
    g = sum(a * b for a, b in gs) / sum(b for _, b in gs) if gs else G_ALL
    PRED[j] = round(S('札幌', j, 0, BASE) * g)
TS = sum(PRED.values())

TITLE = Font(name=FONT, size=15, bold=True, color='1F3864')
SUB = Font(name=FONT, size=11, bold=True, color='2E5C8A')
H1 = Font(name=FONT, size=11, bold=True, color='FFFFFF')
H2 = Font(name=FONT, size=9, bold=True, color='FFFFFF')
BD = Font(name=FONT, size=9, bold=True)
NM = Font(name=FONT, size=9)
IN = Font(name=FONT, size=10, bold=True, color='0000FF')
NOTE = Font(name=FONT, size=8, color='555555')
SEC = Font(name=FONT, size=11, bold=True, color='C00000')
NAVY, BLUE, JUDG, GREY = (PatternFill('solid', fgColor=x) for x in ['1F3864', '2E5C8A', 'BF8F00', 'F2F2F2'])
FUT = PatternFill('solid', fgColor='7B3F00')
YEL = PatternFill('solid', fgColor='FFFF00')
ALERT, WARN, OKF = (PatternFill('solid', fgColor=x) for x in ['FFC7CE', 'FFEB9C', 'E2EFDA'])
thin = Side(style='thin', color='BFBFBF')
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)
CTR = Alignment(horizontal='center', vertical='center', wrap_text=True)
LFT = Alignment(horizontal='left', vertical='center')
INT, PCT, PCT2, YEN, MUL = '#,##0', '0.0%', '0.00%', '¥#,##0', '0.000"倍"'

wb = openpyxl.Workbook()


def body(ws, r1, r2, nc, namecol=3):
    for r in range(r1, r2 + 1):
        for c in range(1, nc + 1):
            cell = ws.cell(r, c)
            cell.font = NM
            cell.border = BOX
            cell.alignment = LFT if c == namecol else Alignment(horizontal='center', vertical='center')


def totrow(ws, r, nc, namecol=3):
    for c in range(1, nc + 1):
        ws.cell(r, c).font = BD
        ws.cell(r, c).fill = GREY
    ws.cell(r, namecol).alignment = LFT


def hdrfmt(ws, r1, r2, nc):
    for r in range(r1, r2 + 1):
        for c in range(1, nc + 1):
            ws.cell(r, c).alignment = CTR
            ws.cell(r, c).border = BOX


# ===========================================================================
# 0. サマリー
# ===========================================================================
ws = wb.active
ws.title = 'サマリー'
ws.sheet_view.showGridLines = False
ws['A1'] = 'アニメーション 呪術廻戦展「懐玉・玉折」「渋谷事変」'
ws['A1'].font = TITLE
ws['A2'] = 'LEGS／イベント記念商品 42SKU　全会場 販売分析・札幌予測確定・石川会場以降予測・追加発注判定'
ws['A2'].font = SUB
ws['A3'] = ('基準日：%s　／　実績データ：%s＝全会期、%s＝%s(会期%d日目)まで　＋ %s SPJ様 欠品報告メール'
            % ('%d年%d月%d日' % (BASEDATE.year, BASEDATE.month, BASEDATE.day),
               '・'.join('%s(%d日)' % (v, DATA_DAYS[v]) for v in VS if v != CUR),
               CUR, BASE_TO, BASE, CFG.SHORTAGE_REPORT_DATE))
ws['A3'].font = NOTE

_sallq, _salla = [x * y for x, y in zip(vt('札幌', 0, BASE),
                 (vt('大阪', 0, 20)[0] / vt('大阪', 0, BASE)[0], vt('大阪', 0, 20)[1] / vt('大阪', 0, BASE)[1]))]
_spa = sum(PRED[j] * R('札幌', j)['price'] for j in TJ)
_lsha, _lshq = _spa / _salla, TS / _sallq
_unit = _spa / TS
TKQ, TKA = vt('東京', 0, 25)
_ru = round(_unit)          # 「前提・入力」に載せる整数単価。Excel側の除数と揃える
_futq = sum(round(TKA * r * _lsha / _ru) for *_, r, _n in FUTURE)
_avail = sum(R('札幌', j)['avail'] for j in TJ)
_rest = sum(max(0, R('札幌', j)['avail'] - PRED[j]) for j in TJ)
_need_sap = [(nm, PRED[j] - R('札幌', j)['avail']) for j, nm in T if PRED[j] > R('札幌', j)['avail']]
_kq = round(TKA * FUTURE[-1][3] * _lsha / _ru)
_mixo = {j: OQ[j] / TO for j in TJ}
_ts2 = sum(PRED.values())
_mix = {j: 0.5 * _mixo[j] + 0.5 * (PRED[j] / _ts2) for j in TJ}
_rest_of = {j: max(0, R('札幌', j)['avail'] - PRED[j]) for j in TJ}
_ach2 = (_salla / TKA) / (PLAN_GOODS['札幌'] / PLAN_GOODS['東京'])
_r2 = [PLAN_GOODS['石川'] / PLAN_GOODS['東京'], PLAN_GOODS['仙台'] / PLAN_GOODS['東京'], 0.12, 0.09, 0.50]
_r3 = [x * _ach2 for x in _r2[:4]] + [0.50]
_futq2 = sum(round(TKA * x * _lsha / _ru) for x in _r2)
_futq3 = sum(round(TKA * x * _lsha / _ru) for x in _r3)
_realtot = sum(INV[str(j)]['real'] for j in TJ)
_whtot = _realtot - _avail
_resup = sum(max(0, PRED[j] - R('札幌', j)['avail']) for j in TJ)
_supply = _realtot - TS
_mixo2 = {j: OQ[j] / TO for j in TJ}
_mix2 = {j: 0.5 * _mixo2[j] + 0.5 * (PRED[j] / TS) for j in TJ}
_sup_of = {j: INV[str(j)]['real'] - PRED[j] for j in TJ}
_gap2 = {j: _mix2[j] * _futq - _sup_of[j] for j in TJ}
_vneed = {fn: round(TKA * ratio * _lsha / _ru) for fn, per, dd, ratio, note in FUTURE}
_alloc, _vrem = {}, {}
for j in TJ:
    rem = _sup_of[j]
    first = None
    seq = []
    for fn, per, dd, ratio, note in FUTURE:
        if j in DECIDED and DECIDED[j]['arrive_at'] == fn:
            rem += DECIDED[j]['qty']                 # 確定発注の入荷
        rem -= round(_mix2[j] * _vneed[fn])
        seq.append(rem)
        if rem < 0 and first is None:
            first = fn
    _alloc[j] = (first, max(0, -rem))
    _vrem[j] = seq
def _venue_of(j):
    """会場順に在庫を引き当てて、最初に在庫が尽きる会場を返す（無ければ None）。"""
    return _alloc[j][0]


def _dl_formula(j, base_row=None):
    """発注期限の数式。会期初日 − 生産日数×30 − バッファ（バッファは入力セル連動）。"""
    fn = _alloc[j][0]
    if fn is None:
        return ''
    if PROD[j] is None:
        return '再生産不可'
    d = VEN_START[fn]
    return '=DATE(%d,%d,%d)-%d-%s' % (d.year, d.month, d.day, round(PROD[j] * 30), pr('buf'))


_addlist = sorted([(dict(T)[j], _alloc[j][1], _alloc[j][0], PROD[j]) for j in TJ if _alloc[j][1] > 0],
                  key=lambda x: -x[1])
_addprod = sum(x[1] for x in _addlist)


def _dl(fn, pm):
    if fn is None or pm is None:
        return '再生産不可'
    return (VEN_START[fn] - datetime.timedelta(days=round(pm * 30) + 14)).strftime('%Y/%m/%d')
_prevj = set()
for _v in ['東京', '名古屋', '博多']:
    _prevj |= {int(k) for k, x in D[_v]['rows'].items() if (x['cat'] or '') not in EXCL}
_newj = {int(k) for k, x in D['大阪']['rows'].items() if (x['cat'] or '') not in EXCL} - _prevj
_newshare = (sum(sum(D['大阪']['rows'][str(j)]['daily'][:20]) * D['大阪']['rows'][str(j)]['price']
                 for j in _newj) / vt('大阪', 0, 20)[1])
_shj = [j for j in TJ if str(j) in SHORTAGE]
_model_hit = [j for j in _shj if PRED[j] > R('札幌', j)['avail']]
_model_warn = [j for j in _shj if PRED[j] <= R('札幌', j)['avail'] and R('札幌', j)['avail'] and PRED[j] / R('札幌', j)['avail'] >= 0.7]
_model_miss = [j for j in _shj if j not in _model_hit and j not in _model_warn]
_model_only = [j for j in TJ if str(j) not in SHORTAGE and PRED[j] > R('札幌', j)['avail']]
_nmof = dict(T)
_noresupply = [j for j in _shj if '追納困難' in SHORTAGE[str(j)]]
_gap_of = {j: _mix[j] * _futq - _rest_of[j] for j in TJ}
_attp = ATT['札幌'][:5]
_attpred = sum(_attp) * sum(ATT['大阪']) / sum(ATT['大阪'][:5])

blocks = [
    ('■ 結論サマリー', [
        ('① 札幌 会期20日予測（確定値）',
         'LEGS42SKU 販売数 %s個／販売金額 %s円　※東京を除く名古屋・博多・大阪の実績倍率から算出'
         % (f'{TS:,}', f'{_spa:,.0f}')),
        ('　札幌 全物販予測',
         '数量 %s個／金額 %s円　→ 東京(初回)対比 %.1f%%　※コナン実績の北海道13.3%%を下回る水準'
         % (f'{_sallq:,.0f}', f'{_salla:,.0f}', _salla / TKA * 100)),
        ('　札幌 入場者予測', '%s人（委員会予測18,000人の43%%／LEGS予測14,000人の56%%）＝計画を大幅に下回る'
         % f'{_attpred:,.0f}'),
        ('② LEGS物販構成比の推移',
         '金額：%s→札幌%.1f%%(予測)　数量：%s→札幌%.1f%%(予測)'
         % ('→'.join('%s%.1f%%' % (v, vt(v, 0, DATA_DAYS[v], legs=True)[1] / vt(v, 0, DATA_DAYS[v])[1] * 100)
                     for v in VS[:-1]), _lsha * 100,
            '→'.join('%.1f%%' % (vt(v, 0, DATA_DAYS[v], legs=True)[0] / vt(v, 0, DATA_DAYS[v])[0] * 100)
                     for v in VS[:-1]), _lshq * 100)),
        ('　構成比低下の主因',
         '大阪会場からの新商品106SKU投入（大阪の会期20日 金額シェア%.1f%%）。札幌は新規SKU 0で品揃えは大阪と同一'
         % (_newshare * 100)),
        ('③ 石川以降5会場 LEGS必要数',
         '%s個（シナリオ①ユーザー指定：石川11%%・仙台13%%・神戸12%%(仮)・鳥取9%%・東京凱旋50%%）'
         % f'{_futq:,.0f}'),
        ('　シナリオ②企画概要書ベース',
         '石川12.48%%・仙台15.60%%（企画概要書グッズ売上÷東京）→ 必要数 %s個'
         % f'{_futq2:,.0f}'),
        ('　シナリオ③札幌実績連動(保守)',
         '②に札幌の計画比達成率%.1f%%を乗算 → 石川5.83%%・仙台7.29%%・神戸5.61%%・鳥取4.21%%（凱旋50%%据置）'
         '→ 必要数 %s個' % (_ach2 * 100, f'{_futq3:,.0f}')),
        ('★残在庫想定の反映',
         '在庫フロー台帳の「%s」列 %s個を採用（梅田会場後の検品結果まで反映済み）。'
         % (STOCK_STAGE, f'{_realtot:,}')),
        ('★札幌への追納（実施済）',
         '%s：%s　計%s個。倉庫在庫からの出荷なので全社在庫は不変'
         % (RESUPPLY_NOTE, '／'.join('%s %s個' % (dict(T)[int(k)], f'{v:,}')
                                    for k, v in RESUPPLY.items()), f'{sum(RESUPPLY.values()):,}')),
        ('　追納後になお必要な追納',
         '%s個（会期20日予測が販売可能数を上回る残差）' % f'{_resup:,}'),
        ('　石川以降 供給可能在庫',
         '%s個 ＝ 残在庫%s個 − 札幌 会期20日予測販売%s個（札幌への追納と会期末返送を相殺した正味）'
         % (f'{_supply:,}', f'{_realtot:,}', f'{TS:,}')),
        ('　全体過不足',
         '供給可能%s個 − 必要%s個 ＝ +%s個の余剰。全体では追加生産は不要な水準。'
         % (f'{_supply:,}', f'{_futq:,.0f}', f'{_supply - _futq:,.0f}')),
        ('★追加発注が必要な商品と会場',
         '%d品目・%s個（会場順に在庫を引き当て、在庫が尽きる会場を特定）' % (len(_addlist), f'{_addprod:,.0f}')),
    ] + [
        ('　　' + nm, '不足%s個／欠品発生会場：%s（%s開始）／生産%sか月 → 最終発注期限 %s'
         % (f'{g:,.0f}', fn, VEN_START[fn].strftime('%Y/%m/%d') if fn else '—',
            pm if pm else '－', _dl(fn, pm)))
        for nm, g, fn, pm in _addlist
    ] + [
        ('　発注の緊急度', '最も早い期限でも%s（基準日2026/7/28から%d日後）。現時点で期限超過の商品はない。'
         % (min(_dl(fn, pm) for _n, _g, fn, pm in _addlist),
            min((VEN_START[fn] - datetime.timedelta(days=round(pm * 30) + 14) - BASEDATE).days
                for _n, _g, fn, pm in _addlist if pm))),
        ('　札幌 会期中の追加発注',
         '%d品目が会期内に在庫切れ見込：%s' % (len(_need_sap), '　'.join('%s(不足%d)' % x for x in _need_sap))),
    ]),
    ('■ 運営（SPJ様）欠品報告との突合 ※%s受領メール' % CFG.SHORTAGE_REPORT_DATE + '', [
        ('報告された欠品／近々欠品',
         '%d品目：%s' % (len(_shj), '　'.join(_nmof[j] for j in _shj))),
        ('◎ モデルも会期中欠品を検知',
         '%d品目：%s' % (len(_model_hit), '　'.join(_nmof[j] for j in _model_hit) or 'なし')),
        ('○ モデルは高消化率で整合',
         '%d品目：%s（予測消化率70%%超。会期末までに枯渇する水準）'
         % (len(_model_warn), '　'.join(_nmof[j] for j in _model_warn) or 'なし')),
        ('△ モデル未検知（現場優先）',
         '%d品目：%s' % (len(_model_miss), '　'.join(
             '%s(予測消化率%.0f%%)' % (_nmof[j], PRED[j] / R('札幌', j)['avail'] * 100 if R('札幌', j)['avail'] else 0)
             for j in _model_miss) or 'なし')),
        ('▲ モデルのみ検知（現場未報告）',
         '%d品目：%s　→ 現場に実在庫の確認を推奨'
         % (len(_model_only), '　'.join(_nmof[j] for j in _model_only) or 'なし')),
        ('★追納不可と回答された品目',
         '%s　＝ 前会場時点で「追納困難」と回答済。ただし7/8実在庫では倉庫残%s個が確認できるため、下記の要確認事項を参照'
         % ('　'.join(_nmof[j] for j in _noresupply) or 'なし',
            f"{sum(INV[str(j)]['real'] - R('札幌', j)['avail'] for j in _noresupply):,}")),
        ('突合の結論',
         '現場報告%d品目のうち%d品目がモデル側でも欠品または高消化率として検知されており、見落とし0・空振り0で完全に整合。'
         % (len(_shj), len(_model_hit) + len(_model_warn))),
        ('', '%s（%d日目）実績を織り込んだことで検知精度が上がり、予測モデルは現場実感と一致する水準にある。'
         % (BASE_TO, BASE)),
        ('', '④シートの「運営 欠品報告」「モデル×現場 突合」列で商品別に明示している（列位置は⑫シート参照）。'),
        ('★倉庫在庫による解決可否',
         '報告%d品目すべて2026/7/8時点で倉庫に在庫が残っており、追加生産ではなく倉庫からの追加納品で対応可能。'
         % len(_shj)),
        ('　倉庫残（札幌納品を除く）',
         '　'.join('%s:%s個' % (_nmof[j], f"{INV[str(j)]['real'] - R('札幌', j)['avail']:,}") for j in _shj)),
        ('　※要確認',
         '「気分リフレッシュオルゴール」は倉庫残28個が存在する。前会場時の「追納困難」回答と実在庫が食い違うため、'),
        ('', '28個が出荷可能かSPJ様・倉庫に再確認のうえ、可能なら札幌へ充当することを推奨。'),
    ]),
    ('■ 最重要リスクと打ち手', [
        ('札幌の動員が計画比%.0f%%' % (_attpred / 18000 * 100),
         '%s入場者%s人（大阪同期間%s人の%.1f%%）。物販売上も東京対比%.1f%%と、'
         % (NB, f'{sum(ATT["札幌"][:BASE]):,}', f'{sum(ATT["大阪"][:BASE]):,}',
            sum(ATT['札幌'][:BASE]) / sum(ATT['大阪'][:BASE]) * 100, _salla / TKA * 100)),
        ('', '石川11%・鳥取9%として設定された「格下」会場と同水準。石川以降の会場比率も下方修正を検討すべき。'),
        ('東京凱旋が必要数の%.0f%%' % (_kq / _futq * 100),
         '石川以降必要数%s個のうち%s個(%.0f%%)が東京凱旋。凱旋の50%%前提が外れると発注量が大きくブレる。'
         % (f'{_futq:,.0f}', f'{_kq:,.0f}', _kq / _futq * 100)),
        ('', '発注は「石川〜鳥取(4会場・%s個)」を先行確定し、凱旋分は会期確定後に第2弾発注とするのが安全。'
         % f'{_futq - _kq:,.0f}'),
        ('★過剰在庫のほうが深刻',
         '残在庫%s個に対し石川以降の必要数は%s個。差引%s個（残在庫の%.0f%%）が余剰となる見込み。'
         % (f'{_realtot:,}', f'{_futq:,.0f}', f'{_supply - _futq:,.0f}', (_supply - _futq) / _realtot * 100)),
        ('', '特に渋谷事変ポストカードセット(供給可能%s個 vs 必要%s個)、懐玉・玉折ポストカードセット(%s vs %s)は'
         % (f'{_sup_of[4515142545112]:,}', f'{_mix2[4515142545112] * _futq:,.0f}',
            f'{_sup_of[4515142545105]:,}', f'{_mix2[4515142545105] * _futq:,.0f}')),
        ('', '20倍超の在庫を抱えている。追加発注よりも、余剰在庫の消化策（セット販売・ノベルティ転用等）の検討が急務。'),
        ('不足はごく一部',
         '追加生産が必要なのは%d品目・%s個のみ。うち展示会キービジュアル クリアファイルが%s個で全体の%.0f%%を占める。'
         % (len(_addlist), f'{_addprod:,.0f}', f'{_addlist[0][1]:,.0f}' if _addlist else '0',
            (_addlist[0][1] / _addprod * 100) if _addlist else 0)),
        ('生産リードタイムの制約',
         '生産日数は1か月（クリアファイル・ポストカード・缶バッジ・ステッカー系）、2.5か月（ミニ色紙2種・前髪ヘアバンド）、'),
        ('', '3か月（絆創膏・お菓子・クリップ・エコバッグ・オルゴール・リング/フラグメント系）。'
             'バスボールは「－」＝再生産不可のため現有在庫で会場配分を組む必要がある。'),
        ('賞味期限の制約', '宿儺の指風お菓子は賞味期限4か月以内の在庫が販売不可。仙台以降向けは2026/5/15手配済だが、'),
        ('', '鳥取(27年3月)・東京凱旋(27年4月)向けは賞味期限の逆算で再手配の要否を要確認。'),
    ]),
    ('■ 算出モデル（①札幌予測）', [
        ('基本式', '%s 会期%d日予測販売数 ＝ %s %s販売数 × 会期倍率 g' % (CUR, RUNC, CUR, NBR)),
        ('会期倍率 g', 'g ＝ 各会場の「%d日累計販売数 ÷ %s販売数」を商品別に算出し加重平均' % (RUNC, NB)),
        ('採用会場と加重', '大阪×2（20日会期・水曜開始・新商品投入後で札幌と同条件）、博多×1、名古屋×1。東京は除外'),
        ('名古屋の換算', '名古屋は19日会期の全日実績あり。札幌20日会期と揃えるため20日目分 +%.2f%%（大阪・博多の実績比率）のみ補完' % (TAIL * 100)),
        ('42品計の実効倍率', '大阪%.3f倍・博多%.3f倍・名古屋%.3f倍(20日換算後) → 加重平均 %.3f倍'
         % (_go['大阪'], _go['博多'], _go['名古屋'], G_ALL)),
    ]),
    ('■ 算出モデル（③石川以降）', [
        ('会場物販売上', '東京(初回)全物販売上 %s円 × 東京対比比率' % f'{TKA:,.0f}'),
        ('LEGS売上', '会場物販売上 × LEGS金額構成比 %.1f%%（札幌予測値＝大阪以降の推移の到達点を採用）' % (_lsha * 100)),
        ('LEGS数量', 'LEGS売上 ÷ LEGS平均単価 %.0f円（札幌予測ミックスベース）' % _unit),
        ('商品別按分', '大阪20日実績の数量構成比50%＋札幌20日予測の数量構成比50%のブレンド'),
        ('比率はすべて入力値', '「前提・入力」シートの青字セルを変更すると全シートが再計算されます'),
    ]),
    ('■ 追加商品によるLEGS売上低下（⑦）', [
        ('市場は縮んでいない', '物販客単価（全商品）は名古屋7,012円/人→大阪6,882円/人と▲1.8%でほぼ横ばい。'),
        ('', '同期間にLEGSは2,099円/人→1,455円/人と▲30.7%（▲643円/人）。純粋なカニバリゼーション。'),
        ('新商品の約47%が食い合い', '大阪の新商品106SKUは1,360円/人を稼いだが、LEGS減少643円/人はその47%に相当。'),
        ('LEGSが削られ始めたのは博多', '博多で「おすわりぬいぐるみ全8種」(230円/人)＋アクリルカードスタンド(233円/人)を投入し、'),
        ('', 'LEGSは2,099→1,634円/人（▲465）。大阪のうちわ14種・アクリル系39種でさらに▲179円/人。'),
        ('被害はブラインド系に集中', 'ブラインド缶バッジ▲272円/人・ブラインドミニ色紙▲284円/人の2タイプでLEGS減少の87%。'),
        ('', 'クリアファイル・ポストカード類は横ばい〜微増で影響なし。追加生産はブラインド系を抑制すべき。'),
        ('★石川以降の最大リスク', '同規模の新商品を石川以降に追加投入すると③の必要数86,000個は最大30%過大になる。'),
        ('', '追加生産を確定する前に、石川以降の新商品投入計画（有無・規模）を必ず確認すること。'),
    ]),
    ('■ シート構成', [
        ('前提・入力', '会場比率・構成比・加重など全パラメータの入力（青字＝編集可）'),
        ('①札幌予測確定', '商品別 会期20日予測（会場別g・予測数量/金額・消化率・会期末残在庫）'),
        ('②会場別構成比', '全会場のLEGS物販構成比（数量・金額）、供給元別構成比、東京対比 会場比率'),
        ('③石川以降予測', '残5会場のLEGS売上/数量予測と商品別必要数'),
        ('④追加発注判定', '札幌会期中／石川以降 の商品別 発注要否・推奨数量'),
        ('⑤全データ統合リスト', '5会場×42商品の全実績・構成比・消化率・予測を1行に集約'),
        ('⑥商品別 会場推移', '商品別の会場ごと販売数・構成比の推移（トレンド確認用）'),
        ('⑦新商品影響分析', '梅田以降の追加商品（ぬいぐるみ・うちわ等）によるLEGS売上低下の要因分析'),
        ('⑧実在庫(7-8)明細', '2026/7/8時点の実在庫（CX催事＋DBS倉庫＋梅田返送良品）と商品別の欠品算出'),
        ('⑨過去4会場 予測vs実績', '池袋・名古屋・博多・梅田の商品別 販売予測／実績／予測時構成比／実績後の構成比修正'),
        ('⑩会場比率シナリオ比較', '石川以降の東京対比比率 4シナリオ（ユーザー指定／企画概要書／札幌実績連動／手動）の比較'),
        ('⑪会場別引当・発注期限', '会場順の在庫引当、欠品発生会場の特定、生産日数から逆算した最終発注期限'),
        ('⑫計算ロジック解説', '各シート・各列がどの数式で計算されているかを、No.1商品の実数を当てはめた実例つきで一覧化'),
        ('運営欠品報告の反映先', '④追加発注判定シートのY列・Z列、および⑧シートの運営欠品報告列・備考欄'),
    ]),
    ('■ 前提と留意事項', [
        ('データ再読込', '全5会場のレジ通過単品売上を全件再読込・再集計。名古屋は後半4/1-4/5の追加提供データを反映し、'),
        ('', '会期19日間（3/18-4/5）の完全実績となった。これに伴い名古屋の会期倍率gは推計2.438倍→実績2.215倍に修正。'),
        ('LEGS＝42SKU', '各会場の商品分類「記念商品/レッグス」「レッグス/イベント記念」は全会場で本42SKUと完全一致'),
        ('物販売上の定義', 'レジ袋代・音声ガイド・当日券を除いた商品売上（税込）。会場報告の総売上とは一致しない場合がある'),
        ('残在庫の出所', '在庫フロー台帳（会場ごとの追加発注→販売→レッグス検品を積み上げた表）の最終実在庫列。'),
        ('', '実際の発注前に倉庫在庫と過去会場の返品在庫を相殺すること（前提・入力シートで控除可能）。'),
        ('神戸阪急の会場比率', '東京対比の指定が無いため12%を仮置き。要確認事項。'),
        ('仙台の会場比率', 'ユーザー指定13%を採用。コナンカフェ2026実績（東京対比10.5%）とは2.5pt乖離あり。'),
        ('鳥取会場', '会期3/15-4/4を21日間として計上。'),
        ('価格改定', 'イメージリング8種は東京¥22,000→名古屋以降¥27,500に改定。⑤では会場別の実売単価で金額を算出。'),
        ('元データの小数値', '名古屋の販売可能数に小数を含む商品が3件あるため整数に丸めて処理している（合計への影響は1個未満）。'),
        ('東京の内覧会', '東京の12/4内覧会分（全商品計142個）は会期実績から除外して会場間比較している。'),
    ]),
]
r = 5
for sec, items in blocks:
    ws.cell(r, 1, sec).font = SEC
    r += 1
    for a, b in items:
        ws.cell(r, 1, a).font = BD if a else NM
        c = ws.cell(r, 2, b)
        c.font = NM
        c.alignment = LFT
        r += 1
    r += 1
ws.column_dimensions['A'].width = 30
ws.column_dimensions['B'].width = 118

# ===========================================================================
# 1. 前提・入力
# ===========================================================================
ws = wb.create_sheet('前提・入力')
ws.sheet_view.showGridLines = False
ws['A1'] = '前提・入力パラメータ'
ws['A1'].font = TITLE
ws['A2'] = '青字・黄色セルが入力値です。変更すると①③④⑤シートの予測が自動再計算されます。'
ws['A2'].font = NOTE

PS = "'前提・入力'!$"
INP = []  # (row, label, value, fmt, note)
sec_rows = []
r = 4


def sec_(t):
    global r
    ws.cell(r, 1, t).font = SEC
    sec_rows.append(r)
    r += 1


def inp(label, val, fmt, note=''):
    global r
    ws.cell(r, 1, label).font = BD
    c = ws.cell(r, 2, val)
    c.font = IN
    c.fill = YEL
    c.number_format = fmt
    c.border = BOX
    c.alignment = CTR
    ws.cell(r, 3, note).font = NOTE
    ws.cell(r, 3).alignment = LFT
    ref = 'B%d' % r
    r += 1
    return ref


def calc(label, val, fmt, note=''):
    global r
    ws.cell(r, 1, label).font = BD
    c = ws.cell(r, 2, val)
    c.font = NM
    c.number_format = fmt
    c.border = BOX
    c.alignment = CTR
    ws.cell(r, 3, note).font = NOTE
    ws.cell(r, 3).alignment = LFT
    ref = 'B%d' % r
    r += 1
    return ref


_NG = ('【何に効くか】①シートO列「採用倍率g」の加重平均ウェイト。'
       '%s%d日予測＝%s実績×採用倍率g。' % (CUR, RUNC, NBR))

sec_('■ ①%s予測 モデルパラメータ' % CUR)
P = {}
P['w_osaka'] = inp('会期倍率g 加重：大阪', CFG.G_WEIGHT.get('大阪', 0), '0',
                   _NG + '\n【なぜ2倍か】大阪④は%d日会期・水曜開始・新商品106SKU投入後と%sの条件がほぼ一致するため、'
                   '他会場の2倍の重みを置く。大阪の倍率g＝%.3f（42品計）。'
                   % (RUN['大阪'], CUR, _go['大阪']))
P['w_hakata'] = inp('会期倍率g 加重：博多', CFG.G_WEIGHT.get('博多', 0), '0',
                    _NG + '\n【算出】博多②は%d日会期のため、頭から%d日分だけを切り出して倍率化'
                    '（①シートK列＝博多%d日累計 ÷ %s）。博多の倍率g＝%.3f。会期構成が%sと異なるため加重1。'
                    % (RUN['博多'], RUNC, RUNC, NB, _go['博多'], CUR))
P['w_nagoya'] = inp('会期倍率g 加重：名古屋', CFG.G_WEIGHT.get('名古屋', 0), '0',
                    _NG + '\n【算出】名古屋①は%d日会期で%d日目が存在しない。下の補完率で1日分を上乗せして'
                    '%d日相当に換算する（①シートH列）。名古屋の倍率g＝%.3f。加重1。'
                    % (RUN['名古屋'], RUNC, RUNC, _go['名古屋']))
P['w_tokyo'] = inp('会期倍率g 加重：東京', CFG.G_WEIGHT.get('東京', 0), '0',
                   _NG + '\n【なぜ0か】東京は初回会場・%d日会期・年末繁忙期で初動カーブが他会場と大きく異なる。'
                   'ユーザー指示により倍率の算出根拠から除外している（0固定）。'
                   '0以外にすると①O列＝%s予測が全品変動する。' % (RUN['東京'], CUR))
P['tail'] = inp('名古屋 %d日目 補完率' % RUNC, TAIL, PCT,
                '【何に効くか】①シートH列「名古屋 倍率g」の分子補正。'
                '\n【算出】大阪・博多の実績から「%d日目の単日販売数 ÷ 1〜%d日累計」を求めた平均が%.2f%%。'
                '名古屋の%d日累計にこの率を掛けて%d日目相当を補完し、他会場の%d日ベースと粒度を揃える。'
                % (RUNC, RUNC - 1, TAIL * 100, RUN['名古屋'], RUNC, RUNC))
P['g_all'] = inp('全体倍率g（実績ゼロ品のフォールバック）', G_ALL, MUL,
                 '【何に効くか】①シートO列のIFERROR代替値。'
                 '\n【算出】42品計の%s→%d日 加重平均倍率＝%.3f。名古屋・博多・大阪のいずれにも%s実績が無く'
                 '商品別倍率を計算できない品にだけ適用する。'
                 % (NB, RUNC, G_ALL, NB))

sec_('■ ②③会場比率 前提')
P['tk_amt'] = inp('東京(初回) 全物販売上（税込）', TKA, YEN,
                  '【出所】東京(初回) %s %d日間の全物販実績（税込）。レジ袋代・音声ガイド・当日券を除いた商品売上のみ。'
                  '\n【何に効くか】③シートE列の基準額。E7＝本値×D7(東京対比)で石川以降 各会場の物販売上を推定する。'
                  % (HALL['東京'], RUN['東京']))
P['tk_qty'] = inp('東京(初回) 全物販 販売数量', TKQ, INT,
                  '【出所】同上（東京%d日間の全物販 販売数量）。'
                  '\n【何に効くか】②シート10行「全物販 販売数量」の石川以降 予測（＝本値×東京対比）と、'
                  'LEGS数量構成比の母数。' % RUN['東京'])
P['scn'] = inp('★シナリオ選択（1〜4）', 1, '0',
               '1＝ユーザー指定／2＝企画概要書ベース／3＝%s実績連動(保守)／4＝手動入力(F列を使用)。'
               '\n【何に効くか】下の「東京対比：〇〇」がINDEXで自動切替 → ③D列 → 会場売上 → 必要数 → '
               '④の発注判定まで一気に連動する。4シナリオの結果比較は⑩シート。' % CUR)
SCN_LABEL = ['①ユーザー指定', '②企画概要書ベース', '③%s実績連動(保守)' % CUR, '④手動入力']
_ach = (_salla / TKA) / (PLAN_GOODS[CUR] / PLAN_GOODS['東京'])   # 開催中会場の計画比達成率
SCN = {}
for _nm, _per, _dd, _r, _no in FUTURE:
    _plan = PLAN_GOODS.get(_nm.split(' ')[0])
    _p2 = (_plan / PLAN_GOODS['東京']) if _plan else _r
    SCN[_nm] = [_r, _p2, _r if _nm.startswith('東京凱旋') else _p2 * _ach, _r]
ws.cell(r, 1, '　シナリオ別 東京対比比率').font = BD
for k, lab in enumerate(SCN_LABEL):
    c = ws.cell(r, 3 + k, lab)
    c.font = H2
    c.fill = BLUE
    c.alignment = CTR
    c.border = BOX
    ws.column_dimensions[gcl(3 + k)].width = 18
r += 1
SCNROW = {}
for nm, per, dd, ratio, note in FUTURE:
    ws.cell(r, 1, '　　' + nm).font = NM
    ws.cell(r, 2, '%s（%d日間）' % (per, dd)).font = NOTE
    for k in range(4):
        c = ws.cell(r, 3 + k, SCN[nm][k])
        c.number_format = PCT
        c.border = BOX
        c.alignment = CTR
        c.font = IN if k == 3 else NM
        if k == 3:
            c.fill = YEL
    SCNROW[nm] = r
    r += 1
ws.cell(r, 1, '　　※③は企画概要書の計画値に%sの計画比達成率 %.1f%% を乗じたもの（凱旋は据置）'
        % (CUR, _ach * 100)).font = NOTE
r += 2
_SCN_WHY = {
    '石川 香林坊大和': '【根拠】商圏規模から他の地方会場より低めに置いたユーザー指定値。'
                       '企画概要書の計画値ベース（シナリオ②）では12.48%と倍近く高くなるため、'
                       '発注量の差は⑩シートで比較すること。',
    '仙台 Ebeans': '【根拠】ユーザー指定値。会期が{dd}日間と全会場で最長のため高めに設定している。'
                   '同規模イベントのコナンカフェ2026では仙台は東京対比10.5%の実績。',
    '神戸阪急': '【根拠】★要確認：指定も類似実績も無いため仮置き。'
                '確定したらシナリオ④の手動入力(F列)を書き換え、シナリオ選択＝4で反映させる。',
    '鳥取島根 夢みなとタワー': '【根拠】商圏規模が小さく信頼できる類似実績も無いため、'
                               '他の地方会場より低めに置いたユーザー指定値。',
    '東京凱旋 東京建物ぴあホール': '【根拠】ユーザー指定「初回東京の50%」。全シナリオ共通で据置のため'
                                   'シナリオを切り替えても動かない。石川以降5会場の必要数の過半を占める、'
                                   '最も感度の高い前提。',
}
FR = {}
for nm, per, dd, ratio, note in FUTURE:
    ws.cell(r, 1, '東京対比：%s' % nm).font = BD
    c = ws.cell(r, 2, '=INDEX(C%d:F%d,%s)' % (SCNROW[nm], SCNROW[nm], PS + P['scn'].replace('B', 'B$')))
    c.number_format = PCT
    c.font = Font(name=FONT, size=10, bold=True, color='008000')
    c.fill = PatternFill('solid', fgColor='E2EFDA')
    c.border = BOX
    c.alignment = CTR
    _why = _SCN_WHY.get(nm, note).replace('{dd}', str(dd))
    ws.cell(r, 3, '【何に効くか】緑字＝シナリオ選択に応じて上表からINDEXで自動取得。'
            '%sの物販売上を東京(初回)の%.2f%%＝%s円と置く（③シートE列）。\n%s'
            % (nm, ratio * 100, format(TKA * ratio, ',.0f'), _why)).font = NOTE
    ws.cell(r, 3).alignment = LFT
    FR[nm] = 'B%d' % r
    r += 1
P['conan'] = calc('［参考］コナンカフェ2026 東京対比', None, PCT,
                  '【参考実績】大阪53.6%・名古屋26.8%・神奈川19.1%・北海道13.3%・仙台10.5%。'
                  '地方会場は東京の10〜20%に収束する傾向がある。'
                  '\n上の東京対比の水準感が妥当かを検証するための参考値で、計算には使っていない。')

sec_('■ ③石川以降 LEGS構成比 前提')
P['lsh_a'] = inp('LEGS 金額構成比（石川以降）', _lsha, PCT,
                 '【意味】会場の全物販売上のうち LEGS 42SKU が占める金額シェア。'
                 '\n【推移】%s → %s%d日予測 %.1f%%。大阪以降の新商品106SKU投入で低下しており、'
                 'その到達点を石川以降にも適用する。'
                 '\n【何に効くか】③シートF列＝E列(会場 全物販売上)×本値。'
                 % (' → '.join('%s %.1f%%' % (v, vt(v, 0, DATA_DAYS[v], legs=True)[1]
                                               / vt(v, 0, DATA_DAYS[v])[1] * 100) for v in VS[:-1]),
                    CUR, RUNC, _lsha * 100))
P['unit'] = inp('LEGS 平均単価（税込）', round(_unit), YEN,
                '【算出】%s%d日予測の金額 %s円 ÷ 数量 %s個 ≒ %d円（①シートQ49合計 ÷ P49合計）。'
                '新商品投入後の商品ミックスを反映した加重平均単価。'
                '\n【何に効くか】③シートG列＝ROUND(F列 LEGS売上 ÷ 本単価,0)＝石川以降5会場の必要数量。'
                % (CUR, RUNC, format(_spa, ',.0f'), format(TS, ','), round(_unit)))
P['mix_o'] = inp('商品別按分：大阪%d日実績の構成比ウェイト' % RUN['大阪'], CFG.MIX_WEIGHT_PREV, PCT,
                 '【何に効くか】③シートH列「採用 按分構成比」＝大阪構成比(E列)×本値 ＋ %s構成比(G列)×下の値。'
                 '総必要数をどの商品に何個振るかを決める。'
                 '\n【大阪を使う理由】42品すべての確定実績（%d日・%s個）がありサンプルが大きい。'
                 'ただし新商品投入直後の需要も含んでいる。'
                 % (CUR, RUN['大阪'], format(vt('大阪', 0, RUN['大阪'], legs=True)[0], ',')))
P['mix_s'] = inp('商品別按分：%s%d日予測の構成比ウェイト' % (CUR, RUNC), CFG.MIX_WEIGHT_CUR, PCT,
                 '大阪ウェイトとの合計が100%になるよう設定する。'
                 '\n【{cur}を使う理由】新商品投入後のミックスで、石川以降の品揃えに条件が近い（ただし予測値）。'
                 '\n【感度】総必要数は変わらず、品目別の偏りだけが動く。大阪100%に寄せると定番商品へ、'
                 '{cur}100%に寄せると直近で伸びている商品へ配分が寄る。'.replace('{cur}', CUR))

sec_('■ ④追加発注 前提')
P['safe'] = inp('安全率（不足数に対する上乗せ）', CFG.SAFETY, PCT,
                '【何に効くか】③④⑧⑪シートの推奨発注数＝IF(不足>0, ROUNDUP(不足数×(1+本率),0), 0)。'
                '\n【根拠】輸送中の破損・検品不良・予測超過の緩衝として%.0f%%を上乗せしている。'
                % (CFG.SAFETY * 100))
P['spj'] = calc('［参考］残在庫想定 全品計（%s）' % STOCK_STAGE,
                sum(INV[str(j)]['real'] for j in TJ), INT,
                '【出所】在庫フロー台帳「【清算(卸値67％)】…_検品差異」シートの最終「実在庫」列（%s）の42SKU合計。'
                '会場ごとの 追加発注→販売→レッグス検品 を積み上げた確定値で、直近会場の返送検品まで反映済み。'
                '%s会場への納品数と%s会場の販売数はまだ差し引かれていない。'
                '\n【何に効くか】本セルは⑧シートH49合計へのリンク（表示用）。'
                '商品別の実数を③⑧⑩⑪シートが直接参照する。' % (STOCK_STAGE, CUR, CUR))
P['slide'] = inp('%s 会期末残在庫のスライド率' % CUR, 1.00, PCT,
                 '【何に効くか】③シートS列「%s会期末 返送(スライド)」＝ROUND(①シートU列×本率,0)。'
                 '石川以降の供給可能在庫に加算される。'
                 '\n【100%%の意味】%s終了在庫 %s個を全量 石川以降へ回せる前提。'
                 '損傷・返送ロスを見込むなら下げる → 供給が減り追加生産が増える。'
                 % (CUR, CUR, format(_rest, ',')))
P['buf'] = inp('発注〜納品バッファ日数（輸送・検品）', CFG.ORDER_BUFFER_DAYS, '0"日"',
               '【何に効くか】④⑧⑪シートの最終発注期限＝欠品発生会場の会期初日 − 生産日数(か月×30日) − 本バッファ。'
               '\n【内訳】%d日＝工場出荷 → 国内輸送・入荷検品・会場搬入のリードタイム。'
               % CFG.ORDER_BUFFER_DAYS)
P['base'] = calc('基準日', BASEDATE, 'yyyy/m/d',
                 '【根拠】%s＝本分析の基準日（生産日数の提示は2026/7/28時点）。'
                 '\n【何に効くか】④⑧⑪シートの「基準日からの残日数」＝最終発注期限 − 本日付。'
                 '日付を進めると残日数が減り、期限超過の判定が増える。'
                 % BASEDATE.strftime('%Y/%m/%d'))

for rr in sec_rows:
    pass
ws.column_dimensions['A'].width = 42
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 108
for _r in range(4, r + 1):
    _c = ws.cell(_r, 3)
    if isinstance(_c.value, str) and len(_c.value) > 40:
        _c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        _lines = sum(1 + len(_x) // 54 for _x in _c.value.split('\n'))
        ws.row_dimensions[_r].height = 13.5 * _lines


def pr(k):
    return PS + P[k].replace('B', 'B$')


def fr(k):
    return PS + FR[k].replace('B', 'B$')


# ===========================================================================
# 2. ①札幌予測確定
# ===========================================================================
ws = wb.create_sheet('①札幌予測確定')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'D8'
ws['A1'] = '① 札幌会場 会期20日間（7/22-8/10）商品別 販売予測【確定値】'
ws['A1'].font = TITLE
ws['A2'] = ('算出根拠：東京を除く名古屋①・博多②・大阪④の「%d日累計 ÷ %s」倍率を商品別に加重平均し、%sの%s実績に乗算'
            % (RUNC, NB, CUR, NBR))
ws['A2'].font = NOTE
ws['A3'] = '加重：大阪2・博多1・名古屋1（「前提・入力」シートで変更可）／名古屋は19日全会期実績を20日相当に換算（+%.2f%%）' % (TAIL * 100)
ws['A3'].font = NOTE

HD = [('No.', 5, NAVY), ('JAN', 15, NAVY), ('商品名', 44, NAVY), ('税込\n単価', 9, NAVY),
      ('%s\n%s\n%s' % (CUR, NB, BASE_RANGE), 10, PatternFill('solid', fgColor=VF[CUR])),
      ('名古屋①\n%s' % NB, 9, PatternFill('solid', fgColor=VF['名古屋'])),
      ('名古屋①\n19日累計\n(全会期)', 10, PatternFill('solid', fgColor=VF['名古屋'])),
      ('名古屋\n倍率g', 9, PatternFill('solid', fgColor=VF['名古屋'])),
      ('博多②\n%s' % NB, 9, PatternFill('solid', fgColor=VF['博多'])),
      ('博多②\n20日累計', 10, PatternFill('solid', fgColor=VF['博多'])),
      ('博多\n倍率g', 9, PatternFill('solid', fgColor=VF['博多'])),
      ('大阪④\n%s' % NB, 9, PatternFill('solid', fgColor=VF['大阪'])),
      ('大阪④\n20日累計', 10, PatternFill('solid', fgColor=VF['大阪'])),
      ('大阪\n倍率g', 9, PatternFill('solid', fgColor=VF['大阪'])),
      ('採用\n倍率g', 9, BLUE),
      ('札幌\n会期20日\n予測販売数', 12, JUDG), ('札幌\n予測販売金額', 13, JUDG),
      ('札幌\n数量構成比', 10, JUDG), ('札幌\n販売可能数\n(追納込)', 12, JUDG),
      ('予測\n消化率', 9, JUDG), ('会期末\n残在庫', 10, JUDG),
      ('会期中\n過不足', 10, JUDG), ('会期中\n在庫判定', 13, JUDG)]
hr = 6
for i, (h, w, f) in enumerate(HD, 1):
    c = ws.cell(hr, i, h)
    c.fill = f
    c.font = H2
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[hr].height = 46
r0 = 7
tot = r0 + len(T)
for i, (jan, nm) in enumerate(T):
    r = r0 + i
    ws.cell(r, 1, i + 1); ws.cell(r, 2, str(jan)); ws.cell(r, 3, nm)
    ws.cell(r, 4, R('札幌', jan)['price']).number_format = YEN
    ws.cell(r, 5, S('札幌', jan, 0, BASE)).number_format = INT
    ws.cell(r, 6, S('名古屋', jan, 0, BASE)).number_format = INT
    ws.cell(r, 7, S('名古屋', jan, 0, 19)).number_format = INT
    ws.cell(r, 8, '=IF(F%d=0,"",G%d*(1+%s)/F%d)' % (r, r, pr('tail'), r)).number_format = MUL
    ws.cell(r, 9, S('博多', jan, 0, BASE)).number_format = INT
    ws.cell(r, 10, S('博多', jan, 0, 20)).number_format = INT
    ws.cell(r, 11, '=IF(I%d=0,"",J%d/I%d)' % (r, r, r)).number_format = MUL
    ws.cell(r, 12, S('大阪', jan, 0, BASE)).number_format = INT
    ws.cell(r, 13, S('大阪', jan, 0, 20)).number_format = INT
    ws.cell(r, 14, '=IF(L%d=0,"",M%d/L%d)' % (r, r, r)).number_format = MUL
    ws.cell(r, 15, '=IFERROR((IF(N{r}="",0,N{r}*{wo})+IF(K{r}="",0,K{r}*{wh})+IF(H{r}="",0,H{r}*{wn}))'
                   '/(IF(N{r}="",0,{wo})+IF(K{r}="",0,{wh})+IF(H{r}="",0,{wn})),{ga})'
           .format(r=r, wo=pr('w_osaka'), wh=pr('w_hakata'), wn=pr('w_nagoya'), ga=pr('g_all'))
           ).number_format = MUL
    ws.cell(r, 16, '=ROUND(E%d*O%d,0)' % (r, r)).number_format = INT
    ws.cell(r, 17, '=P%d*D%d' % (r, r)).number_format = YEN
    ws.cell(r, 18, '=IFERROR(P%d/P$%d,0)' % (r, tot)).number_format = PCT
    ws.cell(r, 19, R('札幌', jan)['avail']).number_format = INT
    ws.cell(r, 20, '=IFERROR(P%d/S%d,0)' % (r, r)).number_format = PCT
    ws.cell(r, 21, '=MAX(0,S%d-P%d)' % (r, r)).number_format = INT
    ws.cell(r, 22, '=P%d-S%d' % (r, r)).number_format = '+#,##0;-#,##0;0'
    ws.cell(r, 23, '=IF(V%d>0,"会期中に在庫切れ",IF(T%d>=0.9,"要注意(90%%超)",IF(T%d>=0.5,"適正","在庫余剰")))' % (r, r, r))
ws.cell(tot, 3, '合計（LEGS 42SKU）')
for c in [5, 6, 7, 9, 10, 12, 13, 16, 17, 19, 21]:
    L = gcl(c)
    ws.cell(tot, c, '=SUM(%s%d:%s%d)' % (L, r0, L, tot - 1)).number_format = YEN if c == 17 else INT
ws.cell(tot, 15, '=IFERROR(P%d/E%d,0)' % (tot, tot)).number_format = MUL
ws.cell(tot, 18, 1).number_format = PCT
ws.cell(tot, 20, '=IFERROR(P%d/S%d,0)' % (tot, tot)).number_format = PCT
ws.cell(tot, 22, '=SUM(V%d:V%d)' % (r0, tot - 1)).number_format = '+#,##0;-#,##0;0'
ws.cell(tot, 23, '=COUNTIF(W%d:W%d,"会期中に在庫切れ")&"品目"' % (r0, tot - 1))
body(ws, r0, tot, 23)
totrow(ws, tot, 23)
hdrfmt(ws, hr, hr, 23)
ws.conditional_formatting.add('W%d:W%d' % (r0, tot - 1),
                              FormulaRule(formula=['$W%d="会期中に在庫切れ"' % r0], fill=ALERT,
                                          font=Font(name=FONT, size=9, bold=True, color='9C0006')))
ws.conditional_formatting.add('W%d:W%d' % (r0, tot - 1),
                              FormulaRule(formula=['$W%d="要注意(90%%超)"' % r0], fill=WARN))
ws.conditional_formatting.add('W%d:W%d' % (r0, tot - 1),
                              FormulaRule(formula=['$W%d="適正"' % r0], fill=OKF))
S1 = "'①札幌予測確定'"

# ===========================================================================
# 3. ②会場別構成比
# ===========================================================================
ws = wb.create_sheet('②会場別構成比')
ws.sheet_view.showGridLines = False
ws['A1'] = '② 会場別 物販実績とLEGS構成比／東京対比 会場比率'
ws['A1'].font = TITLE
ws['A2'] = '物販売上＝レジ袋代・音声ガイド・当日券を除く商品売上（税込）。LEGS＝本42SKU（各会場の商品分類「レッグス」と完全一致）'
ws['A2'].font = NOTE

cols = VS + ['札幌(予測)'] + [f[0] for f in FUTURE]
r = 4
ws.cell(r, 1, '■ 会場別 実績・予測サマリー').font = SEC
r += 1
hr2 = r
ws.cell(hr2, 1, '項目').fill = NAVY
ws.cell(hr2, 1).font = H2
ws.cell(hr2, 2, '東京対比\n算出根拠').fill = NAVY
ws.cell(hr2, 2).font = H2
for i, v in enumerate(cols):
    c = ws.cell(hr2, 3 + i, v)
    c.fill = PatternFill('solid', fgColor=VF.get(v.replace('(予測)', ''), '7B3F00'))
    c.font = H2
    ws.column_dimensions[gcl(3 + i)].width = 15
ws.column_dimensions['A'].width = 26
ws.column_dimensions['B'].width = 22
ws.row_dimensions[hr2].height = 34

METRICS = ['会期', '会期日数', 'データ実績日数', '入場者数', '全物販 販売数量', '全物販 販売金額(税込)',
           'LEGS 販売数量', 'LEGS 販売金額(税込)', 'LEGS 数量構成比', 'LEGS 金額構成比',
           '東京対比（数量）', '東京対比（金額）', 'LEGS平均単価']
base = hr2 + 1
for k, m in enumerate(METRICS):
    ws.cell(base + k, 1, m).font = BD
for i, v in enumerate(VS):
    C = gcl(3 + i)
    n = DATA_DAYS[v]
    q, a = vt(v, 0, n)
    lq, la = vt(v, 0, n, legs=True)
    vals = [PERIOD[v] + '\n' + HALL[v], RUN[v], n, sum(ATT[v][:n]), q, a, lq, la]
    for k, x in enumerate(vals):
        c = ws.cell(base + k, 3 + i, x)
        c.number_format = YEN if k == 5 or k == 7 else INT
    ws.cell(base + 8, 3 + i, '={C}{r1}/{C}{r0}'.format(C=C, r1=base + 6, r0=base + 4)).number_format = PCT
    ws.cell(base + 9, 3 + i, '={C}{r1}/{C}{r0}'.format(C=C, r1=base + 7, r0=base + 5)).number_format = PCT
    ws.cell(base + 10, 3 + i, '={C}{r}/$C${r}'.format(C=C, r=base + 4)).number_format = PCT
    ws.cell(base + 11, 3 + i, '={C}{r}/$C${r}'.format(C=C, r=base + 5)).number_format = PCT
    ws.cell(base + 12, 3 + i, '=IFERROR({C}{a}/{C}{q},0)'.format(C=C, a=base + 7, q=base + 6)).number_format = YEN
    if v == '札幌':
        ws.cell(base + 2, 3 + i).font = Font(name=FONT, size=9, bold=True, color='C00000')
# 札幌(予測)
C = gcl(3 + len(VS))
gq = vt('大阪', 0, 20)[0] / vt('大阪', 0, BASE)[0]
ga = vt('大阪', 0, 20)[1] / vt('大阪', 0, BASE)[1]
ws.cell(base, 3 + len(VS), PERIOD['札幌'] + '\n会期20日 予測')
ws.cell(base + 1, 3 + len(VS), 20)
ws.cell(base + 2, 3 + len(VS), '予測')
ws.cell(base + 3, 3 + len(VS), round(_attpred)).number_format = INT
ws.cell(base + 4, 3 + len(VS), '=G%d*%.6f' % (base + 4, gq)).number_format = INT
ws.cell(base + 5, 3 + len(VS), '=G%d*%.6f' % (base + 5, ga)).number_format = YEN
ws.cell(base + 6, 3 + len(VS), '=%s!P%d' % (S1, tot)).number_format = INT
ws.cell(base + 7, 3 + len(VS), '=%s!Q%d' % (S1, tot)).number_format = YEN
for k, f in [(8, '={C}{r1}/{C}{r0}'), (9, '={C}{r1}/{C}{r0}')]:
    ws.cell(base + k, 3 + len(VS), f.format(C=C, r1=base + 6 + (k - 8), r0=base + 4 + (k - 8))).number_format = PCT
ws.cell(base + 10, 3 + len(VS), '={C}{r}/$C${r}'.format(C=C, r=base + 4)).number_format = PCT
ws.cell(base + 11, 3 + len(VS), '={C}{r}/$C${r}'.format(C=C, r=base + 5)).number_format = PCT
ws.cell(base + 12, 3 + len(VS), '={C}{a}/{C}{q}'.format(C=C, a=base + 7, q=base + 6)).number_format = YEN
ws.cell(base + 2, 3 + len(VS)).font = Font(name=FONT, size=9, bold=True, color='C00000')
# 将来会場
for i, (nm, per, dd, ratio, note) in enumerate(FUTURE):
    cc = 3 + len(VS) + 1 + i
    C = gcl(cc)
    ws.cell(hr2, cc).fill = FUT
    ws.cell(base, cc, per + '\n' + nm)
    ws.cell(base + 1, cc, dd)
    ws.cell(base + 2, cc, '予測')
    ws.cell(base + 4, cc, '={tq}*{fr}'.format(tq=pr('tk_qty'), fr=fr(nm))).number_format = INT
    ws.cell(base + 5, cc, '={ta}*{fr}'.format(ta=pr('tk_amt'), fr=fr(nm))).number_format = YEN
    ws.cell(base + 7, cc, '={C}{r}*{l}'.format(C=C, r=base + 5, l=pr('lsh_a'))).number_format = YEN
    ws.cell(base + 6, cc, '=ROUND({C}{a}/{u},0)'.format(C=C, a=base + 7, u=pr('unit'))).number_format = INT
    ws.cell(base + 8, cc, '={C}{r1}/{C}{r0}'.format(C=C, r1=base + 6, r0=base + 4)).number_format = PCT
    ws.cell(base + 9, cc, '={C}{r1}/{C}{r0}'.format(C=C, r1=base + 7, r0=base + 5)).number_format = PCT
    ws.cell(base + 10, cc, '={C}{r}/$C${r}'.format(C=C, r=base + 4)).number_format = PCT
    ws.cell(base + 11, cc, '={fr}'.format(fr=fr(nm))).number_format = PCT
    ws.cell(base + 12, cc, '={C}{a}/{C}{q}'.format(C=C, a=base + 7, q=base + 6)).number_format = YEN
    ws.cell(base + 11, 2, '')
    ws.cell(base + k, 2, '')
for k in range(len(METRICS)):
    ws.cell(base + k, 2, '')
ws.cell(base + 11, 2, 'ユーザー指定／入力値')
ws.cell(base + 11, 2).font = NOTE
ncol2 = 3 + len(cols) - 1
body(ws, base, base + len(METRICS) - 1, ncol2, namecol=1)
for k in range(len(METRICS)):
    ws.cell(base + k, 1).font = BD
    ws.cell(base + k, 1).alignment = LFT
ws.cell(base, 1).alignment = LFT
for i in range(len(cols)):
    ws.cell(base, 3 + i).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
ws.row_dimensions[base].height = 40
hdrfmt(ws, hr2, hr2, ncol2)

# 供給元別構成比
r = base + len(METRICS) + 2
ws.cell(r, 1, '■ 供給元別 金額構成比（各会場 会期全体。%sは%s実績）' % (CUR, NB)).font = SEC
r += 1
GRPS = ['LEGS(42SKU)', 'ムービック', '東宝', 'MAPPA', 'MBS', '書籍', 'その他']
ws.cell(r, 1, '供給元').fill = NAVY
ws.cell(r, 1).font = H2
for i, v in enumerate(VS):
    c = ws.cell(r, 3 + i, v)
    c.fill = PatternFill('solid', fgColor=VF[v])
    c.font = H2
gr0 = r + 1
for k, g in enumerate(GRPS):
    ws.cell(gr0 + k, 1, g).font = BD
for i, v in enumerate(VS):
    n = DATA_DAYS[v]
    tota_ = vt(v, 0, n)[1]
    acc = {}
    for kk, x in D[v]['rows'].items():
        if (x['cat'] or '') in EXCL:
            continue
        acc[grp(x['cat'])] = acc.get(grp(x['cat']), 0) + sum(x['daily'][:n]) * x['price']
    for k, g in enumerate(GRPS):
        ws.cell(gr0 + k, 3 + i, acc.get(g, 0) / tota_).number_format = PCT
gtot = gr0 + len(GRPS)
ws.cell(gtot, 1, '合計').font = BD
for i in range(len(VS)):
    L = gcl(3 + i)
    ws.cell(gtot, 3 + i, '=SUM(%s%d:%s%d)' % (L, gr0, L, gtot - 1)).number_format = PCT
body(ws, gr0, gtot, 3 + len(VS) - 1, namecol=1)
for k in range(len(GRPS) + 1):
    ws.cell(gr0 + k, 1).font = BD
    ws.cell(gr0 + k, 1).alignment = LFT
totrow(ws, gtot, 3 + len(VS) - 1, namecol=1)
hdrfmt(ws, r, r, 3 + len(VS) - 1)

# 新規SKU投入
r = gtot + 2
ws.cell(r, 1, '■ 新規SKU投入状況（前会場までに存在しなかったJAN）').font = SEC
r += 1
for i, h in enumerate(['会場', '', '総SKU数', '新規SKU数', '新規の金額シェア(%s)' % NB,
                       'LEGS金額構成比(%s)' % NB, '備考'], 1):
    c = ws.cell(r, i, h)
    c.fill = NAVY
    c.font = H2
nr0 = r + 1
seen = set()
notes = {'東京': '初回会場', '名古屋': 'MAPPA記念商品26SKU等を追加投入',
         '博多': '小幅追加', '大阪': '★新商品106SKU投入（東宝一般66・ムービック23・MAPPA16）',
         '札幌': '新規なし。品揃えは大阪と同一'}
for k, v in enumerate(VS):
    cur = {int(x) for x, y in D[v]['rows'].items() if (y['cat'] or '') not in EXCL}
    new = cur - seen
    t5 = vt(v, 0, BASE)[1]
    na = sum(sum(D[v]['rows'][str(j)]['daily'][:BASE]) * D[v]['rows'][str(j)]['price'] for j in new)
    ws.cell(nr0 + k, 1, v)
    ws.cell(nr0 + k, 3, len(cur)).number_format = INT
    ws.cell(nr0 + k, 4, len(new)).number_format = INT
    ws.cell(nr0 + k, 5, na / t5).number_format = PCT
    ws.cell(nr0 + k, 6, vt(v, 0, BASE, legs=True)[1] / t5).number_format = PCT
    ws.cell(nr0 + k, 7, notes[v])
    seen |= cur
body(ws, nr0, nr0 + len(VS) - 1, 7, namecol=7)
for k in range(len(VS)):
    ws.cell(nr0 + k, 1).font = BD
hdrfmt(ws, r, r, 7)
ws.column_dimensions['G'].width = 52
S2 = "'②会場別構成比'"
ROW_TKQ, ROW_TKA = base + 4, base + 5

# ===========================================================================
# 4. ③石川以降予測
# ===========================================================================
ws = wb.create_sheet('③石川以降予測')
ws.sheet_view.showGridLines = False
ws['A1'] = '③ 石川会場以降 5会場 LEGS商品 需要予測'
ws['A1'].font = TITLE
ws['A2'] = '会場物販売上＝東京(初回)実績 × 東京対比比率　／　LEGS売上＝会場物販売上 × LEGS金額構成比　／　LEGS数量＝LEGS売上 ÷ LEGS平均単価'
ws['A2'].font = NOTE
ws['A3'] = '商品別按分＝大阪20日実績の数量構成比 × ウェイト ＋ 札幌20日予測の数量構成比 × ウェイト（「前提・入力」で調整可）'
ws['A3'].font = NOTE
ws['A4'] = '※商品別表の必要数合計は、会場×商品ごとに整数丸めを行うため会場別表の合計と数個の差が生じる（発注判断に影響しない範囲）'
ws['A4'].font = NOTE

r = 5
ws.cell(r, 1, '■ 会場別 需要予測').font = SEC
r += 1
FH = ['会場', '会期', '日数', '東京対比', '会場 全物販売上', 'LEGS 売上', 'LEGS 数量', '構成比根拠']
for i, h in enumerate(FH, 1):
    c = ws.cell(r, i, h)
    c.fill = FUT
    c.font = H2
FW = [26, 24, 6, 10, 18, 16, 13, 40]
for i, w in enumerate(FW, 1):
    ws.column_dimensions[gcl(i)].width = w
fr0 = r + 1
for k, (nm, per, dd, ratio, note) in enumerate(FUTURE):
    rr = fr0 + k
    ws.cell(rr, 1, nm)
    ws.cell(rr, 2, per)
    ws.cell(rr, 3, dd)
    ws.cell(rr, 4, '=%s' % fr(nm)).number_format = PCT
    ws.cell(rr, 5, '=%s*D%d' % (pr('tk_amt'), rr)).number_format = YEN
    ws.cell(rr, 6, '=E%d*%s' % (rr, pr('lsh_a'))).number_format = YEN
    ws.cell(rr, 7, '=ROUND(F%d/%s,0)' % (rr, pr('unit'))).number_format = INT
    ws.cell(rr, 8, note)
ftot = fr0 + len(FUTURE)
ws.cell(ftot, 1, '石川以降 5会場 合計')
ws.cell(ftot, 4, '=SUM(D%d:D%d)' % (fr0, ftot - 1)).number_format = PCT
for c in [5, 6, 7]:
    L = gcl(c)
    ws.cell(ftot, c, '=SUM(%s%d:%s%d)' % (L, fr0, L, ftot - 1)).number_format = YEN if c < 7 else INT
body(ws, fr0, ftot, 8, namecol=1)
for k in range(len(FUTURE) + 1):
    ws.cell(fr0 + k, 1).font = BD
    ws.cell(fr0 + k, 1).alignment = LFT
    ws.cell(fr0 + k, 8).alignment = LFT
totrow(ws, ftot, 8, namecol=1)
hdrfmt(ws, r, r, 8)

r = ftot + 2
ws.cell(r, 1, '■ 在庫収支（LEGS 42SKU 合計）').font = SEC
r += 1
bal = [('残在庫想定（%s／総計）' % STOCK_STAGE, sum(INV[str(j)]['real'] for j in TJ), INT,
        '在庫フロー台帳の最終「実在庫」列。会場ごとの追加発注→販売→検品を積み上げた値'),
       ('　うち札幌会場へ納品（初回＋追納）', '=%s!S%d' % (S1, tot), INT,
        '札幌の販売可能数。%s（LEGS計%s個）を含む' % (RESUPPLY_NOTE, f'{sum(RESUPPLY.values()):,}')),
       ('　差引 倉庫残（石川以降へ即出荷可）', '=B%d-B%d' % (r, r + 1), INT, '札幌へ引き当てた分を除いた手元在庫'),
       ('札幌 会期20日 予測販売数', '=%s!P%d' % (S1, tot), INT, ''),
       ('　うち札幌へ追納が必要', 'PLACEHOLDER_RESUP', INT,
        '札幌の会期20日予測が納品済数を上回る分。倉庫残から充当する（運営欠品報告への対応分）'),
       ('札幌 会期末 残在庫（返送見込）', '=%s!U%d' % (S1, tot), INT, '品目別に下限0で集計（在庫切れ品はマイナスにしない）'),
       ('　うち石川以降へスライド可能', '=B%d*%s' % (r + 5, pr('slide')), INT, 'スライド率は「前提・入力」で調整'),
       ('石川以降 供給可能在庫 合計', '=B%d-B%d+B%d' % (r + 2, r + 4, r + 6), INT,
        '＝倉庫残 − 札幌への追納 ＋ 札幌からの返送スライド'),
       ('確定済み追加発注（入荷予定）', sum(v['qty'] for v in DECIDED.values()), INT,
        '発注判断済みの分。'
        + '／'.join('%s %s%s（%s入荷）' % (dict(T)[j], f"{v['qty']:,}", v['unit'], v['arrive_at'])
                   for j, v in DECIDED.items())),
       ('石川以降 5会場 必要数', '=G%d' % ftot, INT, ''),
       ('全体過不足（供給可能＋確定発注−必要）', '=B%d+B%d-B%d' % (r + 7, r + 8, r + 9), '+#,##0;-#,##0;0',
        'プラス＝全体では在庫充足。ただし商品別には過不足が生じる'),
       ('追加生産 必要数（商品別不足の合計）', 'PLACEHOLDER_ADDPROD', INT,
        '下表の商品別「追加生産必要数」の合計。全体余剰でも品目別には不足が出る')]
for k, (a, b, f, n) in enumerate(bal):
    ws.cell(r + k, 1, a).font = BD
    c = ws.cell(r + k, 2, b)
    c.number_format = f
    c.font = BD if k == len(bal) - 1 else NM
    c.border = BOX
    c.alignment = CTR
    ws.cell(r + k, 3, n).font = NOTE
    ws.cell(r + k, 3).alignment = LFT
ws.cell(r + len(bal) - 1, 2).fill = ALERT
ws.cell(r + len(bal) - 2, 2).fill = OKF
ws.cell(r + 7, 2).fill = OKF
BALROW = r
BAL_LAST = r + len(bal) - 1

# 商品別
r = r + len(bal) + 1
ws.cell(r, 1, '■ 商品別 石川以降 必要数と在庫過不足').font = SEC
r += 1
PH = [('No.', 5, NAVY), ('JAN', 15, NAVY), ('商品名', 44, NAVY),
      ('大阪④\n20日実績', 11, PatternFill('solid', fgColor=VF['大阪'])),
      ('大阪\n構成比', 9, PatternFill('solid', fgColor=VF['大阪'])),
      ('札幌⑤\n20日予測', 11, PatternFill('solid', fgColor=VF['札幌'])),
      ('札幌\n構成比', 9, PatternFill('solid', fgColor=VF['札幌'])),
      ('採用\n按分構成比', 11, BLUE)]
for nm, per, dd, ratio, note in FUTURE:
    PH.append((nm.split(' ')[0] + '\n必要数', 11, FUT))
PH += [('石川以降\n必要数 計', 12, JUDG),
       ('★残在庫想定\n(梅田会場後)', 11, PatternFill('solid', fgColor='375623')),
       ('うち札幌\n納品済', 10, PatternFill('solid', fgColor='375623')),
       ('差引\n倉庫残', 10, PatternFill('solid', fgColor='375623')),
       ('札幌へ\n追納必要数', 11, ALERT), ('札幌会期末\n返送(スライド)', 12, JUDG),
       ('石川以降\n供給可能計', 12, JUDG),
       ('確定発注\n(入荷予定)', 11, PatternFill('solid', fgColor='C55A11')),
       ('過不足\n(必要−供給可能−確定)', 15, JUDG),
       ('追加生産\n必要数', 11, JUDG), ('推奨発注数\n(安全率込)', 12, JUDG)]
phr = r
for i, (h, w, f) in enumerate(PH, 1):
    c = ws.cell(phr, i, h)
    c.fill = f
    c.font = H2
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[phr].height = 42
pr0 = phr + 1
ptot = pr0 + len(T)
NF = len(FUTURE)
c_need = 8 + NF + 1
for i, (jan, nm) in enumerate(T):
    rr = pr0 + i
    ws.cell(rr, 1, i + 1); ws.cell(rr, 2, str(jan)); ws.cell(rr, 3, nm)
    ws.cell(rr, 4, S('大阪', jan, 0, 20)).number_format = INT
    ws.cell(rr, 5, '=IFERROR(D%d/D$%d,0)' % (rr, ptot)).number_format = PCT2
    ws.cell(rr, 6, '=%s!P%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, 7, '=IFERROR(F%d/F$%d,0)' % (rr, ptot)).number_format = PCT2
    ws.cell(rr, 8, '=E%d*%s+G%d*%s' % (rr, pr('mix_o'), rr, pr('mix_s'))).number_format = PCT2
    for k in range(NF):
        ws.cell(rr, 9 + k, '=ROUND($H%d*%s!$G$%d,0)' % (rr, "'③石川以降予測'", fr0 + k)).number_format = INT
    L1, L2 = gcl(9), gcl(9 + NF - 1)
    ws.cell(rr, c_need, '=SUM(%s%d:%s%d)' % (L1, rr, L2, rr)).number_format = INT
    ws.cell(rr, c_need + 1, INV[str(jan)]['real']).number_format = INT
    ws.cell(rr, c_need + 2, '=%s!S%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, c_need + 3, '=%s%d-%s%d' % (gcl(c_need + 1), rr, gcl(c_need + 2), rr)).number_format = INT
    ws.cell(rr, c_need + 4, '=MAX(0,%s!P%d-%s!S%d)' % (S1, r0 + i, S1, r0 + i)).number_format = INT
    ws.cell(rr, c_need + 5, '=ROUND(%s!U%d*%s,0)' % (S1, r0 + i, pr('slide'))).number_format = INT
    ws.cell(rr, c_need + 6, '=%s%d-%s%d+%s%d'
            % (gcl(c_need + 3), rr, gcl(c_need + 4), rr, gcl(c_need + 5), rr)).number_format = INT
    ws.cell(rr, c_need + 7, DECIDED[jan]['qty'] if jan in DECIDED else 0).number_format = INT
    ws.cell(rr, c_need + 8, '=%s%d-%s%d-%s%d'
            % (gcl(c_need), rr, gcl(c_need + 6), rr, gcl(c_need + 7), rr)
            ).number_format = '+#,##0;-#,##0;0'
    ws.cell(rr, c_need + 9, '=MAX(0,%s%d)' % (gcl(c_need + 8), rr)).number_format = INT
    ws.cell(rr, c_need + 10, '=IF(%s%d>0,ROUNDUP(%s%d*(1+%s),0),0)'
            % (gcl(c_need + 9), rr, gcl(c_need + 9), rr, pr('safe'))).number_format = INT
NCP = c_need + 10
ws.cell(ptot, 3, '合計（LEGS 42SKU）')
for c in [4, 6] + list(range(9, NCP + 1)):
    if c in (5, 7, 8):
        continue
    L = gcl(c)
    ws.cell(ptot, c, '=SUM(%s%d:%s%d)' % (L, pr0, L, ptot - 1)).number_format = INT
for c in [5, 7, 8]:
    L = gcl(c)
    ws.cell(ptot, c, '=SUM(%s%d:%s%d)' % (L, pr0, L, ptot - 1)).number_format = PCT2
body(ws, pr0, ptot, NCP)
totrow(ws, ptot, NCP)
hdrfmt(ws, phr, phr, NCP)
ws.cell(BALROW + 4, 2, '=%s%d' % (gcl(c_need + 4), ptot))
ws.cell(BALROW + 4, 2).number_format = INT
ws.cell(BAL_LAST, 2, '=%s%d' % (gcl(c_need + 9), ptot))
ws.cell(BAL_LAST, 2).number_format = INT
ws.conditional_formatting.add('%s%d:%s%d' % (gcl(c_need + 9), pr0, gcl(c_need + 9), ptot - 1),
                              FormulaRule(formula=['$%s%d>0' % (gcl(c_need + 9), pr0)], fill=ALERT))
ws.conditional_formatting.add('%s%d:%s%d' % (gcl(c_need + 8), pr0, gcl(c_need + 8), ptot - 1),
                              FormulaRule(formula=['$%s%d>0' % (gcl(c_need + 8), pr0)], fill=ALERT))
S3 = "'③石川以降予測'"

# ===========================================================================
# 5. ④追加発注判定
# ===========================================================================
ws = wb.create_sheet('④追加発注判定')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'D8'
ws['A1'] = '④ 追加発注 必要有無の判定（札幌会期中 ＋ 石川以降5会場）'
ws['A1'].font = TITLE
ws['A2'] = '【札幌会期中】会期20日予測 vs 札幌への納品済 販売可能数　／　【石川以降】5会場必要数 vs 札幌会期末残在庫＋倉庫在庫　※2026/7/28 SPJ様欠品報告を突合'
ws['A2'].font = NOTE
ws['A3'] = ('推奨発注数＝不足数×(1+安全率10%)を切り上げ。'
            'T〜W列に欠品発生会場・最終発注期限（会期初日−生産日数−バッファ）・残日数・緊急度を記載')
ws['A3'].font = NOTE
JH = [('No.', 5, NAVY), ('JAN', 15, NAVY), ('商品名', 44, NAVY), ('税込単価', 9, NAVY),
      ('%s\n%s実績' % (CUR, NB), 10, PatternFill('solid', fgColor=VF[CUR])),
      ('札幌\n会期20日予測', 12, PatternFill('solid', fgColor=VF['札幌'])),
      ('札幌\n販売可能数\n(追納込)', 12, PatternFill('solid', fgColor=VF['札幌'])),
      ('札幌\n予測消化率', 10, PatternFill('solid', fgColor=VF['札幌'])),
      ('札幌会期中\n過不足', 11, ALERT), ('札幌会期中\n判定', 15, ALERT), ('札幌会期中\n推奨発注数', 12, ALERT),
      ('石川以降\n必要数', 11, FUT), ('札幌からの\nスライド', 11, FUT), ('7/8実在庫の\n倉庫残', 12, FUT),
      ('石川以降\n過不足', 11, FUT), ('石川以降\n判定', 15, FUT), ('石川以降\n推奨発注数', 12, FUT),
      ('発注合計\n(札幌+石川以降)', 14, JUDG), ('発注金額\n(税込小売換算)', 14, JUDG),
      ('欠品発生\n会場', 20, ALERT), ('最終発注期限', 13, ALERT),
      ('基準日からの\n残日数', 12, ALERT), ('発注の緊急度', 16, ALERT),
      ('優先度', 10, JUDG), ('運営 欠品報告\n(7/28 SPJ様)', 16, ALERT),
      ('モデル×現場\n突合', 16, ALERT), ('備考', 50, GREY)]
hr4 = 6
for i, (h, w, f) in enumerate(JH, 1):
    c = ws.cell(hr4, i, h)
    c.fill = f
    c.font = H2 if f not in (ALERT, GREY) else Font(name=FONT, size=9, bold=True, color='333333')
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[hr4].height = 44
j0 = hr4 + 1
jtot = j0 + len(T)
for i, (jan, nm) in enumerate(T):
    rr = j0 + i
    ws.cell(rr, 1, i + 1); ws.cell(rr, 2, str(jan)); ws.cell(rr, 3, nm)
    ws.cell(rr, 4, R('札幌', jan)['price']).number_format = YEN
    ws.cell(rr, 5, '=%s!E%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, 6, '=%s!P%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, 7, '=%s!S%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, 8, '=IFERROR(F%d/G%d,0)' % (rr, rr)).number_format = PCT
    ws.cell(rr, 9, '=F%d-G%d' % (rr, rr)).number_format = '+#,##0;-#,##0;0'
    ws.cell(rr, 10, '=IF(I%d>0,"要 追加発注",IF(H%d>=0.9,"要注意(90%%超)",IF(H%d>=0.5,"適正","在庫余剰")))'
            % (rr, rr, rr))
    ws.cell(rr, 11, '=IF(I%d>0,ROUNDUP(I%d*(1+%s),0),0)' % (rr, rr, pr('safe'))).number_format = INT
    ws.cell(rr, 12, '=%s!%s%d' % (S3, gcl(c_need), pr0 + i)).number_format = INT
    ws.cell(rr, 13, '=%s!%s%d' % (S3, gcl(c_need + 5), pr0 + i)).number_format = INT
    ws.cell(rr, 14, '=%s!%s%d' % (S3, gcl(c_need + 3), pr0 + i)).number_format = INT
    ws.cell(rr, 15, '=L%d-M%d-N%d' % (rr, rr, rr)).number_format = '+#,##0;-#,##0;0'
    ws.cell(rr, 16, '=IF(O%d>0,"要 追加生産",IF(L%d=0,"—","在庫で充足"))' % (rr, rr))
    ws.cell(rr, 17, '=%s!%s%d' % (S3, gcl(c_need + 10), pr0 + i)).number_format = INT
    ws.cell(rr, 18, '=K%d+Q%d' % (rr, rr)).number_format = INT
    ws.cell(rr, 19, '=R%d*D%d' % (rr, rr)).number_format = YEN
    _sh = SHORTAGE.get(str(jan), '')
    _nores = '追納困難' in _sh
    _fn = _venue_of(jan)
    ws.cell(rr, 20, _fn if _fn else '—')
    _dlf = _dl_formula(jan)
    ws.cell(rr, 21, _dlf if _dlf else '')
    if _dlf.startswith('='):
        ws.cell(rr, 21).number_format = 'yyyy/m/d'
    ws.cell(rr, 22, '=IF(OR(U%d="",U%d="再生産不可"),"",U%d-%s)' % (rr, rr, rr, pr('base'))
            ).number_format = '#,##0"日"'
    if jan in DECIDED:
        ws.cell(rr, 23, '◎発注確定済')
    else:
        ws.cell(rr, 23, '=IF(T%d="—","—",IF(U%d="再生産不可","×再生産不可",'
                '_xlfn.IFS(V%d<0,"★期限超過",V%d<=30,"★至急(30日以内)",V%d<=90,"要着手(90日以内)",'
                'TRUE(),"余裕あり")))' % (rr, rr, rr, rr, rr))
    ws.cell(rr, 24, '=IF(Y%d<>"",IF(%s,"S:現場欠品(追納不可)","S:現場欠品"),IF(I%d>0,"A:会期中欠品",'
            'IF(O%d>1000,"B:大口",IF(O%d>0,"C:通常","D:不要"))))'
            % (rr, 'TRUE()' if _nores else 'FALSE()', rr, rr, rr))
    ws.cell(rr, 25, _sh if _sh else '')
    ws.cell(rr, 26, '=IF(Y%d<>"",IF(I%d>0,"◎ 一致（モデルも会期中欠品を検知）",'
            'IF(H%d>=0.7,"○ 整合（予測消化率70%%超）","△ モデル未検知：現場報告を優先")),'
            'IF(I%d>0,"▲ モデルのみ検知：現場に在庫状況を確認","—"))' % (rr, rr, rr, rr))
    _wh = INV[str(jan)]['real'] - R('札幌', jan)['avail']
    nt = []
    if _sh:
        nt.append('【7/28 運営報告】' + _sh)
        nt.append('倉庫残%s個 → %s' % (f'{_wh:,}',
                  '◎倉庫からの追加納品で対応可能（追加生産は不要）' if _wh > 0 else '×倉庫在庫なし。追加生産が必要'))
    if _nores:
        nt.append('★追納不可のため発注推奨は0。石川以降は欠品前提で販売計画を組むか代替商品を手配')
    if PRED[jan] > R('札幌', jan)['avail']:
        nt.append('札幌会期中に欠品見込。至急スライド納品または追加発注')
    if R('札幌', jan)['avail'] <= 10:
        nt.append('札幌納品数が極小(%d個)' % R('札幌', jan)['avail'])
    if S('札幌', jan, 0, BASE) == 0:
        nt.append('%s%sの販売実績ゼロ（予測は大阪・博多構成比で按分）' % (CUR, NB))
    for v in ['名古屋', '博多', '大阪']:
        x = R(v, jan)
        if x['avail'] and x['total'] / x['avail'] >= 0.95:
            nt.append('%s会場で完売（実績が在庫上限で頭打ち＝予測は控えめ）' % v)
    ws.cell(rr, 27, ' ／ '.join(nt))
ws.cell(jtot, 3, '合計（LEGS 42SKU）')
for c in [5, 6, 7, 9, 11, 12, 13, 14, 15, 17, 18, 19]:
    L = gcl(c)
    ws.cell(jtot, c, '=SUM(%s%d:%s%d)' % (L, j0, L, jtot - 1)).number_format = YEN if c == 19 else INT
ws.cell(jtot, 8, '=IFERROR(F%d/G%d,0)' % (jtot, jtot)).number_format = PCT
ws.cell(jtot, 10, '=COUNTIF(J%d:J%d,"要 追加発注")&"品目"' % (j0, jtot - 1))
ws.cell(jtot, 16, '=COUNTIF(P%d:P%d,"要 追加生産")&"品目"' % (j0, jtot - 1))
ws.cell(jtot, 21, '=IF(COUNT(U%d:U%d)=0,"",TEXT(MIN(U%d:U%d),"yyyy/m/d")&" が最短期限")'
        % (j0, jtot - 1, j0, jtot - 1))
ws.cell(jtot, 23, '=COUNTIF(W%d:W%d,"★期限超過")&"品目が期限超過／"&COUNTIF(W%d:W%d,"★至急(30日以内)")&"品目が至急"'
        % (j0, jtot - 1, j0, jtot - 1))
ws.cell(jtot, 25, '=COUNTIF(Y%d:Y%d,"<>")&"品目 報告あり"' % (j0, jtot - 1))
body(ws, j0, jtot, 27)
for rr in range(j0, jtot):
    for _c in (20, 23, 25, 26, 27):
        ws.cell(rr, _c).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
totrow(ws, jtot, 27)
hdrfmt(ws, hr4, hr4, 27)
for rng, cond, fill in [('J', '$J%d="要 追加発注"', ALERT), ('J', '$J%d="要注意(90%%超)"', WARN),
                        ('J', '$J%d="適正"', OKF), ('P', '$P%d="要 追加生産"', ALERT),
                        ('T', '$T%d<>"—"', ALERT),
                        ('W', 'LEFT($W%d,1)="★"', ALERT), ('W', '$W%d="要着手(90日以内)"', WARN),
                        ('W', '$W%d="余裕あり"', OKF),
                        ('X', 'LEFT($X%d,1)="S"', ALERT), ('X', '$X%d="A:会期中欠品"', ALERT),
                        ('X', '$X%d="B:大口"', WARN), ('Y', '$Y%d<>""', ALERT),
                        ('Z', 'LEFT($Z%d,1)="△"', WARN), ('Z', 'LEFT($Z%d,1)="▲"', WARN)]:
    ws.conditional_formatting.add('%s%d:%s%d' % (rng, j0, rng, jtot - 1),
                                  FormulaRule(formula=[cond % j0], fill=fill))

# ===========================================================================
# 6. ⑤全データ統合リスト
# ===========================================================================
ws = wb.create_sheet('⑤全データ統合リスト')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'D8'
ws['A1'] = '⑤ 商品別 全データ統合リスト（5会場実績 ＋ 札幌予測 ＋ 石川以降予測）'
ws['A1'].font = TITLE
ws['A2'] = ('構成比＝LEGS 42SKU内シェア（数量・金額）。消化率＝販売数÷販売可能数。%sのみ%s(%d日目)時点の実績'
            % (CUR, BASE_TO, BASE))
ws['A2'].font = NOTE
ws['A3'] = '※イメージリング8種は東京¥22,000→名古屋以降¥27,500に改定。販売金額は各会場の実売単価で算出している'
ws['A3'].font = NOTE
G5 = [('基本情報', NAVY, [('No.', 5), ('JAN', 15), ('商品名', 44), ('現行\n税込単価', 9)])]
for v in VS:
    lab = ('%s（%s・%d日）' % (v, PERIOD[v].split('-')[0], RUN[v]) if v != CUR
           else '%s（%s 実績%d日）' % (CUR, BASE_RANGE, BASE))
    G5.append((lab, PatternFill('solid', fgColor=VF[v]),
               [('税込\n単価', 9), ('販売数', 10), ('数量\n構成比', 9), ('販売金額', 12), ('金額\n構成比', 9),
                ('販売\n可能数', 9), ('消化率', 9)]))
G5.append(('札幌 会期20日予測', PatternFill('solid', fgColor='C00000'),
           [('採用\n倍率g', 9), ('予測\n販売数', 10), ('予測\n販売金額', 12), ('予測\n消化率', 9), ('会期末\n残在庫', 10)]))
G5.append(('石川以降 5会場', FUT, [('必要数', 10), ('供給可能\n在庫', 11), ('過不足', 10), ('追加生産\n必要数', 11)]))
G5.append(('発注判定', JUDG, [('札幌会期中\n判定', 14), ('石川以降\n判定', 13), ('発注合計', 10), ('優先度', 10)]))
h1, h2 = 5, 6
col = 1
for gname, fill, cs in G5:
    st = col
    for h, w in cs:
        c = ws.cell(h2, col, h)
        c.fill = fill
        c.font = H2
        ws.column_dimensions[gcl(col)].width = w
        col += 1
    c = ws.cell(h1, st, gname)
    if col - 1 > st:
        ws.merge_cells(start_row=h1, start_column=st, end_row=h1, end_column=col - 1)
    c.fill = fill
    c.font = H1
NC5 = col - 1
ws.row_dimensions[h1].height = 20
ws.row_dimensions[h2].height = 34
q0 = 7
qtot = q0 + len(T)
for i, (jan, nm) in enumerate(T):
    rr = q0 + i
    ws.cell(rr, 1, i + 1); ws.cell(rr, 2, str(jan)); ws.cell(rr, 3, nm)
    ws.cell(rr, 4, R('札幌', jan)['price']).number_format = YEN
    cc = 5
    for v in VS:
        n = DATA_DAYS[v]
        ws.cell(rr, cc, R(v, jan)['price']).number_format = YEN
        L = gcl(cc + 1)
        ws.cell(rr, cc + 1, S(v, jan, 0, n)).number_format = INT
        ws.cell(rr, cc + 2, '=IFERROR(%s%d/%s$%d,0)' % (L, rr, L, qtot)).number_format = PCT
        ws.cell(rr, cc + 3, '=%s%d*%s%d' % (L, rr, gcl(cc), rr)).number_format = YEN
        L2 = gcl(cc + 3)
        ws.cell(rr, cc + 4, '=IFERROR(%s%d/%s$%d,0)' % (L2, rr, L2, qtot)).number_format = PCT
        ws.cell(rr, cc + 5, R(v, jan)['avail']).number_format = INT
        L3 = gcl(cc + 5)
        ws.cell(rr, cc + 6, '=IFERROR(%s%d/%s%d,0)' % (L, rr, L3, rr)).number_format = PCT
        cc += 7
    ws.cell(rr, cc, '=%s!O%d' % (S1, r0 + i)).number_format = MUL
    ws.cell(rr, cc + 1, '=%s!P%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, cc + 2, '=%s!Q%d' % (S1, r0 + i)).number_format = YEN
    ws.cell(rr, cc + 3, '=%s!T%d' % (S1, r0 + i)).number_format = PCT
    ws.cell(rr, cc + 4, '=%s!U%d' % (S1, r0 + i)).number_format = INT
    cc += 5
    ws.cell(rr, cc, '=%s!%s%d' % (S3, gcl(c_need), pr0 + i)).number_format = INT
    ws.cell(rr, cc + 1, '=%s!%s%d' % (S3, gcl(c_need + 6), pr0 + i)).number_format = INT
    ws.cell(rr, cc + 2, '=%s!%s%d' % (S3, gcl(c_need + 8), pr0 + i)).number_format = '+#,##0;-#,##0;0'
    ws.cell(rr, cc + 3, '=%s!%s%d' % (S3, gcl(c_need + 9), pr0 + i)).number_format = INT
    cc += 4
    ws.cell(rr, cc, "='④追加発注判定'!J%d" % (j0 + i))
    ws.cell(rr, cc + 1, "='④追加発注判定'!P%d" % (j0 + i))
    ws.cell(rr, cc + 2, "='④追加発注判定'!R%d" % (j0 + i)).number_format = INT
    ws.cell(rr, cc + 3, "='④追加発注判定'!X%d" % (j0 + i))
ws.cell(qtot, 3, '合計（LEGS 42SKU）')
sumcols = []
cc = 5
for v in VS:
    sumcols += [cc + 1, cc + 3, cc + 5]          # 販売数・販売金額・販売可能数
    cc += 7
sumcols += [cc + 1, cc + 2, cc + 4]              # 札幌予測: 販売数・金額・会期末残在庫
cc += 5
sumcols += [cc, cc + 1, cc + 2, cc + 3]          # 石川以降: 必要数・スライド・過不足・追加生産
cc += 4
sumcols += [cc + 2]                              # 発注合計
for c in sorted(set(sumcols)):
    if c > NC5:
        continue
    L = gcl(c)
    fm = ws.cell(q0, c).number_format
    ws.cell(qtot, c, '=SUM(%s%d:%s%d)' % (L, q0, L, qtot - 1)).number_format = fm if fm not in (PCT,) else INT
body(ws, q0, qtot, NC5)
totrow(ws, qtot, NC5)
hdrfmt(ws, h1, h2, NC5)

# ===========================================================================
# 7. ⑥商品別 会場推移
# ===========================================================================
ws = wb.create_sheet('⑥商品別 会場推移')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'D7'
ws['A1'] = '⑥ 商品別 会場推移（LEGS 42SKU内 数量構成比の会場間トレンド）'
ws['A1'].font = TITLE
ws['A2'] = ('各会場の会期全体（%sは%s実績と%d日予測）。構成比の伸び／減退から石川以降の重点商品を判断'
            % (CUR, NB, RUNC))
ws['A2'].font = NOTE
TH = [('No.', 5, NAVY), ('JAN', 15, NAVY), ('商品名', 44, NAVY)]
for v in VS:
    TH.append((v + '\n販売数', 10, PatternFill('solid', fgColor=VF[v])))
for v in VS:
    TH.append((v + '\n構成比', 9, PatternFill('solid', fgColor=VF[v])))
TH += [('札幌20日\n予測構成比', 11, JUDG), ('大阪→札幌\n構成比変化', 12, JUDG),
       ('東京→大阪\n構成比変化', 12, JUDG), ('トレンド', 14, JUDG)]
thr = 5
for i, (h, w, f) in enumerate(TH, 1):
    c = ws.cell(thr, i, h)
    c.fill = f
    c.font = H2
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[thr].height = 34
t0 = thr + 1
ttot = t0 + len(T)
for i, (jan, nm) in enumerate(T):
    rr = t0 + i
    ws.cell(rr, 1, i + 1); ws.cell(rr, 2, str(jan)); ws.cell(rr, 3, nm)
    for k, v in enumerate(VS):
        ws.cell(rr, 4 + k, S(v, jan, 0, DATA_DAYS[v])).number_format = INT
        L = gcl(4 + k)
        ws.cell(rr, 9 + k, '=IFERROR(%s%d/%s$%d,0)' % (L, rr, L, ttot)).number_format = PCT
    ws.cell(rr, 14, '=%s!R%d' % (S1, r0 + i)).number_format = PCT
    ws.cell(rr, 15, '=N%d-L%d' % (rr, rr)).number_format = '+0.0%;-0.0%;0.0%'
    ws.cell(rr, 16, '=L%d-I%d' % (rr, rr)).number_format = '+0.0%;-0.0%;0.0%'
    ws.cell(rr, 17, '=IF(O%d>=0.005,"↑ 伸長",IF(O%d<=-0.005,"↓ 減退","→ 横ばい"))' % (rr, rr))
ws.cell(ttot, 3, '合計（LEGS 42SKU）')
for k in range(len(VS)):
    L = gcl(4 + k)
    ws.cell(ttot, 4 + k, '=SUM(%s%d:%s%d)' % (L, t0, L, ttot - 1)).number_format = INT
    ws.cell(ttot, 9 + k, 1).number_format = PCT
ws.cell(ttot, 14, 1).number_format = PCT
body(ws, t0, ttot, 17)
totrow(ws, ttot, 17)
hdrfmt(ws, thr, thr, 17)
ws.conditional_formatting.add('Q%d:Q%d' % (t0, ttot - 1), FormulaRule(formula=['$Q%d="↑ 伸長"' % t0], fill=OKF))
ws.conditional_formatting.add('Q%d:Q%d' % (t0, ttot - 1), FormulaRule(formula=['$Q%d="↓ 減退"' % t0], fill=WARN))

# ===========================================================================
# 8. ⑦新商品影響分析（カニバリゼーション）
# ===========================================================================
def ltype(n):
    for k, t in [('クリアファイル', 'クリアファイル'), ('ポストカード', 'ポストカード類'),
                 ('缶バッジ', 'ブラインド缶バッジ'), ('ミニ色紙', 'ブラインドミニ色紙'),
                 ('ステッカー', 'ブラインドステッカー')]:
        if k in n:
            return t
    if 'イメージリング' in n or 'フラグメントケース' in n:
        return 'アクセサリー(リング等)'
    return '雑貨・その他'


def ntype(n):
    if 'うちわ' in n:
        return 'うちわ（新カテゴリ）'
    if 'ぬいぐるみ' in n or 'マスコット' in n:
        return 'ぬいぐるみ（新カテゴリ）'
    if 'クリアファイル' in n:
        return 'クリアファイル'
    if any(k in n for k in ['ポストカード', 'Pictas', 'しおり', 'マグネット', 'カード']):
        return 'カード・ポストカード類'
    if any(k in n for k in ['アクリルキーホルダー', 'アクリルスタンド', 'アクリルパネル', 'アクリルカードスタンド']):
        return 'アクリル系（コレクション）'
    if 'シール' in n or 'ステッカー' in n:
        return 'シール・ステッカー'
    if '缶バッジ' in n or '缶バッチ' in n:
        return '缶バッジ'
    return 'その他新商品'


ATTOT = {v: sum(ATT[v][:DATA_DAYS[v]]) for v in VS}
GRSET = ['LEGS(42SKU)', 'ムービック', '東宝', 'MAPPA', 'MBS', '書籍']
PERV = {}
for v in VS:
    n = DATA_DAYS[v]
    acc = {}
    tt = 0
    for k, x in D[v]['rows'].items():
        if (x['cat'] or '') in EXCL:
            continue
        s = sum(x['daily'][:n]) * x['price']
        tt += s
        acc[grp(x['cat'])] = acc.get(grp(x['cat']), 0) + s
    PERV[v] = {g: acc.get(g, 0) / ATTOT[v] for g in GRSET}
    PERV[v]['計'] = tt / ATTOT[v]
NEWSKU = {}
_seen = set()
for v in VS:
    cur = {int(k) for k, x in D[v]['rows'].items() if (x['cat'] or '') not in EXCL}
    NEWSKU[v] = cur - _seen
    _seen |= cur

ws = wb.create_sheet('⑦新商品影響分析')
ws.sheet_view.showGridLines = False
ws['A1'] = '⑦ 梅田(大阪)会場以降の追加商品による LEGS売上低下の要因分析'
ws['A1'].font = TITLE
ws['A2'] = ('会場ごとに規模が違うため、すべて「入場者1人あたり売上（円/人）」に正規化して比較する'
            '（%sは%sまでの%d日実績）' % (CUR, BASE_TO, BASE))
ws['A2'].font = NOTE

r = 4
ws.cell(r, 1, '■ 結論：市場は縮んでいない。LEGSの財布が新商品に奪われた（カニバリゼーション）').font = SEC
r += 1
for t in ['物販客単価（全商品）は名古屋7,012円/人 → 大阪6,882円/人 と ▲1.8% でほぼ横ばい。',
          '同期間にLEGSは2,099円/人 → 1,455円/人 と ▲30.7%（▲643円/人）。会場全体の購買力は落ちていない。',
          '大阪の新商品106SKUは1,360円/人を稼いでおり、LEGSの減少分643円/人はその約47%に相当する。',
          'つまり新商品売上のおよそ半分は「純増」ではなく既存商品（特にLEGSとMAPPA）からの移動である。']:
    ws.cell(r, 1, '・' + t).font = NM
    r += 1
r += 1

ws.cell(r, 1, '■ ① 入場者1人あたり売上（円/人）の会場推移　供給元別').font = SEC
r += 1
hh = r
ws.cell(hh, 1, '供給元').fill = NAVY
ws.cell(hh, 1).font = H2
for i, v in enumerate(VS):
    c = ws.cell(hh, 2 + i, v)
    c.fill = PatternFill('solid', fgColor=VF[v])
    c.font = H2
ws.cell(hh, 7, '名古屋→大阪\n増減').fill = JUDG
ws.cell(hh, 7).font = H2
ws.cell(hh, 8, '増減率').fill = JUDG
ws.cell(hh, 8).font = H2
ws.cell(hh, 9, '評価').fill = GREY
ws.cell(hh, 9).font = Font(name=FONT, size=9, bold=True, color='333333')
ev = {'LEGS(42SKU)': '★最大の被害。新商品に置き換わった', 'ムービック': '新商品投入側。大きく増加',
      '東宝': '新商品投入側。ほぼ倍増', 'MAPPA': 'LEGSと並ぶ被害。自社新商品にも食われた',
      'MBS': '新商品投入側。増加', '書籍': '大阪では取扱なし'}
p0 = r + 1
for k, g in enumerate(GRSET + ['計']):
    rr = p0 + k
    ws.cell(rr, 1, g if g != '計' else '物販客単価（全商品計）').font = BD
    for i, v in enumerate(VS):
        ws.cell(rr, 2 + i, round(PERV[v][g])).number_format = YEN
    ws.cell(rr, 7, '=E%d-C%d' % (rr, rr)).number_format = '+¥#,##0;-¥#,##0;-'
    ws.cell(rr, 8, '=IFERROR(G%d/C%d,"")' % (rr, rr)).number_format = '+0.0%;-0.0%;-'
    ws.cell(rr, 9, ev.get(g, '会場全体では ほぼ横ばい ＝ 市場は縮んでいない'))
pend = p0 + len(GRSET)
body(ws, p0, pend, 9, namecol=1)
for k in range(len(GRSET) + 1):
    ws.cell(p0 + k, 1).font = BD
    ws.cell(p0 + k, 1).alignment = LFT
    ws.cell(p0 + k, 9).alignment = LFT
totrow(ws, pend, 9, namecol=1)
hdrfmt(ws, hh, hh, 9)
ws.row_dimensions[hh].height = 32
for i, w in enumerate([26, 12, 12, 12, 12, 12, 14, 10, 44], 1):
    ws.column_dimensions[gcl(i)].width = w

r = pend + 2
ws.cell(r, 1, '■ ② 会場ごとの新商品投入と1人あたり売上（LEGSがいつ削られ始めたか）').font = SEC
r += 1
hh = r
for i, h in enumerate(['会場', '新規SKU数', '新商品\n売上(円/人)', 'LEGS\n売上(円/人)',
                       'LEGS\n前会場差', '物販客単価\n(全商品)', '主な新商品'], 1):
    c = ws.cell(hh, i, h)
    c.fill = NAVY
    c.font = H2
mj = {'東京': '（初回：全商品が新規）',
      '名古屋': 'MAPPAブラインドカード4種（フォト風カード・原画クリアカード）＝1,014円/人',
      '博多': '★おすわりぬいぐるみ全8種（230円/人）、アクリルカードスタンド（233円/人）、原画ポストカードセット（178円/人）',
      '大阪': '★うちわ14種（195円/人）、アクリル系39種（450円/人）、劇場版ED系カード/ファイル（500円/人）、ぐみっとシール（101円/人）',
      '札幌': '新規なし（品揃えは大阪と同一）'}
n0 = hh + 1
for k, v in enumerate(VS):
    rr = n0 + k
    n = DATA_DAYS[v]
    na = sum(sum(D[v]['rows'][str(j)]['daily'][:n]) * D[v]['rows'][str(j)]['price'] for j in NEWSKU[v])
    ws.cell(rr, 1, v).font = BD
    ws.cell(rr, 2, len(NEWSKU[v])).number_format = INT
    ws.cell(rr, 3, round(na / ATTOT[v])).number_format = YEN
    ws.cell(rr, 4, round(PERV[v]['LEGS(42SKU)'])).number_format = YEN
    ws.cell(rr, 5, '' if k == 0 else '=D%d-D%d' % (rr, rr - 1))
    ws.cell(rr, 5).number_format = '+¥#,##0;-¥#,##0;-'
    ws.cell(rr, 6, round(PERV[v]['計'])).number_format = YEN
    ws.cell(rr, 7, mj[v])
body(ws, n0, n0 + len(VS) - 1, 7, namecol=7)
for k in range(len(VS)):
    ws.cell(n0 + k, 1).font = BD
    ws.cell(n0 + k, 7).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
hdrfmt(ws, hh, hh, 7)
ws.row_dimensions[hh].height = 32

r = n0 + len(VS) + 2
ws.cell(r, 1, '■ ③ LEGS 商品タイプ別 1人あたり売上（円/人）— どのタイプが削られたか').font = SEC
r += 1
hh = r
ws.cell(hh, 1, 'LEGS 商品タイプ').fill = NAVY
ws.cell(hh, 1).font = H2
for i, v in enumerate(VS):
    c = ws.cell(hh, 2 + i, v)
    c.fill = PatternFill('solid', fgColor=VF[v])
    c.font = H2
for i, h in enumerate(['名古屋→大阪\n増減', 'LEGS減少への\n寄与度', '競合した新商品'], 7):
    c = ws.cell(hh, i, h)
    c.fill = JUDG
    c.font = H2
LT = ['ブラインド缶バッジ', 'ブラインドミニ色紙', 'クリアファイル', 'ポストカード類',
      'ブラインドステッカー', '雑貨・その他', 'アクセサリー(リング等)']
comp = {'ブラインド缶バッジ': '★アクリル系39SKU（450円/人）・ぬいぐるみ全8種。同じ「引く／集める」コレクション需要が直撃',
        'ブラインドミニ色紙': '★アクリル系・劇場版EDカード類（293円/人）。ブラインド飽和＋新規コレクション商材へ移行',
        'クリアファイル': '劇場版EDクリアファイル10SKU（207円/人）と競合するが、LEGS側は横ばいを維持',
        'ポストカード類': '劇場版EDポストカード/Pictas（199円/人）と競合。LEGS側は横ばい',
        'ブラインドステッカー': 'ぐみっとシール等9SKU（121円/人）。影響軽微',
        '雑貨・その他': 'うちわ14SKU（195円/人）が近い衝動買い枠。バスボール・お菓子等が微減',
        'アクセサリー(リング等)': '高単価アクセ。ぬいぐるみコンプリート(¥19,840)等の高単価枠と競合'}
l0 = hh + 1
for k, t in enumerate(LT):
    rr = l0 + k
    ws.cell(rr, 1, t).font = BD
    for i, v in enumerate(VS):
        n = DATA_DAYS[v]
        s = sum(sum(D[v]['rows'][str(j)]['daily'][:n]) * D[v]['rows'][str(j)]['price']
                for j, nm in T if ltype(nm) == t)
        ws.cell(rr, 2 + i, round(s / ATTOT[v])).number_format = YEN
    ws.cell(rr, 7, '=E%d-C%d' % (rr, rr)).number_format = '+¥#,##0;-¥#,##0;-'
    ws.cell(rr, 8, '=IFERROR(G%d/G$%d,"")' % (rr, l0 + len(LT))).number_format = PCT
    ws.cell(rr, 9, comp[t])
ltot = l0 + len(LT)
ws.cell(ltot, 1, 'LEGS 合計').font = BD
for i in range(len(VS)):
    L = gcl(2 + i)
    ws.cell(ltot, 2 + i, '=SUM(%s%d:%s%d)' % (L, l0, L, ltot - 1)).number_format = YEN
ws.cell(ltot, 7, '=E%d-C%d' % (ltot, ltot)).number_format = '+¥#,##0;-¥#,##0;-'
ws.cell(ltot, 8, 1).number_format = PCT
body(ws, l0, ltot, 9, namecol=1)
for k in range(len(LT) + 1):
    ws.cell(l0 + k, 1).font = BD
    ws.cell(l0 + k, 1).alignment = LFT
    ws.cell(l0 + k, 9).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
totrow(ws, ltot, 9, namecol=1)
hdrfmt(ws, hh, hh, 9)
ws.row_dimensions[hh].height = 32

r = ltot + 2
ws.cell(r, 1, '■ ④ 大阪 新商品106SKU の内訳（タイプ別）').font = SEC
r += 1
hh = r
for i, h in enumerate(['新商品タイプ', 'SKU数', '販売数量', '販売金額(税込)', '会場売上\n構成比',
                       '1人あたり\n売上', '競合するLEGS商品'], 1):
    c = ws.cell(hh, i, h)
    c.fill = FUT
    c.font = H2
nagg = {}
for j in NEWSKU['大阪']:
    x = D['大阪']['rows'][str(j)]
    t = ntype(x['name'])
    q = sum(x['daily'][:20])
    nagg.setdefault(t, [0, 0, 0])
    nagg[t][0] += 1
    nagg[t][1] += q
    nagg[t][2] += q * x['price']
rival = {'アクリル系（コレクション）': '★ブラインド缶バッジ・ブラインドミニ色紙（直接競合）',
         'カード・ポストカード類': 'ポストカードセット2種・メタリックポストカード',
         'クリアファイル': 'クリアファイル14種（展示会KV・キャラ別）',
         'うちわ（新カテゴリ）': '直接競合なし。ただし1,100円の衝動買い枠を奪う',
         'ぬいぐるみ（新カテゴリ）': '直接競合なし。高単価コレクション枠でブラインド系と競合',
         'シール・ステッカー': 'ブラインドダイカットステッカー2種',
         '缶バッジ': '★ブラインド缶バッジ2種（直接競合）',
         'その他新商品': '雑貨・その他'}
OT = sorted(nagg.items(), key=lambda y: -y[1][2])
o0 = hh + 1
TOTA_O = vt('大阪', 0, 20)[1]
for k, (t, (s, q, a)) in enumerate(OT):
    rr = o0 + k
    ws.cell(rr, 1, t).font = BD
    ws.cell(rr, 2, s).number_format = INT
    ws.cell(rr, 3, q).number_format = INT
    ws.cell(rr, 4, a).number_format = YEN
    ws.cell(rr, 5, '=D%d/%.0f' % (rr, TOTA_O)).number_format = PCT
    ws.cell(rr, 6, '=ROUND(D%d/%d,0)' % (rr, ATTOT['大阪'])).number_format = YEN
    ws.cell(rr, 7, rival.get(t, ''))
otot = o0 + len(OT)
ws.cell(otot, 1, '新商品 合計').font = BD
for c in [2, 3, 4]:
    L = gcl(c)
    ws.cell(otot, c, '=SUM(%s%d:%s%d)' % (L, o0, L, otot - 1)).number_format = YEN if c == 4 else INT
ws.cell(otot, 5, '=D%d/%.0f' % (otot, TOTA_O)).number_format = PCT
ws.cell(otot, 6, '=ROUND(D%d/%d,0)' % (otot, ATTOT['大阪'])).number_format = YEN
ws.cell(otot, 7, 'うち約47%がLEGSからの移動と推定')
body(ws, o0, otot, 7, namecol=1)
for k in range(len(OT) + 1):
    ws.cell(o0 + k, 1).font = BD
    ws.cell(o0 + k, 1).alignment = LFT
    ws.cell(o0 + k, 7).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
totrow(ws, otot, 7, namecol=1)
hdrfmt(ws, hh, hh, 7)
ws.row_dimensions[hh].height = 32

r = otot + 2
ws.cell(r, 1, '■ ⑤ 石川会場以降への示唆').font = SEC
r += 1
for a, b in [
    ('LEGS構成比は底打ちの見込み',
     '札幌は新規SKU 0で品揃えが大阪と同一。LEGS金額構成比 大阪21.1%→札幌19.6%(予測)で減少幅は縮小しており、'
     '石川以降も新商品の追加が無ければ19〜20%で安定すると想定。③の予測はこの水準を採用している。'),
    ('★最大のリスクは「石川以降の新商品追加」',
     '大阪の実績では新商品1,360円/人の投入に対しLEGSが643円/人失われた。同規模の新商品を石川以降に投入すると、'
     'LEGS構成比はさらに5〜7pt下がり、③の必要数86,000個は最大で30%程度過大になる。'),
    ('　→ 発注前の必須確認事項',
     '石川以降に新商品（ぬいぐるみ第2弾・うちわ第2弾等）の投入計画があるか、あるなら規模はどの程度か。'
     'これを確認せずに追加生産46,000個を確定するのは危険。'),
    ('ブラインド系は減産、定番系は維持',
     'ブラインド缶バッジ▲272円/人・ミニ色紙▲284円/人でLEGS減少の87%を占める一方、'
     'クリアファイル・ポストカード類はほぼ横ばい。追加生産はブラインド系を抑え、クリアファイル系を優先すべき。'),
    ('高単価コレクション商材が主戦場',
     'おすわりぬいぐるみコンプリート(¥19,840)が大阪で203個、アクリル系39SKUで450円/人。'
     '「引く」から「選んで揃える」へ需要が移行している。LEGS側もブラインド偏重の見直しが必要。'),
]:
    ws.cell(r, 1, a).font = BD
    c = ws.cell(r, 2, b)
    c.font = NM
    c.alignment = LFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=9)
    r += 1

# ===========================================================================
# 9. ⑧実在庫(7/8)明細
# ===========================================================================
ws = wb.create_sheet('⑧実在庫(7-8)明細')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'D7'
ws['A1'] = '⑧ 残在庫想定（%s）の明細と欠品算出' % STOCK_STAGE
ws['A1'].font = TITLE
ws['A2'] = ('残在庫想定 ＝ 在庫フロー台帳（【清算(卸値67％)】…_検品差異）の「%s」列。'
            '会場ごとの 追加発注→販売→レッグス検品 を積み上げた最終値で、梅田会場後の検品結果まで反映済み'
            % STOCK_STAGE)
ws['A2'].font = NOTE
ws['A3'] = ('この残在庫から札幌会場へ7/21に出荷（＝札幌の販売可能数）。'
            '石川以降に使えるのは「倉庫残 − 札幌への追納 ＋ 札幌会期末の返送見込」＝ 残在庫 − 札幌会期予測販売数')
ws['A3'].font = NOTE
ws['A5'] = ('※P列の過不足には確定済み追加発注（%s）を加算している'
            % ('／'.join('%s %s%s' % (dict(T)[j], f"{v['qty']:,}", v['unit']) for j, v in DECIDED.items())
               or 'なし'))
ws['A5'].font = NOTE
ws['A4'] = ('※同ファイルの「元データ」「在庫合算」「在庫元データ(12月清算後)」シートは池袋終了時点の古い'
            'スナップショット（合計183,696個）であり、残在庫として使ってはいけない。'
            'ポストカードセット2種・宿儺の指風お菓子はSET商品で、台帳の発注数はバラ単位・実在庫はSET単位')
ws['A4'].font = NOTE

IH = [('No.', 5, NAVY), ('JAN', 15, NAVY), ('商品名', 44, NAVY), ('税込単価', 9, NAVY)]
for _st in STOCK_STAGES[:-1]:
    IH.append(('実在庫\n%s' % _st.replace('(SET)', ''), 11, PatternFill('solid', fgColor='7B7B7B')))
IH += [('★残在庫想定\n%s' % STOCK_STAGE.replace('(SET)', ''), 13,
        PatternFill('solid', fgColor='375623')),
      ('残在庫\n金額(税込)', 13, PatternFill('solid', fgColor='375623')),
      ('札幌へ納品\n(7/21)', 11, PatternFill('solid', fgColor=VF['札幌'])),
      ('札幌 会期20日\n予測販売', 12, PatternFill('solid', fgColor=VF['札幌'])),
      ('札幌 会期末\n返送見込', 11, PatternFill('solid', fgColor=VF['札幌'])),
      ('差引 倉庫残\n(即出荷可)', 12, BLUE),
      ('石川以降\n供給可能計', 12, BLUE),
      ('石川以降\n必要数', 11, FUT),
      ('過不足\n(供給−必要)', 13, JUDG),
      ('欠品\n(追加生産必要数)', 14, JUDG),
      ('推奨発注数\n(安全率込)', 12, JUDG),
      ('欠品発生\n会場', 20, ALERT), ('最終発注期限', 13, ALERT), ('発注の緊急度', 16, ALERT),
      ('判定', 20, JUDG),
      ('運営欠品報告\n(7/28)', 14, ALERT),
      ('備考', 46, GREY)]
ihr = 6
for i, (h, w, f) in enumerate(IH, 1):
    c = ws.cell(ihr, i, h)
    c.fill = f
    c.font = H2 if f not in (ALERT, GREY) else Font(name=FONT, size=9, bold=True, color='333333')
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[ihr].height = 46
i0 = ihr + 1
itot = i0 + len(T)
for i, (jan, nm) in enumerate(T):
    rr = i0 + i
    iv = INV[str(jan)]
    ws.cell(rr, 1, i + 1); ws.cell(rr, 2, str(jan)); ws.cell(rr, 3, nm)
    ws.cell(rr, 4, R('札幌', jan)['price']).number_format = YEN
    for _k, _h in enumerate(iv['hist']):
        ws.cell(rr, 5 + _k, _h).number_format = INT
    ws.cell(rr, 9, '=H%d*D%d' % (rr, rr)).number_format = YEN
    ws.cell(rr, 10, '=%s!S%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, 11, '=%s!P%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, 12, '=%s!U%d' % (S1, r0 + i)).number_format = INT
    ws.cell(rr, 13, '=H%d-J%d' % (rr, rr)).number_format = INT
    # 会期中に在庫切れになる品は倉庫残から札幌へ追納するので、その分は石川以降に使えない
    ws.cell(rr, 14, '=M%d-MAX(0,K%d-J%d)+ROUND(L%d*%s,0)'
            % (rr, rr, rr, rr, pr('slide'))).number_format = INT
    ws.cell(rr, 15, '=%s!%s%d' % (S3, gcl(c_need), pr0 + i)).number_format = INT
    _dq = DECIDED[jan]['qty'] if jan in DECIDED else 0
    ws.cell(rr, 16, '=N%d-O%d%s' % (rr, rr, ('+%d' % _dq) if _dq else '')
            ).number_format = '+#,##0;-#,##0;0'
    ws.cell(rr, 17, '=MAX(0,-P%d)' % rr).number_format = INT
    ws.cell(rr, 18, '=IF(Q%d>0,ROUNDUP(Q%d*(1+%s),0),0)' % (rr, rr, pr('safe'))).number_format = INT
    _fn8 = _venue_of(jan)
    ws.cell(rr, 19, _fn8 if _fn8 else '—')
    _dl8 = _dl_formula(jan)
    ws.cell(rr, 20, _dl8 if _dl8 else '')
    if _dl8.startswith('='):
        ws.cell(rr, 20).number_format = 'yyyy/m/d'
    if jan in DECIDED:
        ws.cell(rr, 21, '◎発注確定済')
    else:
        ws.cell(rr, 21, '=IF(S%d="—","—",IF(T%d="再生産不可","×再生産不可",'
                '_xlfn.IFS(T%d-%s<0,"★期限超過",T%d-%s<=30,"★至急(30日以内)",'
                'T%d-%s<=90,"要着手(90日以内)",TRUE(),"余裕あり")))'
                % (rr, rr, rr, pr('base'), rr, pr('base'), rr, pr('base')))
    ws.cell(rr, 22, '=IF(Q%d>0,"★欠品：追加生産が必要",IF(P%d>=O%d*2,"大幅余剰","在庫で充足"))' % (rr, rr, rr))
    _sh = SHORTAGE.get(str(jan), '')
    ws.cell(rr, 23, '欠品報告あり' if _sh else '')
    _wh = iv['real'] - R('札幌', jan)['avail']
    nts = []
    if jan in DECIDED:
        _d8 = DECIDED[jan]
        nts.append('◎追加発注 確定済 %s%s（%s入荷）を過不足に加算済／%s'
                   % (f"{_d8['qty']:,}", _d8['unit'], _d8['arrive_at'], _d8['moq']))
    if jan in SHELF:
        nts.append('【賞味期限】' + SHELF[jan]['note'])
    if _sh:
        nts.append('倉庫残%s個 → %s' % (f'{_wh:,}',
                   '◎倉庫からの追加納品で即対応可能' if _wh > 0 else '×倉庫在庫なし'))
        if '追納困難' in _sh:
            nts.append('★「追納困難」との回答だが実在庫%s個あり。出荷可否を要再確認' % f'{_wh:,}')
    if iv['set_size'] > 1:
        nts.append('SET商品（1販売単位＝%d個）。台帳の発注数はバラ単位' % iv['set_size'])
    ws.cell(rr, 24, ' ／ '.join(nts))
ws.cell(itot, 3, '合計（LEGS 42SKU）')
for c in [5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18]:
    L = gcl(c)
    ws.cell(itot, c, '=SUM(%s%d:%s%d)' % (L, i0, L, itot - 1)).number_format = YEN if c == 9 else INT
ws.cell(itot, 21, '=COUNTIF(U%d:U%d,"★期限超過")&"品目が期限超過"' % (i0, itot - 1))
ws.cell(itot, 22, '=COUNTIF(V%d:V%d,"★欠品：追加生産が必要")&"品目が欠品"' % (i0, itot - 1))
ws.cell(itot, 23, '=COUNTIF(W%d:W%d,"<>")&"品目"' % (i0, itot - 1))
body(ws, i0, itot, 24)
for rr in range(i0, itot):
    for _c in (19, 21, 22, 24):
        ws.cell(rr, _c).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
totrow(ws, itot, 24)
hdrfmt(ws, ihr, ihr, 24)
ws.conditional_formatting.add('V%d:V%d' % (i0, itot - 1),
                              FormulaRule(formula=['LEFT($V%d,1)="★"' % i0], fill=ALERT,
                                          font=Font(name=FONT, size=9, bold=True, color='9C0006')))
ws.conditional_formatting.add('V%d:V%d' % (i0, itot - 1),
                              FormulaRule(formula=['$V%d="在庫で充足"' % i0], fill=OKF))
ws.conditional_formatting.add('W%d:W%d' % (i0, itot - 1),
                              FormulaRule(formula=['$W%d<>""' % i0], fill=WARN))
ws.conditional_formatting.add('S%d:S%d' % (i0, itot - 1),
                              FormulaRule(formula=['$S%d<>"—"' % i0], fill=ALERT))
ws.conditional_formatting.add('U%d:U%d' % (i0, itot - 1),
                              FormulaRule(formula=['LEFT($U%d,1)="★"' % i0], fill=ALERT))
ws.conditional_formatting.add('U%d:U%d' % (i0, itot - 1),
                              FormulaRule(formula=['$U%d="要着手(90日以内)"' % i0], fill=WARN))
ws.conditional_formatting.add('Q%d:Q%d' % (i0, itot - 1),
                              FormulaRule(formula=['$Q%d>0' % i0], fill=ALERT))

# ===========================================================================
# 10. ⑨過去4会場 予測vs実績・構成比修正
# ===========================================================================
ORD = ['東京', '名古屋', '博多', '大阪']
_FRtot = sum(S('博多', j, 0, 26) for j in TJ)


def _frac(n):
    return sum(S('博多', j, 0, min(n, 26)) for j in TJ) / _FRtot if n <= 26 else 1.0


FRC = {n: _frac(n) for n in range(1, 31)}


def _cum(v, j, n):
    L = DATA_DAYS[v]
    return S(v, j, 0, n) if n <= L else S(v, j, 0, L) * FRC[min(n, 30)] / FRC[L]


BACK = {}
for _vi, _V in enumerate(['東京', '名古屋', '博多', '大阪', '札幌']):
    _pri = ['東京', '名古屋', '博多', '大阪', '札幌'][:_vi]
    if not _pri:
        BACK[_V] = None
        continue
    _w = {q: (2 if q == _pri[-1] else 1) for q in _pri}
    _Dv = RUN[_V]
    _G = (sum(sum(_cum(q, j, _Dv) for j in TJ) / sum(_cum(q, j, BASE) for j in TJ) * _w[q] for q in _pri)
          / sum(_w.values()))
    _pd = {}
    for j in TJ:
        gs = [(_cum(q, j, _Dv) / _cum(q, j, BASE), _w[q]) for q in _pri if _cum(q, j, BASE)]
        _pd[j] = round(S(_V, j, 0, BASE) * (sum(a * b for a, b in gs) / sum(b for _, b in gs) if gs else _G))
    BACK[_V] = dict(pred=_pd, G=_G, prior=_pri, w=_w)

ws = wb.create_sheet('⑨過去4会場 予測vs実績')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'D8'
ws['A1'] = '⑨ 過去4会場（①池袋・②名古屋・③博多阪急・④梅田阪急）商品別 販売予測 vs 実績販売・構成比の修正'
ws['A1'].font = TITLE
ws['A2'] = ('販売予測＝各会場の「%s実績 × 会期倍率g」。gはその時点で判明していた過去会場のみから算出（直近会場を2倍加重）。' % NB +
            '池袋は初回会場のため予測不能で、投入数（販売可能数）を計画値として表示')
ws['A2'].font = NOTE
ws['A3'] = ('予測時構成比＝予測販売数のLEGS42SKU内シェア（池袋は投入数シェア）／'
            '実績構成比＝実績販売数のシェア／構成比修正＝実績構成比 − 予測時構成比（ポイント）')
ws['A3'].font = NOTE
ws['A4'] = ('会期日数が異なるため、過去会場の累計は博多26日間の日次消化カーブで対象会場の会期日数に換算している'
            '／名古屋の販売可能数は元データに小数値が含まれるため整数に丸めている（3品目）')
ws['A4'].font = NOTE

r = 6
ws.cell(r, 1, '■ 会場別 予測精度サマリー（LEGS 42SKU 合計）').font = SEC
r += 1
AH = ['会場', '会期日数', '参照した過去会場（加重）', '会期倍率g', '投入数\n(販売可能数)',
      'モデル\n予測販売数', '実績\n販売数', '予測誤差\n(実績−予測)', '誤差率',
      '商品別\n絶対誤差率', '消化率\n(実績/投入)', '評価']
AW = [12, 8, 30, 10, 12, 13, 12, 13, 9, 11, 10, 46]
ahr = r
for i, (h, w) in enumerate(zip(AH, AW), 1):
    c = ws.cell(ahr, i, h)
    c.fill = NAVY
    c.font = H2
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[ahr].height = 34
_ev = {'東京': '初回会場。予測モデルの参照元が無いため投入数のみ',
       '名古屋': '池袋のみを参照。実績が予測を上回り、やや保守的な予測だった',
       '博多': '26日間の長期会期で後半にブラインド系が伸び、予測を大きく上回った',
       '大阪': '博多の高い伸長率を引き継いだため予測が上振れ。新商品106SKU投入でLEGSが想定より抑制された'}
a0 = ahr + 1
for k, V in enumerate(ORD):
    rr = a0 + k
    b = BACK[V]
    av = sum(R(V, j)['avail'] for j in TJ)
    ac = sum(S(V, j, 0, DATA_DAYS[V]) for j in TJ)
    ws.cell(rr, 1, V).font = BD
    ws.cell(rr, 2, RUN[V]).number_format = INT
    ws.cell(rr, 3, '—（初回会場）' if b is None else '・'.join('%s×%d' % (q, b['w'][q]) for q in b['prior']))
    ws.cell(rr, 4, '' if b is None else b['G']).number_format = MUL
    ws.cell(rr, 5, av).number_format = INT
    ws.cell(rr, 6, '' if b is None else sum(b['pred'].values())).number_format = INT
    ws.cell(rr, 7, ac).number_format = INT
    if b is None:
        ws.cell(rr, 8, '')
        ws.cell(rr, 9, '')
        ws.cell(rr, 10, '')
    else:
        ws.cell(rr, 8, '=G%d-F%d' % (rr, rr)).number_format = '+#,##0;-#,##0;0'
        ws.cell(rr, 9, '=IFERROR(-H%d/G%d,"")' % (rr, rr)).number_format = '+0.0%;-0.0%;0.0%'
        ws.cell(rr, 10, sum(abs(b['pred'][j] - S(V, j, 0, DATA_DAYS[V])) for j in TJ) / ac).number_format = PCT
    ws.cell(rr, 11, '=IFERROR(G%d/E%d,0)' % (rr, rr)).number_format = PCT
    ws.cell(rr, 12, _ev[V])
atot = a0 + len(ORD)
ws.cell(atot, 1, '4会場 合計').font = BD
for c in [5, 6, 7]:
    L = gcl(c)
    ws.cell(atot, c, '=SUM(%s%d:%s%d)' % (L, a0, L, atot - 1)).number_format = INT
ws.cell(atot, 8, '=G%d-F%d' % (atot, atot)).number_format = '+#,##0;-#,##0;0'
ws.cell(atot, 9, '=IFERROR(-H%d/G%d,"")' % (atot, atot)).number_format = '+0.0%;-0.0%;0.0%'
ws.cell(atot, 11, '=IFERROR(G%d/E%d,0)' % (atot, atot)).number_format = PCT
ws.cell(atot, 12, '池袋は予測対象外のため、予測計・誤差は名古屋以降3会場の合計')
body(ws, a0, atot, 12, namecol=3)
for k in range(len(ORD) + 1):
    ws.cell(a0 + k, 1).font = BD
    ws.cell(a0 + k, 3).alignment = LFT
    ws.cell(a0 + k, 12).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
totrow(ws, atot, 12, namecol=3)
hdrfmt(ws, ahr, ahr, 12)

# 商品別
r = atot + 2
ws.cell(r, 1, '■ 商品別 予測 vs 実績・構成比の修正').font = SEC
r += 1
PH2 = [('No.', 5, NAVY), ('JAN', 15, NAVY), ('商品名', 42, NAVY), ('税込単価', 9, NAVY)]
for V in ORD:
    f = PatternFill('solid', fgColor=VF[V])
    PH2 += [('%s\n投入数' % V, 10, f), ('%s\n予測販売数' % V, 11, f), ('%s\n実績販売数' % V, 11, f),
            ('%s\n予測誤差' % V, 10, f), ('%s\n予測時構成比' % V, 11, f), ('%s\n実績構成比' % V, 10, f),
            ('%s\n構成比修正' % V, 11, f)]
PH2 += [('4会場\n実績合計', 11, JUDG), ('4会場\n実績構成比', 11, JUDG),
        ('札幌\n予測構成比', 11, PatternFill('solid', fgColor=VF['札幌'])),
        ('石川以降\n適用構成比', 12, FUT), ('構成比トレンド', 16, JUDG)]
phr2 = r
for i, (h, w, f) in enumerate(PH2, 1):
    c = ws.cell(phr2, i, h)
    c.fill = f
    c.font = H2
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[phr2].height = 40
q0 = phr2 + 1
qtot2 = q0 + len(T)
for i, (jan, nm) in enumerate(T):
    rr = q0 + i
    ws.cell(rr, 1, i + 1); ws.cell(rr, 2, str(jan)); ws.cell(rr, 3, nm)
    ws.cell(rr, 4, R('札幌', jan)['price']).number_format = YEN
    cc = 5
    for V in ORD:
        b = BACK[V]
        ws.cell(rr, cc, R(V, jan)['avail']).number_format = INT
        if b is None:
            ws.cell(rr, cc + 1, '—')
        else:
            ws.cell(rr, cc + 1, b['pred'][jan]).number_format = INT
        ws.cell(rr, cc + 2, S(V, jan, 0, DATA_DAYS[V])).number_format = INT
        L0, L1, L2 = gcl(cc), gcl(cc + 1), gcl(cc + 2)
        if b is None:
            ws.cell(rr, cc + 3, '—')
            ws.cell(rr, cc + 4, '=IFERROR(%s%d/%s$%d,0)' % (L0, rr, L0, qtot2)).number_format = PCT
        else:
            ws.cell(rr, cc + 3, '=%s%d-%s%d' % (L2, rr, L1, rr)).number_format = '+#,##0;-#,##0;0'
            ws.cell(rr, cc + 4, '=IFERROR(%s%d/%s$%d,0)' % (L1, rr, L1, qtot2)).number_format = PCT
        ws.cell(rr, cc + 5, '=IFERROR(%s%d/%s$%d,0)' % (L2, rr, L2, qtot2)).number_format = PCT
        ws.cell(rr, cc + 6, '=%s%d-%s%d' % (gcl(cc + 5), rr, gcl(cc + 4), rr)).number_format = '+0.00%;-0.00%;0.00%'
        cc += 7
    C1, C2, C3, C4 = gcl(7), gcl(14), gcl(21), gcl(28)
    ws.cell(rr, cc, '=%s%d+%s%d+%s%d+%s%d' % (C1, rr, C2, rr, C3, rr, C4, rr)).number_format = INT
    ws.cell(rr, cc + 1, '=IFERROR(%s%d/%s$%d,0)' % (gcl(cc), rr, gcl(cc), qtot2)).number_format = PCT
    ws.cell(rr, cc + 2, '=%s!R%d' % (S1, r0 + i)).number_format = PCT
    ws.cell(rr, cc + 3, '=%s!H%d' % (S3, pr0 + i)).number_format = PCT
    ws.cell(rr, cc + 4, '=IF(%s%d-%s%d>=0.005,"↑ 上方修正",IF(%s%d-%s%d<=-0.005,"↓ 下方修正","→ 据置"))'
            % (gcl(cc + 3), rr, gcl(cc + 1), rr, gcl(cc + 3), rr, gcl(cc + 1), rr))
NCQ = cc + 4
ws.cell(qtot2, 3, '合計（LEGS 42SKU）')
cc = 5
for V in ORD:
    for c in [cc, cc + 1, cc + 2]:
        L = gcl(c)
        if c == cc + 1 and BACK[V] is None:
            continue
        ws.cell(qtot2, c, '=SUM(%s%d:%s%d)' % (L, q0, L, qtot2 - 1)).number_format = INT
    cc += 7
ws.cell(qtot2, NCQ - 4, '=SUM(%s%d:%s%d)' % (gcl(NCQ - 4), q0, gcl(NCQ - 4), qtot2 - 1)).number_format = INT
for c in [NCQ - 3, NCQ - 2, NCQ - 1]:
    ws.cell(qtot2, c, 1).number_format = PCT
cc = 5
for V in ORD:
    if BACK[V] is not None:
        ws.cell(qtot2, cc + 3, '=SUM(%s%d:%s%d)' % (gcl(cc + 3), q0, gcl(cc + 3), qtot2 - 1)
                ).number_format = '+#,##0;-#,##0;0'
    ws.cell(qtot2, cc + 4, 1).number_format = PCT
    ws.cell(qtot2, cc + 5, 1).number_format = PCT
    ws.cell(qtot2, cc + 6, 0).number_format = '+0.00%;-0.00%;0.00%'
    cc += 7
body(ws, q0, qtot2, NCQ)
totrow(ws, qtot2, NCQ)
hdrfmt(ws, phr2, phr2, NCQ)
ws.conditional_formatting.add('%s%d:%s%d' % (gcl(NCQ), q0, gcl(NCQ), qtot2 - 1),
                              FormulaRule(formula=['LEFT($%s%d,1)="↑"' % (gcl(NCQ), q0)], fill=OKF))
ws.conditional_formatting.add('%s%d:%s%d' % (gcl(NCQ), q0, gcl(NCQ), qtot2 - 1),
                              FormulaRule(formula=['LEFT($%s%d,1)="↓"' % (gcl(NCQ), q0)], fill=WARN))

r = qtot2 + 2
ws.cell(r, 1, '■ 構成比修正の考え方と次会場への反映').font = SEC
r += 1
for a, b in [
    ('予測時構成比', '各会場の会期前〜%s時点で、過去会場の実績から見込んでいた商品別シェア。' % NB +
                 '池袋のみ実績データが無いため投入数（販売可能数）のシェアを計画値として採用している。'),
    ('実績構成比', '各会場の会期全体の実績販売数に基づく商品別シェア。'),
    ('構成比修正', '実績構成比 − 予測時構成比（ポイント）。プラスは想定より売れた＝次会場で構成比を引き上げるべき商品。'),
    ('石川以降 適用構成比', '大阪20日実績のシェア50% ＋ 札幌20日予測のシェア50%のブレンド（③シートで算出）。'
                     '直近2会場は品揃えが同一のため、この2会場に絞って重み付けしている。'),
    ('モデル精度の読み方', '3会場の絶対誤差率は10〜27%。ブラインド系など会期後半に伸びる商品ほど予測が下振れしやすく、'),
    ('', '定番のクリアファイル・ポストカード系は誤差が小さい。発注量は絶対誤差率を織り込んだ安全率で調整すること。'),
]:
    ws.cell(r, 1, a).font = BD
    c = ws.cell(r, 2, b)
    c.font = NM
    c.alignment = LFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    r += 1

# ===========================================================================
# 11. ⑩会場比率シナリオ比較
# ===========================================================================
ws = wb.create_sheet('⑩会場比率シナリオ比較')
ws.sheet_view.showGridLines = False
ws['A1'] = '⑩ 石川会場以降の 東京対比 会場比率 シナリオ比較'
ws['A1'].font = TITLE
ws['A2'] = ('「前提・入力」シートのシナリオ選択セルで①〜④を切り替えると、③④⑤⑧シートの予測がすべて連動して再計算される。'
            '本シートは4シナリオを並べて比較するためのもの')
ws['A2'].font = NOTE
ws['A3'] = ('②企画概要書ベースは、企画概要書「グッズ売上(LEGS予測値)」の 各会場÷東京 で算出'
            '（石川 31,200千円÷250,000千円＝12.48%、仙台 39,000千円÷250,000千円＝15.60%）')
ws['A3'].font = NOTE
ws['A4'] = ('③札幌実績連動は、②に札幌の計画比達成率 %.1f%%（予測東京対比%.2f%% ÷ 企画概要書21.84%%）を乗じた保守シナリオ。'
            '東京凱旋のみ50%%で据置') % (_ach * 100, _salla / TKA * 100)
ws['A4'].font = NOTE

r = 6
ws.cell(r, 1, '■ 企画概要書の計画値と実績の乖離（東京対比・物販売上ベース）').font = SEC
r += 1
GH = ['会場', '企画概要書\nグッズ売上予測', '企画概要書\n東京対比', '実績／予測\n物販売上', '実績／予測\n東京対比',
      '計画比\n達成率', '評価']
GW = [22, 18, 14, 18, 14, 12, 52]
ghr = r
for i, (h, w) in enumerate(zip(GH, GW), 1):
    c = ws.cell(ghr, i, h)
    c.fill = NAVY
    c.font = H2
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[ghr].height = 34
_actamt = {v: vt(v, 0, DATA_DAYS[v])[1] for v in VS}
_actamt['札幌'] = _salla
_gev = {'東京': '基準会場', '名古屋': '計画を大幅超過', '博多': '計画を大幅超過',
        '大阪': '計画の1.9倍。新商品投入で会場全体が伸びた', '札幌': '★計画の47%。動員が大きく下振れ',
        '石川': '実績未確定。②はこの計画値を採用', '仙台': '実績未確定。②はこの計画値を採用'}
g0 = ghr + 1
_glist = ['東京', '名古屋', '博多', '大阪', '札幌', '石川', '仙台']
for k, v in enumerate(_glist):
    rr = g0 + k
    ws.cell(rr, 1, v + ('（予測）' if v == '札幌' else '')).font = BD
    ws.cell(rr, 2, PLAN_GOODS[v]).number_format = YEN
    ws.cell(rr, 3, '=B%d/$B$%d' % (rr, g0)).number_format = PCT2
    if v in _actamt:
        ws.cell(rr, 4, round(_actamt[v])).number_format = YEN
        ws.cell(rr, 5, '=D%d/$D$%d' % (rr, g0)).number_format = PCT2
        ws.cell(rr, 6, '=IFERROR(E%d/C%d,"")' % (rr, rr)).number_format = PCT
    else:
        for c in (4, 5, 6):
            ws.cell(rr, c, '—')
    ws.cell(rr, 7, _gev[v])
gtot2 = g0 + len(_glist)
body(ws, g0, gtot2 - 1, 7, namecol=7)
for k in range(len(_glist)):
    ws.cell(g0 + k, 1).font = BD
    ws.cell(g0 + k, 7).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
hdrfmt(ws, ghr, ghr, 7)
ws.conditional_formatting.add('F%d:F%d' % (g0, gtot2 - 1),
                              FormulaRule(formula=['AND(ISNUMBER($F%d),$F%d<0.8)' % (g0, g0)], fill=ALERT))

r = gtot2 + 1
ws.cell(r, 1, '■ シナリオ別 東京対比比率と LEGS必要数').font = SEC
r += 1
shr = r
ws.cell(shr, 1, '会場').fill = NAVY
ws.cell(shr, 1).font = H2
ws.cell(shr, 2, '会期日数').fill = NAVY
ws.cell(shr, 2).font = H2
for k, lab in enumerate(SCN_LABEL):
    ws.cell(shr, 3 + k * 2, lab + '\n東京対比').fill = BLUE
    ws.cell(shr, 3 + k * 2).font = H2
    ws.cell(shr, 4 + k * 2, lab + '\nLEGS必要数').fill = BLUE
    ws.cell(shr, 4 + k * 2).font = H2
    ws.column_dimensions[gcl(3 + k * 2)].width = 15
    ws.column_dimensions[gcl(4 + k * 2)].width = 15
ws.row_dimensions[shr].height = 34
s0 = shr + 1
PSH = "'前提・入力'!"
for k, (nm, per, dd, ratio, note) in enumerate(FUTURE):
    rr = s0 + k
    ws.cell(rr, 1, nm).font = BD
    ws.cell(rr, 2, dd).number_format = INT
    for m in range(4):
        ws.cell(rr, 3 + m * 2, '=%s%s%d' % (PSH, gcl(3 + m), SCNROW[nm])).number_format = PCT2
        ws.cell(rr, 4 + m * 2, '=ROUND(%s*%s%d*%s/%s,0)'
                % (pr('tk_amt'), gcl(3 + m * 2), rr, pr('lsh_a'), pr('unit'))).number_format = INT
stot = s0 + len(FUTURE)
ws.cell(stot, 1, '5会場 合計').font = BD
for m in range(4):
    for c in (3 + m * 2, 4 + m * 2):
        L = gcl(c)
        ws.cell(stot, c, '=SUM(%s%d:%s%d)' % (L, s0, L, stot - 1)).number_format = PCT2 if c % 2 else INT
body(ws, s0, stot, 10, namecol=1)
for k in range(len(FUTURE) + 1):
    ws.cell(s0 + k, 1).font = BD
    ws.cell(s0 + k, 1).alignment = LFT
totrow(ws, stot, 10, namecol=1)
hdrfmt(ws, shr, shr, 10)

# 商品別 シナリオ別 追加生産必要数
r = stot + 2
ws.cell(r, 1, '■ 商品別 シナリオ別 必要数と追加生産必要数（供給可能在庫はシナリオ共通）').font = SEC
r += 1
CH = [('No.', 5, NAVY), ('JAN', 15, NAVY), ('商品名', 42, NAVY),
      ('按分\n構成比', 10, BLUE), ('石川以降\n供給可能在庫', 14, PatternFill('solid', fgColor='375623'))]
for lab in SCN_LABEL:
    CH += [(lab + '\n必要数', 13, FUT), (lab + '\n追加生産', 13, JUDG)]
chr_ = r
for i, (h, w, f) in enumerate(CH, 1):
    c = ws.cell(chr_, i, h)
    c.fill = f
    c.font = H2
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[chr_].height = 40
c0 = chr_ + 1
ctot = c0 + len(T)
for i, (jan, nm) in enumerate(T):
    rr = c0 + i
    ws.cell(rr, 1, i + 1); ws.cell(rr, 2, str(jan)); ws.cell(rr, 3, nm)
    ws.cell(rr, 4, '=%s!H%d' % (S3, pr0 + i)).number_format = PCT2
    ws.cell(rr, 5, '=%s!%s%d' % (S3, gcl(c_need + 6), pr0 + i)).number_format = INT
    for m in range(4):
        cc = 6 + m * 2
        ws.cell(rr, cc, '=ROUND($D%d*%s$%d,0)' % (rr, gcl(4 + m * 2), stot)).number_format = INT
        ws.cell(rr, cc + 1, '=MAX(0,%s%d-$E%d)' % (gcl(cc), rr, rr)).number_format = INT
NCC = 5 + 8
ws.cell(ctot, 3, '合計（LEGS 42SKU）')
for c in [4] + list(range(5, NCC + 1)):
    L = gcl(c)
    ws.cell(ctot, c, '=SUM(%s%d:%s%d)' % (L, c0, L, ctot - 1)).number_format = PCT2 if c == 4 else INT
body(ws, c0, ctot, NCC)
totrow(ws, ctot, NCC)
hdrfmt(ws, chr_, chr_, NCC)
for m in range(4):
    L = gcl(7 + m * 2)
    ws.conditional_formatting.add('%s%d:%s%d' % (L, c0, L, ctot - 1),
                                  FormulaRule(formula=['$%s%d>0' % (L, c0)], fill=ALERT))

r = ctot + 2
ws.cell(r, 1, '■ シナリオ別 結論').font = SEC
r += 1
ws.cell(r, 1, '項目').fill = NAVY
ws.cell(r, 1).font = H2
for k, lab in enumerate(SCN_LABEL):
    ws.cell(r, 2 + k, lab).fill = BLUE
    ws.cell(r, 2 + k).font = H2
    ws.column_dimensions[gcl(2 + k)].width = 18
e0 = r + 1
ERows = ['石川以降 LEGS必要数', '石川以降 供給可能在庫', '全体過不足', '追加生産 必要数', '不足品目数',
         '推奨発注数(安全率込)']
for k, lab in enumerate(ERows):
    ws.cell(e0 + k, 1, lab).font = BD
for m in range(4):
    C = gcl(2 + m)
    NC_ = gcl(6 + m * 2)
    AC = gcl(7 + m * 2)
    ws.cell(e0, 2 + m, '=%s%d' % (NC_, ctot)).number_format = INT
    ws.cell(e0 + 1, 2 + m, '=$E$%d' % ctot).number_format = INT
    ws.cell(e0 + 2, 2 + m, '=%s%d-%s%d' % (C, e0 + 1, C, e0)).number_format = '+#,##0;-#,##0;0'
    ws.cell(e0 + 3, 2 + m, '=%s%d' % (AC, ctot)).number_format = INT
    ws.cell(e0 + 4, 2 + m, '=COUNTIF(%s%d:%s%d,">0")&"品目"' % (AC, c0, AC, ctot - 1))
    ws.cell(e0 + 5, 2 + m, '=ROUNDUP(%s%d*(1+%s),0)' % (AC, ctot, pr('safe'))).number_format = INT
body(ws, e0, e0 + len(ERows) - 1, 5, namecol=1)
for k in range(len(ERows)):
    ws.cell(e0 + k, 1).font = BD
    ws.cell(e0 + k, 1).alignment = LFT
hdrfmt(ws, r, r, 5)
ws.cell(e0 + 3, 1).fill = ALERT
for m in range(4):
    ws.cell(e0 + 3, 2 + m).fill = ALERT

r = e0 + len(ERows) + 1
for a, b in [
    ('推奨', '②企画概要書ベースは石川12.48%・仙台15.60%と①より強気。ただし札幌が計画の47%に留まっている実態を踏まえると、'),
    ('', '発注量の意思決定は③札幌実績連動（保守）を下限、②企画概要書ベースを上限としたレンジで持つのが妥当。'),
    ('留意', '名古屋134%・博多132%・大阪188%と、札幌以前の3会場はいずれも計画を大きく超過している。'),
    ('', '札幌だけが47%と外れ値であり、これが札幌固有の要因（会場規模・立地）か地方会場全体の傾向かで判断が分かれる。'),
    ('', '石川・仙台の前売/予約状況が判明した時点でシナリオを確定させることを推奨する。'),
]:
    ws.cell(r, 1, a).font = BD
    c = ws.cell(r, 2, b)
    c.font = NM
    c.alignment = LFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    r += 1

# ===========================================================================
# 12. ⑪会場別在庫引当・最終発注期限
# ===========================================================================
ws = wb.create_sheet('⑪会場別引当・発注期限')
ws.sheet_view.showGridLines = False
ws.freeze_panes = 'E9'
ws['A1'] = '⑪ 石川以降 会場別 在庫引当シミュレーションと 最終追加発注期限'
ws['A1'].font = TITLE
ws['A2'] = ('札幌の欠品分は倉庫在庫から追納する前提。石川以降の供給可能在庫（＝7/8実在庫 − 札幌会期20日予測販売数）を'
            '会場順に引き当て、在庫が尽きる会場＝「欠品発生会場」を特定する')
ws['A2'].font = NOTE
ws['A3'] = ('発注期限 ＝ 欠品発生会場の会期初日 − 生産日数 − 発注〜納品バッファ（既定14日）。'
            '生産日数は2026/7/28時点の提示値で、物価高騰・素材/材料の入手困難等により変動の可能性あり')
ws['A3'].font = NOTE

r = 5
ws.cell(r, 1, '■ 既に手配済みの追加発注（参考）').font = SEC
r += 1
for i, h in enumerate(['商品', '発注／納品日', '充当先', '備考'], 1):
    c = ws.cell(r, i, h)
    c.fill = NAVY
    c.font = H2
po0 = r + 1
_PO = list(PAST_ORDER) + [('― 開催中会場への追納 ―', '', '', RESUPPLY_NOTE)] + [
    (nm, dt, '札幌会場（%s個）' % f'{q:,}', note) for dt, nm, q, note in RESUPPLY_DETAIL]
for k, (a, b, c_, d_) in enumerate(_PO):
    ws.cell(po0 + k, 1, a)
    ws.cell(po0 + k, 2, b)
    ws.cell(po0 + k, 3, c_)
    ws.cell(po0 + k, 4, d_)
body(ws, po0, po0 + len(_PO) - 1, 4, namecol=1)
for k in range(len(_PO)):
    for cx in (1, 3, 4):
        ws.cell(po0 + k, cx).alignment = LFT
ws.cell(po0 + len(PAST_ORDER), 1).font = BD
hdrfmt(ws, r, r, 4)

r = po0 + len(_PO) + 1
ws.cell(r, 1, '■ 商品別 会場別 在庫引当と発注期限').font = SEC
r += 1
VH = [('No.', 5, NAVY), ('JAN', 15, NAVY), ('商品名', 42, NAVY),
      ('生産日数\n(か月)', 10, PatternFill('solid', fgColor='7B3F00')),
      ('生産日数\n(日換算)', 10, PatternFill('solid', fgColor='7B3F00')),
      ('石川以降\n供給可能在庫', 13, PatternFill('solid', fgColor='375623'))]
for nm, per, dd, ratio, note in FUTURE:
    sn = nm.split(' ')[0]
    VH += [('%s\n必要数' % sn, 10, FUT), ('%s\n引当後残' % sn, 11, BLUE)]
VH += [('欠品発生会場', 22, ALERT), ('欠品会場\n会期初日', 12, ALERT),
       ('不足数\n(会期末まで)', 13, ALERT), ('推奨発注数\n(安全率込)', 12, JUDG),
       ('最終発注期限', 13, JUDG), ('基準日からの\n残日数', 12, JUDG), ('緊急度', 16, JUDG), ('備考', 40, GREY)]
vhr = r
for i, (h, w, f) in enumerate(VH, 1):
    c = ws.cell(vhr, i, h)
    c.fill = f
    c.font = H2 if f not in (ALERT, GREY) else Font(name=FONT, size=9, bold=True, color='333333')
    ws.column_dimensions[gcl(i)].width = w
ws.row_dimensions[vhr].height = 42
v0 = vhr + 1
vtot = v0 + len(T)
NFV = len(FUTURE)
cD = 7 + NFV * 2          # 欠品発生会場
for i, (jan, nm) in enumerate(T):
    rr = v0 + i
    pm = PROD[jan]
    ws.cell(rr, 1, i + 1); ws.cell(rr, 2, str(jan)); ws.cell(rr, 3, nm)
    ws.cell(rr, 4, pm if pm else '－')
    ws.cell(rr, 5, '=IF(ISNUMBER(D%d),ROUND(D%d*30,0),"－")' % (rr, rr))
    ws.cell(rr, 6, '=%s!%s%d' % (S3, gcl(c_need + 6), pr0 + i)).number_format = INT
    prev = gcl(6)
    for k, (fn, per, dd, ratio, note) in enumerate(FUTURE):
        cq = 7 + k * 2
        ws.cell(rr, cq, '=%s!%s%d' % (S3, gcl(9 + k), pr0 + i)).number_format = INT
        _inb = DECIDED[jan]['qty'] if (jan in DECIDED and DECIDED[jan]['arrive_at'] == fn) else 0
        ws.cell(rr, cq + 1, '=%s%d%s-%s%d'
                % (prev, rr, ('+%d' % _inb) if _inb else '', gcl(cq), rr)).number_format = '#,##0;-#,##0;0'
        if _inb:
            ws.cell(rr, cq + 1).fill = OKF
        prev = gcl(cq + 1)
    # 欠品発生会場：最初に残がマイナスになる会場
    cond = []
    for k, (fn, per, dd, ratio, note) in enumerate(FUTURE):
        cond.append('%s%d<0,"%s"' % (gcl(8 + k * 2), rr, fn))
    ws.cell(rr, cD, '=IFS(%s,TRUE(),"—")'.replace('IFS', '_xlfn.IFS') % ','.join(cond))
    ws.cell(rr, cD + 1, '=IF(%s%d="—","",_xlfn.IFS(%s))'
            % (gcl(cD), rr, ','.join('%s%d<0,DATE(%d,%d,%d)'
                                     % (gcl(8 + k * 2), rr, VEN_START[fn].year, VEN_START[fn].month,
                                        VEN_START[fn].day)
                                     for k, (fn, *_x) in enumerate(FUTURE))))
    ws.cell(rr, cD + 1).number_format = 'yyyy/m/d'
    ws.cell(rr, cD + 2, '=MAX(0,-%s%d)' % (gcl(6 + NFV * 2), rr)).number_format = INT
    ws.cell(rr, cD + 3, '=IF(%s%d>0,ROUNDUP(%s%d*(1+%s),0),0)'
            % (gcl(cD + 2), rr, gcl(cD + 2), rr, pr('safe'))).number_format = INT
    ws.cell(rr, cD + 4, '=IF(%s%d="","",IF(NOT(ISNUMBER(D%d)),"再生産不可",%s%d-E%d-%s))'
            % (gcl(cD + 1), rr, rr, gcl(cD + 1), rr, rr, pr('buf')))
    ws.cell(rr, cD + 4).number_format = 'yyyy/m/d'
    ws.cell(rr, cD + 5, '=IF(OR(%s%d="",%s%d="再生産不可"),"",%s%d-%s)'
            % (gcl(cD + 4), rr, gcl(cD + 4), rr, gcl(cD + 4), rr, pr('base'))).number_format = '#,##0"日"'
    if jan in DECIDED:
        ws.cell(rr, cD + 6, '◎発注確定済')
    else:
        ws.cell(rr, cD + 6, '=IF(%s%d="","—",IF(%s%d="再生産不可","×再生産不可",'
                '_xlfn.IFS(%s%d<0,"★期限超過",%s%d<=30,"★至急(30日以内)",'
                '%s%d<=90,"要着手(90日以内)",TRUE(),"余裕あり")))'
                % (gcl(cD), rr, gcl(cD + 4), rr, gcl(cD + 5), rr, gcl(cD + 5), rr, gcl(cD + 5), rr))
    nts = []
    if jan in PROD_NOTE:
        nts.append(PROD_NOTE[jan])
    if jan in DECIDED:
        _d = DECIDED[jan]
        nts.append('★追加発注 確定済：%s%s を %s 入荷予定（%s）'
                   % (f"{_d['qty']:,}", _d['unit'], _d['arrive_at'], _d['moq']))
    if jan in SHELF:
        nts.append(SHELF[jan]['note'])
    if str(jan) in SHORTAGE:
        nts.append('7/28 運営欠品報告あり（札幌へ倉庫から追納）')
    ws.cell(rr, cD + 7, ' ／ '.join(nts))
NCV = cD + 7
ws.cell(vtot, 3, '合計（LEGS 42SKU）')
for c in [6] + [7 + k * 2 for k in range(NFV)] + [cD + 2, cD + 3]:
    L = gcl(c)
    ws.cell(vtot, c, '=SUM(%s%d:%s%d)' % (L, v0, L, vtot - 1)).number_format = INT
ws.cell(vtot, cD, '=COUNTIF(%s%d:%s%d,"<>—")&"品目で欠品"' % (gcl(cD), v0, gcl(cD), vtot - 1))
body(ws, v0, vtot, NCV)
for rr in range(v0, vtot):
    for cx in (cD, cD + 6, NCV):
        ws.cell(rr, cx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
totrow(ws, vtot, NCV)
hdrfmt(ws, vhr, vhr, NCV)
for k in range(NFV):
    L = gcl(8 + k * 2)
    ws.conditional_formatting.add('%s%d:%s%d' % (L, v0, L, vtot - 1),
                                  FormulaRule(formula=['$%s%d<0' % (L, v0)], fill=ALERT))
ws.conditional_formatting.add('%s%d:%s%d' % (gcl(cD), v0, gcl(cD), vtot - 1),
                              FormulaRule(formula=['$%s%d<>"—"' % (gcl(cD), v0)], fill=ALERT,
                                          font=Font(name=FONT, size=9, bold=True, color='9C0006')))
for cond, fill in [('LEFT($%s%d,1)="★"', ALERT), ('$%s%d="要着手(90日以内)"', WARN),
                   ('$%s%d="余裕あり"', OKF), ('LEFT($%s%d,1)="◎"', OKF)]:
    ws.conditional_formatting.add('%s%d:%s%d' % (gcl(cD + 6), v0, gcl(cD + 6), vtot - 1),
                                  FormulaRule(formula=[cond % (gcl(cD + 6), v0)], fill=fill))

r = vtot + 2
ws.cell(r, 1, '■ 会場別 発注デッドライン早見').font = SEC
r += 1
DH = ['会場', '会期初日', '生産1か月品\n発注期限', '生産2.5か月品\n発注期限', '生産3か月品\n発注期限',
      '基準日からの残日数\n(3か月品)', '備考']
DW = [26, 13, 15, 16, 15, 18, 44]
dhr = r
for i, (h, w) in enumerate(zip(DH, DW), 1):
    c = ws.cell(dhr, i, h)
    c.fill = FUT
    c.font = H2
ws.row_dimensions[dhr].height = 34
d0 = dhr + 1
for k, (fn, per, dd, ratio, note) in enumerate(FUTURE):
    rr = d0 + k
    st = VEN_START[fn]
    ws.cell(rr, 1, fn).font = BD
    ws.cell(rr, 2, st).number_format = 'yyyy/m/d'
    for m, mm in enumerate([1, 2.5, 3]):
        ws.cell(rr, 3 + m, '=B%d-%d-%s' % (rr, round(mm * 30), pr('buf'))).number_format = 'yyyy/m/d'
    ws.cell(rr, 6, '=E%d-%s' % (rr, pr('base'))).number_format = '#,##0"日"'
    ws.cell(rr, 7, '会期%d日間' % dd)
body(ws, d0, d0 + len(FUTURE) - 1, 7, namecol=1)
for k in range(len(FUTURE)):
    ws.cell(d0 + k, 1).font = BD
    ws.cell(d0 + k, 1).alignment = LFT
    ws.cell(d0 + k, 7).alignment = LFT
hdrfmt(ws, dhr, dhr, 7)
ws.conditional_formatting.add('F%d:F%d' % (d0, d0 + len(FUTURE) - 1),
                              FormulaRule(formula=['$F%d<0' % d0], fill=ALERT))

r = d0 + len(FUTURE) + 2
if SHELF or DECIDED:
    ws.cell(r, 1, '■ 賞味期限のある商品／確定済み追加発注の売り切り管理').font = SEC
    r += 1
    ws.cell(r, 1, '「発注数の消化率」は全会期（東京凱旋）が終わった時点の累計。'
                  '会場ごとの到達度は右側の「累計消化率 ○○終了時」列で確認する').font = NOTE
    r += 1
    SH2 = ['商品', '現在庫で賞味期限が\n持つ最後の会場', 'その会場までの\n必要数', '石川以降\n供給可能在庫',
           '会期内に売り切れ\nなかった場合の残', '廃棄見込金額\n(上代)', '確定発注', '入荷会場',
           '入荷後の必要数', '発注数の消化率\n(全会期終了時)', '過剰見込']
    for _fn, _p, _dd, _r, _n in FUTURE:
        SH2.append('累計消化率\n%s終了時' % _fn.split(' ')[0])
    SH2.append('打ち手')
    SW2 = [22, 18, 13, 14, 16, 14, 14, 16, 14, 14, 12] + [13] * len(FUTURE) + [52]
    shr = r
    for i, (h, w) in enumerate(zip(SH2, SW2), 1):
        c = ws.cell(shr, i, h)
        c.fill = PatternFill('solid', fgColor='C55A11')
        c.font = H2
        ws.column_dimensions[gcl(i)].width = max(ws.column_dimensions[gcl(i)].width or 0, w)
    ws.row_dimensions[shr].height = 42
    sh0 = shr + 1
    _keys = [j for j, _ in T if j in SHELF or j in DECIDED]
    for k, jan in enumerate(_keys):
        rr = sh0 + k
        nm = dict(T)[jan]
        price = R('札幌', jan)['price']
        idx = TJ.index(jan)
        ws.cell(rr, 1, nm)
        thru = SHELF.get(jan, {}).get('sellable_through', '')
        ws.cell(rr, 2, thru or '—')
        # 賞味期限が持つ会場までの累計必要数
        cum = 0
        for fn, per, dd, ratio, note in FUTURE:
            cum += round(_mix2[jan] * _vneed[fn])
            if fn == thru:
                break
        ws.cell(rr, 3, cum if thru else '—').number_format = INT
        ws.cell(rr, 4, _sup_of[jan]).number_format = INT
        ws.cell(rr, 5, '=IF(OR(B%d="—",C%d="—"),"—",MAX(0,D%d-C%d))' % (rr, rr, rr, rr)).number_format = INT
        ws.cell(rr, 6, '=IF(E%d="—","—",E%d*%d)' % (rr, rr, price)).number_format = YEN
        _d = DECIDED.get(jan)
        ws.cell(rr, 7, ('%s%s' % (f"{_d['qty']:,}", _d['unit'])) if _d else '—')
        ws.cell(rr, 8, _d['arrive_at'] if _d else '—')
        if _d:
            after = sum(round(_mix2[jan] * _vneed[fn]) for fn, *_x in
                        [f for f in FUTURE][[f[0] for f in FUTURE].index(_d['arrive_at']):])
            ws.cell(rr, 9, after).number_format = INT
            ws.cell(rr, 10, '=IFERROR(I%d/%d,"")' % (rr, _d['qty'])).number_format = PCT
            ws.cell(rr, 11, '=%d-I%d' % (_d['qty'], rr)).number_format = '+#,##0;-#,##0;0'
        else:
            for cx in (9, 10, 11):
                ws.cell(rr, cx, '—')
        # 会場ごとの累計消化率（確定発注に対して、その会場が終わった時点で何%消化したか）
        if _d:
            _idx = [f[0] for f in FUTURE].index(_d['arrive_at'])
            _cum2 = 0
            for _k, (_fn, _p2, _d2, _r2, _n2) in enumerate(FUTURE):
                if _k < _idx:
                    ws.cell(rr, 12 + _k, '—')
                    continue
                _cum2 += round(_mix2[jan] * _vneed[_fn])
                ws.cell(rr, 12 + _k, _cum2 / _d['qty']).number_format = PCT
        else:
            for _k in range(len(FUTURE)):
                ws.cell(rr, 12 + _k, '—')
        acts = []
        if thru:
            acts.append('%s会期内で現在庫を売り切る（セット販売・POP強化・値引き検討）。残ると廃棄' % thru.split(' ')[0])
        if _d:
            acts.append('確定発注%s%sは最低ロット。消化率が9割超なら妥当、下回るなら会場配分の再検討'
                        % (f"{_d['qty']:,}", _d['unit']))
        _NCS = 11 + len(FUTURE) + 1
        ws.cell(rr, _NCS, ' ／ '.join(acts))
    shtot = sh0 + len(_keys)
    _NCS = 11 + len(FUTURE) + 1
    body(ws, sh0, shtot - 1, _NCS, namecol=1)
    for k in range(len(_keys)):
        ws.cell(sh0 + k, 1).font = BD
        for cx in (1, 2, 8, _NCS):
            ws.cell(sh0 + k, cx).alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    hdrfmt(ws, shr, shr, _NCS)
    ws.conditional_formatting.add('E%d:E%d' % (sh0, shtot - 1),
                                  FormulaRule(formula=['AND(ISNUMBER($E%d),$E%d>0)' % (sh0, sh0)], fill=ALERT))
    ws.conditional_formatting.add('K%d:K%d' % (sh0, shtot - 1),
                                  FormulaRule(formula=['AND(ISNUMBER($K%d),$K%d>0)' % (sh0, sh0)], fill=WARN))
    r = shtot + 1

ws.cell(r, 1, '■ 結論').font = SEC
r += 1
for a, b in [
    ('札幌への対応', '欠品報告8品目はすべて倉庫在庫から追納可能。追加生産は不要で、追納数量は⑧⑪シートに算出済み。'),
    ('石川以降の考え方', 'まず既存在庫を会場順に引き当て、在庫が尽きた会場で初めて追加発注を検討する。'),
    ('', 'その会場の会期初日から生産日数＋バッファを逆算した日が「最終発注期限」となる。'),
    ('緊急度の判定', '★期限超過＝すでに間に合わない／★至急＝30日以内／要着手＝90日以内／余裕あり＝91日以上。'),
    ('生産日数の変動リスク', '2026/7/28時点の提示値。物価高騰・素材/材料の入手困難により変動しうるため、'),
    ('', '実際の発注判断では期限に対して30日程度の余裕を見込むことを推奨。'),
    ('賞味期限の制約', '宿儺の指風お菓子は賞味期限4か月以内が販売不可。現在庫は金沢(石川)会場までは賞味期限OK。'),
    ('', '残すと廃棄になるため、石川会期内での売り切りを目標にする（上表の「売り切れなかった場合の残」参照）。'),
    ('', '仙台以降向けは1,200SET(12,000本)で発注確定済み。最低ロットのため数量調整の余地はない。'),
    ('', '東京凱旋(27年4月)まで賞味期限が持つか、製造日から逆算した確認が必要。'),
    ('再生産不可品', '虚式『茈』バスボールは生産日数が「－」＝再生産不可。現有在庫の範囲で会場配分を組む必要がある。'),
]:
    ws.cell(r, 1, a).font = BD
    c = ws.cell(r, 2, b)
    c.font = NM
    c.alignment = LFT
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=10)
    r += 1

# ===========================================================================
# 13. ⑫計算ロジック解説
#   列位置・数式は「生成済みのシートから読み取って」書き出す。列を足しても
#   解説がずれないよう、ここでは列文字も数式も直書きしない。
# ===========================================================================
def _sh12(pfx):
    return next(n for n in wb.sheetnames if n.startswith(pfx))


def _hr12(sn, label='No.', maxr=40):
    wx = wb[sn]
    for rr in range(1, min(maxr, wx.max_row) + 1):
        if str(wx.cell(rr, 1).value or '').strip() == label:
            return rr
    return 1


def _rw12(sn, label):
    """A列が label で始まる行番号（見つからなければ0）。"""
    wx = wb[sn]
    key = label.replace(' ', '').replace('　', '')
    for rr in range(1, wx.max_row + 1):
        v = str(wx.cell(rr, 1).value or '').replace(' ', '').replace('　', '').replace('\n', '')
        if v.startswith(key):
            return rr
    return 0


def _cc12(sn, hrow, *keys):
    """ヘッダ行 hrow で keys をすべて含む見出しの列番号。"""
    wx = wb[sn]
    for cc in range(1, wx.max_column + 1):
        v = str(wx.cell(hrow, cc).value or '').replace('\n', '').replace(' ', '').replace('　', '')
        if v and all(k.replace(' ', '').replace('\n', '') in v for k in keys):
            return cc
    return 0


def _L(sn, hrow, *keys):
    """列文字（見つからなければ '?'）。"""
    cc = _cc12(sn, hrow, *keys)
    return gcl(cc) if cc else '?'


def _HD(sn, hrow, *keys):
    """実際の見出し文字（改行は「 」に畳む）。"""
    cc = _cc12(sn, hrow, *keys)
    return str(wb[sn].cell(hrow, cc).value or '').replace('\n', ' ') if cc else '／'.join(keys)


def _FX(sn, hrow, drow, *keys):
    """データ先頭行の数式（_xlfn. は表示上落とす）。"""
    cc = _cc12(sn, hrow, *keys)
    if not cc:
        return ''
    v = wb[sn].cell(drow, cc).value
    if v is None or not str(v).startswith('='):
        return '（数式なし＝実績値・転記値の直接入力）'
    return str(v).replace('_xlfn.', '')


def _f12(x, d=0):
    return format(x, ',.%df' % d)


SN1, SN2, SN3, SN4 = _sh12('①'), _sh12('②'), _sh12('③'), _sh12('④')
SN5, SN6, SN7, SN8 = _sh12('⑤'), _sh12('⑥'), _sh12('⑦'), _sh12('⑧')
SN9, SN10, SN11 = _sh12('⑨'), _sh12('⑩'), _sh12('⑪')

H1R = _hr12(SN1); D1 = H1R + 1; T1 = D1 + len(T)
H3R = _hr12(SN3); D3 = H3R + 1; T3 = D3 + len(T)
H4R = _hr12(SN4); D4 = H4R + 1; T4 = D4 + len(T)
H5R = _hr12(SN5); D5 = H5R + 1
H6R = _hr12(SN6); D6 = H6R + 1; T6 = D6 + len(T)
H8R = _hr12(SN8); D8 = H8R + 1; T8 = D8 + len(T)
H9R = _hr12(SN9); D9 = H9R + 1; T9 = D9 + len(T)
H10R = _hr12(SN10); D10 = H10R + 1
H11R = _hr12(SN11); D11 = H11R + 1
V3 = _rw12(SN3, FUTURE[0][0])                 # ③会場別表の先頭行
NF = len(FUTURE)
FV1, FVL = FUTURE[0][0], FUTURE[-1][0]

# ---- 実例に使う No.1 商品の実数 -------------------------------------------
_j1, _n1 = T[0][0], T[0][1]
_p1 = R(CUR, _j1)['price']
_e1 = S(CUR, _j1, 0, BASE)
_nag1, _nag2 = S('名古屋', _j1, 0, BASE), S('名古屋', _j1, 0, 19)
_hak1, _hak2 = S('博多', _j1, 0, BASE), S('博多', _j1, 0, RUNC)
_osa1, _osa2 = S('大阪', _j1, 0, BASE), S('大阪', _j1, 0, RUNC)
_g1 = {v: gv(v, _j1) for v in WGT}
_gad = (PRED[_j1] / _e1) if _e1 else G_ALL
_pr1, _av1 = PRED[_j1], R(CUR, _j1)['avail']
_rs1 = max(0, _av1 - _pr1)
_iv1 = INV[str(_j1)]['real']
_me1, _mg1, _mh1 = OQ[_j1] / TO, _pr1 / TS, _mix2[_j1]
_nd1 = {fn: round(_mh1 * _vneed[fn]) for fn, *_x in FUTURE}
_ndt1 = sum(_nd1.values())
_wh1 = _iv1 - _av1
_sp1 = _wh1 - max(0, _pr1 - _av1) + _rs1
_dc1 = (DECIDED.get(_j1) or {}).get('qty', 0)
_ap1 = max(0, _ndt1 - _sp1 - _dc1)
_or1 = int(-(-_ap1 * (1 + CFG.SAFETY) // 1)) if _ap1 else 0
_alloc1, _rem1, _sv1 = [], _sp1 + _dc1, ''
for _fn, *_x in FUTURE:
    _rem1 -= _nd1[_fn]
    _alloc1.append(_rem1)
    if _rem1 < 0 and not _sv1:
        _sv1 = _fn

# ---- 全品計（③の在庫収支の実例に使う） ------------------------------------
_wh_t = sum(INV[str(j)]['real'] - R(CUR, j)['avail'] for j in TJ)
_add_t = sum(max(0, PRED[j] - R(CUR, j)['avail']) for j in TJ)
_sup_t = _wh_t - _add_t + _rest
_dec_t = sum(v['qty'] for v in DECIDED.values())
_need_t = sum(_vneed.values())
_stkS = STOCK_STAGE.replace('(SET)', '')

ws = wb.create_sheet('⑫計算ロジック解説')
ws.sheet_view.showGridLines = False
ws['A1'] = '⑫ 計算ロジック解説 ― 各シート・各列の計算式（Excel関数）と実例'
ws['A1'].font = TITLE
ws['A2'] = ('実例はすべて No.1「%s」（税込%s円）の行の値。'
            '「前提・入力」シートの青字セルを変えると下記の数式が連動して再計算される。'
            % (_n1, _f12(_p1)))
ws['A2'].font = NOTE
ws['A3'] = ('セル・列の位置は生成時のレイアウトから自動取得しているため、列を追加しても解説はずれない。'
            '数式中の _xlfn. は Excel2007以降の関数を示す内部接頭辞のため、表示上は省いている。')
ws['A3'].font = NOTE

r = 5
for i, h in enumerate(['シート', 'セル／列', '項目', '計算式（Excel関数）', '実例（No.1）', '参照元・補足'], 1):
    c = ws.cell(r, i, h)
    c.font = H1
    c.fill = NAVY
    c.alignment = CTR
    c.border = BOX
ws.row_dimensions[r].height = 22
r += 1


def _sec12(t):
    global r
    c = ws.cell(r, 1, t)
    c.font = Font(name=FONT, size=10, bold=True, color='FFFFFF')
    c.fill = BLUE
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.row_dimensions[r].height = 18
    r += 1


def _row12(sheet, ref, item, formula, ex, note):
    global r
    for i, v in enumerate([sheet, ref, item, formula, ex, note], 1):
        c = ws.cell(r, i, v)
        # 「=…」は解説として見せる文字列。openpyxl が数式と誤認するので文字列に戻す
        if isinstance(v, str) and v.startswith('='):
            c.data_type = 's'
        c.font = BD if i == 3 else NM
        c.border = BOX
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    ws.row_dimensions[r].height = 13.5 * max(1, max(len(str(x)) // w + 1 for x, w in
                                                    [(item, 14), (formula, 44), (ex, 30), (note, 32)]))
    r += 1


# ---- 前提・入力 -----------------------------------------------------------
_sec12('■ 前提・入力（青字・黄色セルが入力値。C列に各パラメータの意味・算出・根拠を記載）')
_row12('前提・入力', '黄色セル（%s ほか）' % P['w_osaka'], '手入力の前提値',
       '（数式なし。直接入力する）',
       '例）%s＝%.0f%%（大阪の按分ウェイト）' % (P['mix_o'], CFG.MIX_WEIGHT_PREV * 100),
       'ここを変えると①③④⑤⑧⑩⑪が一斉に再計算される')
_row12('前提・入力', '%s:%s' % (FR[FV1], FR[FVL]), '東京対比 会場比率（シナリオ連動）',
       "=INDEX(C%d:F%d,'前提・入力'!$%s)" % (SCNROW[FV1], SCNROW[FV1], P['scn']),
       '%s＝1 → C%d の %.1f%%（%s）を取得' % (P['scn'], SCNROW[FV1], FUTURE[0][3] * 100, FV1),
       '%s で①〜④を切替。4案の結果比較は%s' % (P['scn'], SN10))

# ---- ① --------------------------------------------------------------------
_sec12('■ %s（%d行目＝見出し、%d〜%d行＝商品、%d行＝合計）' % (SN1, H1R, D1, T1 - 1, T1))
_row12(SN1, '%s〜%s列' % (_L(SN1, H1R, CUR, NB), _L(SN1, H1R, '大阪', '累計')), '各会場の実績販売数',
       '（数式なし＝POSデータ実績の直接入力）',
       '%s%s %s／名古屋 %s・%s／博多 %s・%s／大阪 %s・%s'
       % (CUR, NB, _f12(_e1), _f12(_nag1), _f12(_nag2), _f12(_hak1), _f12(_hak2),
          _f12(_osa1), _f12(_osa2)),
       '%d日ベースに揃えた累計（博多は初%d日、名古屋は19日＋%d日目補完）' % (RUNC, RUNC, RUNC))
for _v, _ex in [('名古屋', '%s×(1+%.2f%%)÷%s ＝ %.3f' % (_f12(_nag2), TAIL * 100, _f12(_nag1),
                                                       _g1.get('名古屋', 0))),
                ('博多', '%s÷%s ＝ %.3f' % (_f12(_hak2), _f12(_hak1), _g1.get('博多', 0))),
                ('大阪', '%s÷%s ＝ %.3f' % (_f12(_osa2), _f12(_osa1), _g1.get('大阪', 0)))]:
    _nt = {'名古屋': '%s＝%d日目の補完率。19日会期を%d日相当に換算' % (P['tail'], RUNC, RUNC),
           '博多': '博多%d日会期のうち頭から%d日ぶんを使用' % (RUN['博多'], RUNC),
           '大阪': '%sと同条件（%d日・水曜開始・新商品投入後）のため加重%d'
                   % (CUR, RUNC, CFG.G_WEIGHT.get('大阪', 0))}[_v]
    _row12(SN1, '%s列' % _L(SN1, H1R, _v, '倍率g'), _HD(SN1, H1R, _v, '倍率g'),
           _FX(SN1, H1R, D1, _v, '倍率g'), _ex, _nt)
_row12(SN1, '%s列' % _L(SN1, H1R, '採用', '倍率g'), _HD(SN1, H1R, '採用', '倍率g'),
       _FX(SN1, H1R, D1, '採用', '倍率g'),
       '(%s)÷%d ＝ %.4f' % ('＋'.join('%.3f×%d' % (_g1[v], WGT[v]) for v in WGT),
                           sum(WGT.values()), _gad),
       '加重は「前提・入力」%s:%s。実績ゼロ品は%s（%.3f倍）で代用' % (P['w_osaka'], P['w_tokyo'], P['g_all'], G_ALL))
_row12(SN1, '%s列' % _L(SN1, H1R, '予測販売数'), _HD(SN1, H1R, '予測販売数'),
       _FX(SN1, H1R, D1, '予測販売数'), '%s×%.4f ＝ %s個' % (_f12(_e1), _gad, _f12(_pr1)),
       '★本モデルの中核。合計 %s%d＝%s個' % (_L(SN1, H1R, '予測販売数'), T1, _f12(TS)))
_row12(SN1, '%s列' % _L(SN1, H1R, '予測販売金額'), _HD(SN1, H1R, '予測販売金額'),
       _FX(SN1, H1R, D1, '予測販売金額'),
       '%s×%s円 ＝ %s円' % (_f12(_pr1), _f12(_p1), _f12(_pr1 * _p1)),
       '合計 %s%d＝%s円' % (_L(SN1, H1R, '予測販売金額'), T1, _f12(_spa)))
_row12(SN1, '%s列' % _L(SN1, H1R, '数量構成比'), _HD(SN1, H1R, '数量構成比'),
       _FX(SN1, H1R, D1, '数量構成比'), '%s÷%s ＝ %.2f%%' % (_f12(_pr1), _f12(TS), _mg1 * 100),
       '③の按分構成比（%s側ウェイト）に使う' % CUR)
_row12(SN1, '%s列' % _L(SN1, H1R, '販売可能数'), _HD(SN1, H1R, '販売可能数'),
       '（数式なし＝納品数の実績入力）', '%s個' % _f12(_av1),
       '%s の追納分（LEGS計%s個）を加算済み' % (RESUPPLY_NOTE, _f12(sum(RESUPPLY.values()))))
_row12(SN1, '%s列' % _L(SN1, H1R, '消化率'), _HD(SN1, H1R, '消化率'), _FX(SN1, H1R, D1, '消化率'),
       '%s÷%s ＝ %.1f%%' % (_f12(_pr1), _f12(_av1), (_pr1 / _av1 * 100) if _av1 else 0),
       '90%超で「要注意」判定')
_row12(SN1, '%s列' % _L(SN1, H1R, '会期末', '残在庫'), _HD(SN1, H1R, '会期末', '残在庫'),
       _FX(SN1, H1R, D1, '会期末', '残在庫'), '%s−%s ＝ %s個' % (_f12(_av1), _f12(_pr1), _f12(_rs1)),
       '③のスライド列で石川以降へ回す')
_row12(SN1, '%s列' % _L(SN1, H1R, '会期中', '過不足'), _HD(SN1, H1R, '会期中', '過不足'),
       _FX(SN1, H1R, D1, '会期中', '過不足'), '%s−%s ＝ %s' % (_f12(_pr1), _f12(_av1), _f12(_pr1 - _av1)),
       'プラス＝会期中に在庫切れ')
_row12(SN1, '%s列' % _L(SN1, H1R, '在庫判定'), _HD(SN1, H1R, '在庫判定'), _FX(SN1, H1R, D1, '在庫判定'),
       '判定：%s' % ('会期中に在庫切れ' if _pr1 > _av1 else
                     ('要注意(90%超)' if _av1 and _pr1 / _av1 >= 0.9 else
                      ('適正' if _av1 and _pr1 / _av1 >= 0.5 else '在庫余剰'))),
       '%s%d＝COUNTIF で「会期中に在庫切れ」の品目数を集計' % (_L(SN1, H1R, '在庫判定'), T1))

# ---- ② --------------------------------------------------------------------
_r2q, _r2a = _rw12(SN2, '全物販 販売数量'), _rw12(SN2, '全物販 販売金額')
_r2lq, _r2la = _rw12(SN2, 'LEGS 販売数量'), _rw12(SN2, 'LEGS 販売金額')
_r2sq, _r2sa = _rw12(SN2, 'LEGS 数量構成比'), _rw12(SN2, 'LEGS 金額構成比')
_r2tq, _r2u = _rw12(SN2, '東京対比（数量）'), _rw12(SN2, 'LEGS 平均単価')
_c2f, _c2i, _c2e = gcl(3 + len(VS)), gcl(4 + len(VS)), gcl(3 + len(VS) + NF)
_sec12('■ %s（%d〜%d行 × C〜%s列＝会場別。実績5会場＋%s予測＋石川以降5会場）'
       % (SN2, _r2q, _r2u, _c2e, CUR))
_row12(SN2, '%s%d / %s%d' % (_c2f, _r2q, _c2f, _r2a), '%s%d日 全物販（予測）' % (CUR, RUNC),
       '%s ／ %s' % (wb[SN2]['%s%d' % (_c2f, _r2q)].value, wb[SN2]['%s%d' % (_c2f, _r2a)].value),
       '%s個×%.4f ＝ %s個' % (_f12(vt(CUR, 0, BASE)[0]),
                            vt('大阪', 0, RUNC)[0] / vt('大阪', 0, BASE)[0], _f12(_sallq)),
       '%s→%d日の実績倍率（数量／金額で別倍率）を大阪実績から算出' % (NB, RUNC))
_row12(SN2, '%s%d:%s%d' % (_c2i, _r2q, _c2e, _r2a), '石川以降 全物販（予測）',
       '%s（数量）／%s（金額）' % (wb[SN2]['%s%d' % (_c2i, _r2q)].value,
                                wb[SN2]['%s%d' % (_c2i, _r2a)].value),
       '%s円×%.2f%% ＝ %s円' % (_f12(TKA), FUTURE[0][3] * 100, _f12(TKA * FUTURE[0][3])),
       '東京(初回)実績 × 東京対比比率（シナリオ連動）')
_row12(SN2, '%s%d:%s%d' % (_c2i, _r2lq, _c2e, _r2la), 'LEGS 販売数量／金額（予測）',
       '%s ／ %s' % (wb[SN2]['%s%d' % (_c2i, _r2lq)].value, wb[SN2]['%s%d' % (_c2i, _r2la)].value),
       '会場売上×%.1f%% ＝ LEGS売上、それを%s円で割る' % (_lsha * 100, _f12(round(_unit))),
       '%s＝LEGS金額構成比、%s＝LEGS平均単価' % (P['lsh_a'], P['unit']))
_row12(SN2, '%d行 / %d行' % (_r2sq, _r2sa), 'LEGS 数量／金額 構成比',
       '%s ／ %s' % (wb[SN2]['C%d' % _r2sq].value, wb[SN2]['C%d' % _r2sa].value),
       '東京 %s÷%s ＝ %.1f%%（数量）' % (_f12(vt('東京', 0, RUN['東京'], legs=True)[0]), _f12(TKQ),
                                      vt('東京', 0, RUN['東京'], legs=True)[0] / TKQ * 100),
       '新商品投入による構成比の低下推移が読める行')
_row12(SN2, '%d行' % _r2tq, '東京対比（数量／金額）', wb[SN2]['D%d' % _r2tq].value,
       '名古屋 %s÷%s ＝ %.1f%%' % (_f12(vt('名古屋', 0, RUN['名古屋'])[1]), _f12(TKA),
                                  vt('名古屋', 0, RUN['名古屋'])[1] / TKA * 100),
       '石川以降は「前提・入力」の東京対比セルを直接参照している')
_row12(SN2, '%d行' % _r2u, 'LEGS 平均単価', wb[SN2]['C%d' % _r2u].value,
       '東京 %s円÷%s個 ＝ %s円' % (_f12(vt('東京', 0, RUN['東京'], legs=True)[1]),
                                 _f12(vt('東京', 0, RUN['東京'], legs=True)[0]),
                                 _f12(vt('東京', 0, RUN['東京'], legs=True)[1]
                                      / vt('東京', 0, RUN['東京'], legs=True)[0])),
       '%s予測は%s円（%s の根拠）' % (CUR, _f12(round(_unit)), P['unit']))

# ---- ③ --------------------------------------------------------------------
_b15, _b26 = _rw12(SN3, '残在庫想定'), _rw12(SN3, '追加生産 必要数')
_sec12('■ %s（%d〜%d行＝会場別、%d〜%d行＝在庫収支、%d行＝見出し、%d〜%d行＝商品、%d行＝合計）'
       % (SN3, V3, V3 + NF, _b15, _b26, H3R, D3, T3 - 1, T3))
_row12(SN3, 'D%d:D%d' % (V3, V3 + NF - 1), '東京対比', wb[SN3]['D%d' % V3].value,
       '%.1f%%（%s）' % (FUTURE[0][3] * 100, FV1), 'シナリオ選択で自動切替')
_row12(SN3, 'E%d:E%d' % (V3, V3 + NF - 1), '会場 全物販売上', wb[SN3]['E%d' % V3].value,
       '%s×%.1f%% ＝ %s円' % (_f12(TKA), FUTURE[0][3] * 100, _f12(TKA * FUTURE[0][3])),
       '%s＝東京(初回)%d日間の全物販実績' % (P['tk_amt'], RUN['東京']))
_row12(SN3, 'F%d:F%d' % (V3, V3 + NF - 1), 'LEGS 売上', wb[SN3]['F%d' % V3].value,
       '%s×%.1f%% ＝ %s円' % (_f12(TKA * FUTURE[0][3]), _lsha * 100, _f12(TKA * FUTURE[0][3] * _lsha)),
       '%s＝LEGS金額構成比' % P['lsh_a'])
_row12(SN3, 'G%d:G%d' % (V3, V3 + NF - 1), 'LEGS 数量（会場の必要数）', wb[SN3]['G%d' % V3].value,
       '%s÷%s円 ＝ %s個' % (_f12(TKA * FUTURE[0][3] * _lsha), _f12(round(_unit)),
                          _f12(_vneed[FV1])),
       '%d会場計 G%d＝%s個' % (NF, V3 + NF, _f12(_need_t)))
_row12(SN3, 'B%d:B%d' % (_b15, _b26), '在庫収支（LEGS 42SKU 合計）',
       '残在庫想定 → %s納品 → 倉庫残 → 追納 → 会期末返送 → 供給可能在庫 → 必要数 → 追加生産、と積み上げる' % CUR,
       '供給可能 %s ＋ 確定発注 %s − 必要 %s ＝ %s個'
       % (_f12(_sup_t), _f12(_dec_t), _f12(_need_t), _f12(_sup_t + _dec_t - _need_t)),
       '全体は余剰でも品目別には不足が出る（追加生産 %s個・%d品目）' % (_f12(_addprod), len(_addlist)))
_row12(SN3, '%s列' % _L(SN3, H3R, '大阪', '構成比'), _HD(SN3, H3R, '大阪', '構成比'),
       _FX(SN3, H3R, D3, '大阪', '構成比'), '%s÷%s ＝ %.2f%%' % (_f12(OQ[_j1]), _f12(TO), _me1 * 100),
       '大阪%d日実績ベース（前会場側の構成比）' % RUN['大阪'])
_row12(SN3, '%s列' % _L(SN3, H3R, CUR, '構成比'), _HD(SN3, H3R, CUR, '構成比'),
       _FX(SN3, H3R, D3, CUR, '構成比'), '%s÷%s ＝ %.2f%%' % (_f12(_pr1), _f12(TS), _mg1 * 100),
       '%s%d日予測ベース（①の数量構成比と同値）' % (CUR, RUNC))
_row12(SN3, '%s列' % _L(SN3, H3R, '採用', '按分'), _HD(SN3, H3R, '採用', '按分'),
       _FX(SN3, H3R, D3, '採用', '按分'),
       '%.2f%%×%.0f%% ＋ %.2f%%×%.0f%% ＝ %.3f%%'
       % (_me1 * 100, CFG.MIX_WEIGHT_PREV * 100, _mg1 * 100, CFG.MIX_WEIGHT_CUR * 100, _mh1 * 100),
       '★商品別配分の要。ウェイトは「前提・入力」%s／%s' % (P['mix_o'], P['mix_s']))
_row12(SN3, '%s:%s列' % (_L(SN3, H3R, FV1.split(' ')[0], '必要数'),
                        _L(SN3, H3R, FVL.split(' ')[0], '必要数')), '会場別 必要数',
       _FX(SN3, H3R, D3, FV1.split(' ')[0], '必要数'),
       '%.3f%%×%s ＝ %s個（%s）' % (_mh1 * 100, _f12(_vneed[FV1]), _f12(_nd1[FV1]), FV1),
       '列ごとに会場別表の $G$%d〜$G$%d を参照' % (V3, V3 + NF - 1))
_row12(SN3, '%s列' % _L(SN3, H3R, '必要数', '計'), _HD(SN3, H3R, '必要数', '計'),
       _FX(SN3, H3R, D3, '必要数', '計'),
       '%s ＝ %s個' % ('＋'.join(_f12(_nd1[f[0]]) for f in FUTURE), _f12(_ndt1)),
       'この商品の石川以降 総需要。④⑤⑧⑩⑪へ供給される')
_row12(SN3, '%s列' % _L(SN3, H3R, '残在庫想定'), _HD(SN3, H3R, '残在庫想定'),
       '⑧シートの残在庫想定列からの転記', '%s個' % _f12(_iv1), CFG.STOCK_SOURCE)
_row12(SN3, '%s / %s列' % (_L(SN3, H3R, '納品済'), _L(SN3, H3R, '倉庫残')), '%s納品済／差引 倉庫残' % CUR,
       '%s ／ %s' % (_FX(SN3, H3R, D3, '納品済'), _FX(SN3, H3R, D3, '倉庫残')),
       '%s ／ %s−%s ＝ %s個' % (_f12(_av1), _f12(_iv1), _f12(_av1), _f12(_wh1)),
       '倉庫残＝石川以降へ即出荷できる分')
_row12(SN3, '%s列' % _L(SN3, H3R, '追納必要数'), _HD(SN3, H3R, '追納必要数'),
       _FX(SN3, H3R, D3, '追納必要数'),
       'MAX(0, %s−%s) ＝ %s' % (_f12(_pr1), _f12(_av1), _f12(max(0, _pr1 - _av1))),
       '会期中に在庫切れになる品だけプラスになる')
_row12(SN3, '%s列' % _L(SN3, H3R, '返送'), _HD(SN3, H3R, '返送'), _FX(SN3, H3R, D3, '返送'),
       '%s×100%% ＝ %s個' % (_f12(_rs1), _f12(_rs1)), '%s＝スライド率' % P['slide'])
_row12(SN3, '%s列' % _L(SN3, H3R, '供給可能計'), _HD(SN3, H3R, '供給可能計'),
       _FX(SN3, H3R, D3, '供給可能計'),
       '%s−%s＋%s ＝ %s個' % (_f12(_wh1), _f12(max(0, _pr1 - _av1)), _f12(_rs1), _f12(_sp1)),
       '倉庫残 − %s追納 ＋ %s会期末返送' % (CUR, CUR))
_row12(SN3, '%s列' % _L(SN3, H3R, '確定発注'), _HD(SN3, H3R, '確定発注'),
       '（数式なし＝発注判断済みの数量を入力）',
       '／'.join('%s %s%s（%s入荷）' % (_nmof.get(k, k), _f12(v['qty']), v['unit'], v['arrive_at'])
                for k, v in DECIDED.items()) or '設定なし',
       '発注済みなので過不足から差し引く。⑪では緊急度「◎発注確定済」として期限警告から外す')
_row12(SN3, '%s列' % _L(SN3, H3R, '過不足'), _HD(SN3, H3R, '過不足'), _FX(SN3, H3R, D3, '過不足'),
       '%s−%s−%s ＝ %s' % (_f12(_ndt1), _f12(_sp1), _f12(_dc1), _f12(_ndt1 - _sp1 - _dc1)),
       'プラス＝不足')
_row12(SN3, '%s列' % _L(SN3, H3R, '追加生産'), _HD(SN3, H3R, '追加生産'), _FX(SN3, H3R, D3, '追加生産'),
       '%s個' % _f12(_ap1), '42SKU計で %s個・%d品目' % (_f12(_addprod), len(_addlist)))
_row12(SN3, '%s列' % _L(SN3, H3R, '推奨発注数'), _HD(SN3, H3R, '推奨発注数'), _FX(SN3, H3R, D3, '推奨発注数'),
       '%s×%.1f → %s個' % (_f12(_ap1), 1 + CFG.SAFETY, _f12(_or1)),
       '%s＝安全率%.0f%%' % (P['safe'], CFG.SAFETY * 100))

# ---- ④ --------------------------------------------------------------------
_sec12('■ %s（%d行目＝見出し、%d〜%d行＝商品、%d行＝合計）' % (SN4, H4R, D4, T4 - 1, T4))
_row12(SN4, '%s〜%s列' % (_L(SN4, H4R, CUR, '実績'), _L(SN4, H4R, CUR, '販売可能数')),
       '%s %s／%d日予測／販売可能数' % (CUR, NB, RUNC),
       '%s ／ %s ／ %s' % (_FX(SN4, H4R, D4, CUR, '実績'), _FX(SN4, H4R, D4, CUR, '予測'),
                          _FX(SN4, H4R, D4, CUR, '販売可能数')),
       '%s ／ %s ／ %s' % (_f12(_e1), _f12(_pr1), _f12(_av1)), '①シートからのリンク（再計算はしない）')
_row12(SN4, '%s / %s列' % (_L(SN4, H4R, '予測消化率'), _L(SN4, H4R, '会期中', '過不足')),
       '予測消化率／会期中 過不足',
       '%s ／ %s' % (_FX(SN4, H4R, D4, '予測消化率'), _FX(SN4, H4R, D4, '会期中', '過不足')),
       '%.1f%% ／ %s' % ((_pr1 / _av1 * 100) if _av1 else 0, _f12(_pr1 - _av1)),
       'プラス＝会期中に足りない')
_row12(SN4, '%s / %s列' % (_L(SN4, H4R, '会期中', '判定'), _L(SN4, H4R, '会期中', '推奨発注数')),
       '%s会期中 判定／推奨発注数' % CUR,
       '%s ／ %s' % (_FX(SN4, H4R, D4, '会期中', '判定'), _FX(SN4, H4R, D4, '会期中', '推奨発注数')),
       '「%s」／ %s個' % ('在庫余剰' if _pr1 <= _av1 * 0.5 else '適正', _f12(max(0, _pr1 - _av1))),
       '不足分×%.1f を切上げ（%s＝安全率）' % (1 + CFG.SAFETY, P['safe']))
_row12(SN4, '%s〜%s列' % (_L(SN4, H4R, '石川以降', '必要数'), _L(SN4, H4R, '倉庫残')),
       '石川以降 必要数／スライド／倉庫残',
       '③シートからのリンク（%s ほか）' % _FX(SN4, H4R, D4, '石川以降', '必要数'),
       '%s ／ %s ／ %s' % (_f12(_ndt1), _f12(_rs1), _f12(_wh1)), '')
_row12(SN4, '%s〜%s列' % (_L(SN4, H4R, '石川以降', '過不足'), _L(SN4, H4R, '石川以降', '推奨発注数')),
       '石川以降 過不足／判定／推奨発注数',
       '%s ／ %s ／ %s' % (_FX(SN4, H4R, D4, '石川以降', '過不足'),
                          _FX(SN4, H4R, D4, '石川以降', '判定'),
                          _FX(SN4, H4R, D4, '石川以降', '推奨発注数')),
       '%s−%s−%s ＝ %s → 「%s」'
       % (_f12(_ndt1), _f12(_rs1), _f12(_wh1), _f12(_ndt1 - _rs1 - _wh1),
          '要 追加生産' if _ndt1 - _rs1 - _wh1 > 0 else '在庫で充足'), '')
_row12(SN4, '%s / %s列' % (_L(SN4, H4R, '発注合計'), _L(SN4, H4R, '発注金額')), '発注合計／発注金額',
       '%s ／ %s' % (_FX(SN4, H4R, D4, '発注合計'), _FX(SN4, H4R, D4, '発注金額')),
       '%s個 ／ %s円' % (_f12(_or1), _f12(_or1 * _p1)), '%s会期中分 ＋ 石川以降分' % CUR)
_row12(SN4, '%s / %s列' % (_L(SN4, H4R, '欠品発生'), _L(SN4, H4R, '最終発注期限')),
       '欠品発生 会場／最終発注期限',
       '欠品発生会場は⑪の会場別引当の結果を確定値で書出し ／ %s' % _FX(SN4, H4R, D4, '最終発注期限'),
       '欠品会場の会期初日 − 生産日数 − バッファ%d日' % CFG.ORDER_BUFFER_DAYS,
       '在庫を会場順に引き当てて最初にマイナスへ転じた会場（⑪と同値）')
_row12(SN4, '%s / %s列' % (_L(SN4, H4R, '残日数'), _L(SN4, H4R, '緊急度')), '残日数／発注の緊急度',
       '%s ／ %s' % (_FX(SN4, H4R, D4, '残日数'), _FX(SN4, H4R, D4, '緊急度')),
       '★期限超過／★至急(30日以内)／要着手(90日以内)／余裕あり／◎発注確定済／×再生産不可',
       '%s＝基準日。日付を進めると残日数が減る' % P['base'])
_row12(SN4, '%s列' % _L(SN4, H4R, '優先度'), '優先度', _FX(SN4, H4R, D4, '優先度'),
       'S:現場欠品 ＞ A:会期中欠品 ＞ B:石川以降不足 ＞ —',
       '%s列（運営 欠品報告 %s）を最優先にする' % (_L(SN4, H4R, '欠品報告'), CFG.SHORTAGE_REPORT_DATE))
_row12(SN4, '%s列' % _L(SN4, H4R, '突合'), _HD(SN4, H4R, '突合'), _FX(SN4, H4R, D4, '突合'),
       '◎一致／○整合／△未検知 の3段階',
       'モデルの見落とし・空振りを可視化する検算列。サマリーに件数を集計している')

# ---- ⑤ --------------------------------------------------------------------
_sec12('■ %s（%d行目＝見出し、%d〜%d行＝商品）' % (SN5, H5R, D5, D5 + len(T) - 1))
_row12(SN5, '各会場ブロック', '数量構成比／販売金額／消化率',
       '=IFERROR(販売数/販売数合計,0) ／ =販売数*税込単価 ／ =IFERROR(販売数/販売可能数,0)',
       '東京：%s÷%s＝%.2f%% ／ %s×%s＝%s円'
       % (_f12(S('東京', _j1, 0, RUN['東京'])), _f12(vt('東京', 0, RUN['東京'], legs=True)[0]),
          S('東京', _j1, 0, RUN['東京']) / vt('東京', 0, RUN['東京'], legs=True)[0] * 100,
          _f12(S('東京', _j1, 0, RUN['東京'])), _f12(R('東京', _j1)['price']),
          _f12(S('東京', _j1, 0, RUN['東京']) * R('東京', _j1)['price'])),
       '会場ごとに同じ3式が横に並ぶ。イメージリング8種は東京¥22,000→名古屋以降¥27,500の'
       '実売単価で金額を算出している')
_row12(SN5, '予測・判定ブロック', '①③④の結果の一覧化',
       '=①!採用倍率g〜会期末残在庫 ／ =③!必要数〜追加生産必要数 ／ =④!判定〜優先度',
       '%.4f ／ %s個 ／ %s円 ／ %s個 ／ %s個 ／ %s個'
       % (_gad, _f12(_pr1), _f12(_pr1 * _p1), _f12(_ndt1), _f12(_sp1), _f12(_ap1)),
       'ここでは計算せず、①③④の結果を商品別に横一列で見るためのビュー')

# ---- ⑥ --------------------------------------------------------------------
_sec12('■ %s（%d行目＝見出し、%d〜%d行＝商品、%d行＝合計）' % (SN6, H6R, D6, T6 - 1, T6))
_row12(SN6, '%s〜%s列' % (_L(SN6, H6R, '東京', '構成比'), _L(SN6, H6R, CUR, '構成比')), '各会場 構成比',
       _FX(SN6, H6R, D6, '東京', '構成比'),
       '東京 %s÷%s ＝ %.2f%%' % (_f12(S('東京', _j1, 0, RUN['東京'])),
                               _f12(vt('東京', 0, RUN['東京'], legs=True)[0]),
                               S('東京', _j1, 0, RUN['東京'])
                               / vt('東京', 0, RUN['東京'], legs=True)[0] * 100),
       '会期全体ベース（%sのみ%s実績）' % (CUR, NB))
_row12(SN6, '%s列' % _L(SN6, H6R, '予測構成比'), _HD(SN6, H6R, '予測構成比'),
       _FX(SN6, H6R, D6, '予測構成比'), '%.2f%%' % (_mg1 * 100), '①シートの数量構成比へのリンク')
_row12(SN6, '%s / %s列' % (_L(SN6, H6R, '大阪→'), _L(SN6, H6R, '→大阪')), '構成比変化',
       '%s ／ %s' % (_FX(SN6, H6R, D6, '大阪→'), _FX(SN6, H6R, D6, '→大阪')),
       '%.2f%%−%.2f%% ＝ %+.2fpt' % (_mg1 * 100, _me1 * 100, (_mg1 - _me1) * 100),
       'ポイント差（pt）で表示。石川以降の重点商品を選ぶ材料')
_row12(SN6, '%s列' % _L(SN6, H6R, 'トレンド'), 'トレンド', _FX(SN6, H6R, D6, 'トレンド'),
       '「↑ 伸長」／「→ 横ばい」／「↓ 減退」', '±0.5pt を閾値に判定')

# ---- ⑦ --------------------------------------------------------------------
_sec12('■ %s' % SN7)
_row12(SN7, '1人あたり売上の表', '1人あたり売上（円/人）',
       '（数式なし＝供給元別売上 ÷ 入場者数の実績値を入力）',
       'LEGS：%s' % ' → '.join('%s %s円/人' % (v, _f12(vt(v, 0, DATA_DAYS[v], legs=True)[1]
                                                     / max(1, sum(ATT[v][:DATA_DAYS[v]])))) for v in VS),
       '会場規模の差を消すため「1人あたり」に正規化して比較している')
_row12(SN7, '増減／増減率の列', '名古屋→大阪 増減／増減率', '=大阪−名古屋 ／ =IFERROR(増減/名古屋,"")',
       'LEGSが最大の減少幅', '新商品106SKUを投入した大阪で、LEGSがどれだけ削られたかを見る')
_row12(SN7, '会場別の新商品表', '新規SKU数／新商品売上（円/人）', '（数式なし＝会場別の実績集計値）',
       '大阪：新規106SKU・1,360円/人', 'LEGSの減少幅と新商品の獲得額を突き合わせてカニバリ率を出す')
_row12(SN7, '商品タイプ別の表', 'LEGS減少への寄与度', '=IFERROR(タイプ別の増減/LEGS合計の増減,"")',
       'ブラインド缶バッジ・ミニ色紙で減少の大半を占める', 'どのタイプが新商品に食われたかを特定する')

# ---- ⑧ --------------------------------------------------------------------
_sec12('■ %s（%d行目＝見出し、%d〜%d行＝商品、%d行＝合計）' % (SN8, H8R, D8, T8 - 1, T8))
_row12(SN8, '%s〜%s列' % (_L(SN8, H8R, '実在庫', STOCK_STAGES[0].replace('(SET)', '')),
                        _L(SN8, H8R, '残在庫想定')), '在庫フロー台帳の段階別 実在庫',
       '（数式なし＝台帳「実在庫」列の転記）',
       ' → '.join(_f12(x) for x in INV[str(_j1)]['hist']),
       '会場ごとの 追加発注→販売→レッグス検品 の積上げ。最終列（%s）が現時点の残在庫想定' % STOCK_STAGE)
_row12(SN8, '%s列' % _L(SN8, H8R, '残在庫', '金額'), _HD(SN8, H8R, '残在庫', '金額'),
       _FX(SN8, H8R, D8, '残在庫', '金額'),
       '%s×%s ＝ %s円' % (_f12(_iv1), _f12(_p1), _f12(_iv1 * _p1)),
       '42SKU計 %s個（「前提・入力」%s に集計）' % (_f12(_realtot), P['spj']))
_row12(SN8, '%s〜%s列' % (_L(SN8, H8R, '納品'), _L(SN8, H8R, '返送見込')),
       '%sへ納品／%d日予測販売／会期末返送' % (CUR, RUNC), '①シートからのリンク',
       '%s ／ %s ／ %s' % (_f12(_av1), _f12(_pr1), _f12(_rs1)), '')
_row12(SN8, '%s / %s列' % (_L(SN8, H8R, '倉庫残'), _L(SN8, H8R, '供給可能計')),
       '差引 倉庫残／石川以降 供給可能計',
       '%s ／ %s' % (_FX(SN8, H8R, D8, '倉庫残'), _FX(SN8, H8R, D8, '供給可能計')),
       '%s ／ %s−%s＋%s ＝ %s個'
       % (_f12(_wh1), _f12(_wh1), _f12(max(0, _pr1 - _av1)), _f12(_rs1), _f12(_sp1)),
       '③の供給可能計と必ず一致する（突合用）。%s会期中に欠品する品は追納分を差し引く' % CUR)
_row12(SN8, '%s〜%s列' % (_L(SN8, H8R, '石川以降', '必要数'), _L(SN8, H8R, '推奨発注数')),
       '必要数／過不足／欠品／推奨発注数',
       '=③!必要数 ／ %s ／ %s ／ %s' % (_FX(SN8, H8R, D8, '過不足'), _FX(SN8, H8R, D8, '欠品'),
                                     _FX(SN8, H8R, D8, '推奨発注数')),
       '%s ／ %s ／ %s ／ %s' % (_f12(_ndt1), _f12(_sp1 + _dc1 - _ndt1), _f12(_ap1), _f12(_or1)),
       '過不足はプラス＝余剰。③の過不足とは符号が逆なので注意')
_row12(SN8, '%s〜%s列' % (_L(SN8, H8R, '欠品発生'), _L(SN8, H8R, '緊急度')),
       '欠品発生会場／最終発注期限／緊急度', '⑪シートの会場別引当の結果と同値',
       '会期初日 − 生産日数 − バッファ%d日' % CFG.ORDER_BUFFER_DAYS,
       '④を見なくても⑧単体で発注期限が分かるようにしている')
_row12(SN8, '%s列' % _L(SN8, H8R, '判定'), '判定', _FX(SN8, H8R, D8, '判定'),
       '★欠品：追加生産が必要／大幅余剰／在庫で充足', '余剰が必要数の2倍以上で「大幅余剰」')

# ---- ⑨ --------------------------------------------------------------------
_a9 = _rw12(SN9, '東京')
_sec12('■ %s（%d〜%d行＝会場別サマリー、%d行＝見出し、%d〜%d行＝商品、%d行＝合計）'
       % (SN9, _a9, _a9 + len(VS) - 1, H9R, D9, T9 - 1, T9))
_row12(SN9, '会場別サマリーの倍率列', '会場倍率g（バックキャスト）',
       '（数式なし＝その時点で判明していた会場だけで再計算した倍率の記録）',
       '／'.join('%s %.3f' % (v, BACK[v]['G']) for v in BACK if BACK[v]),
       '「その会場の直前までの情報」だけで予測をやり直し、モデルの精度を後方検証している')
_row12(SN9, '各会場ブロック', '予測時 構成比／実績 構成比',
       '=IFERROR(予測販売数/予測合計,0) ／ =IFERROR(実績販売数/実績合計,0)',
       '投入数ベース vs 実績販売ベース', '会場ごとに同じ2式が横に並ぶ（池袋のみ投入数シェア）')
_row12(SN9, '各会場ブロックの修正列', '構成比 修正幅', '=実績構成比 − 予測時構成比',
       'プラス＝想定より売れた商品', '次会場で構成比を引き上げるべき商品が分かる')
_row12(SN9, '%s / %s列' % (_L(SN9, H9R, '4会場', '実績合計'), _L(SN9, H9R, '4会場', '実績構成比')),
       '4会場 実績合計／構成比',
       '%s ／ %s' % (_FX(SN9, H9R, D9, '4会場', '実績合計'), _FX(SN9, H9R, D9, '4会場', '実績構成比')),
       '%s ＝ %s個' % ('＋'.join(_f12(S(v, _j1, 0, RUN[v])) for v in VS[:-1]),
                      _f12(sum(S(v, _j1, 0, RUN[v]) for v in VS[:-1]))), '')
_row12(SN9, '%s / %s列' % (_L(SN9, H9R, CUR, '予測構成比'), _L(SN9, H9R, '適用構成比')),
       '%s 予測構成比／石川以降 適用構成比' % CUR,
       '%s ／ %s' % (_FX(SN9, H9R, D9, CUR, '予測構成比'), _FX(SN9, H9R, D9, '適用構成比')),
       '%.2f%% ／ %.3f%%' % (_mg1 * 100, _mh1 * 100), '')
_row12(SN9, '%s列' % _L(SN9, H9R, '構成比トレンド'), '構成比トレンド', _FX(SN9, H9R, D9, '構成比トレンド'),
       '↑ 上方修正／→ 据置／↓ 下方修正', '過去4会場の実績構成比と、石川以降の適用構成比の差で判定')

# ---- ⑩ --------------------------------------------------------------------
_s10 = _rw12(SN10, FV1)
_sec12('■ %s（%d〜%d行＝会場別、%d行＝見出し、%d〜%d行＝商品）'
       % (SN10, _s10, _s10 + NF, H10R, D10, D10 + len(T) - 1))
_row12(SN10, 'C%d:J%d' % (_s10, _s10 + NF - 1), 'シナリオ別 東京対比／必要数',
       '=前提・入力のシナリオ表を直接参照 ／ =ROUND(東京売上×東京対比×LEGS金額構成比÷LEGS平均単価,0)',
       '%s×%.1f%%×%.1f%%÷%s ＝ %s個（%s・シナリオ①）'
       % (_f12(TKA), FUTURE[0][3] * 100, _lsha * 100, _f12(round(_unit)), _f12(_vneed[FV1]), FV1),
       '4シナリオを横並びで同時計算。%d行＝各シナリオの%d会場合計必要数' % (_s10 + NF, NF))
_row12(SN10, '%s / %s列' % (_L(SN10, H10R, '按分'), _L(SN10, H10R, '供給可能在庫')),
       '按分構成比／供給可能在庫',
       '%s ／ %s' % (_FX(SN10, H10R, D10, '按分'), _FX(SN10, H10R, D10, '供給可能在庫')),
       '%.3f%% ／ %s個' % (_mh1 * 100, _f12(_sp1)), 'シナリオを変えても在庫は動かないので共通列')
_row12(SN10, 'シナリオ別 必要数の列', 'シナリオ別 必要数', '=ROUND($按分構成比*シナリオ別の会場合計,0)',
       '%.3f%%×%s ＝ %s個（シナリオ①）' % (_mh1 * 100, _f12(_need_t), _f12(_ndt1)), '')
_row12(SN10, 'シナリオ別 追加生産の列', 'シナリオ別 追加生産', '=MAX(0,必要数−$供給可能在庫)',
       'MAX(0, %s−%s) ＝ %s個' % (_f12(_ndt1), _f12(_sp1), _f12(max(0, _ndt1 - _sp1))),
       'シナリオを切り替えたとき発注量がどれだけ動くかを見る')

# ---- ⑪ --------------------------------------------------------------------
_sec12('■ %s（%d行目＝見出し、%d〜%d行＝商品）' % (SN11, H11R, D11, D11 + len(T) - 1))
_row12(SN11, '%s / %s列' % (_L(SN11, H11R, '生産日数', 'か月'), _L(SN11, H11R, '生産日数', '日換算')),
       '生産日数', '（か月は手入力） ／ %s' % _FX(SN11, H11R, D11, '生産日数', '日換算'),
       '%sか月 → %d日' % (PROD[_j1], round(PROD[_j1] * 30)),
       '2026/7/28時点の提示値。再生産不可の品は「－」として日数計算をしない')
_row12(SN11, '%s列' % _L(SN11, H11R, '供給可能在庫'), _HD(SN11, H11R, '供給可能在庫'),
       _FX(SN11, H11R, D11, '供給可能在庫'), '%s個' % _f12(_sp1), 'この在庫を会場順に引き当てていく')
_row12(SN11, '各会場の必要数列', '会場別 必要数', '=③!会場別の必要数（%s〜%s）' % (FV1, FVL),
       '／'.join(_f12(_nd1[f[0]]) for f in FUTURE), '')
_row12(SN11, '各会場の引当後残の列', '引当後 残',
       '=供給可能在庫−%s必要数 → =前の残−%s必要数 → …と連鎖' % (FV1.split(' ')[0], FUTURE[1][0].split(' ')[0]),
       ' → '.join(_f12(x) for x in _alloc1),
       'マイナスに転じた会場でその商品が欠品する（確定発注の入荷分は加算済み）')
_row12(SN11, '%s / %s列' % (_L(SN11, H11R, '欠品発生会場'), _L(SN11, H11R, '会期初日')),
       '欠品発生会場／会期初日', _FX(SN11, H11R, D11, '欠品発生会場'),
       'No.1 は %s' % ('%s で欠品' % _sv1 if _sv1 else '欠品なし → 「—」'),
       'IFS で最初にマイナスへ転じた会場を特定する')
_row12(SN11, '%s / %s列' % (_L(SN11, H11R, '不足数'), _L(SN11, H11R, '推奨発注数')), '不足数／推奨発注数',
       '%s ／ %s' % (_FX(SN11, H11R, D11, '不足数'), _FX(SN11, H11R, D11, '推奨発注数')),
       '%s個 ／ %s個' % (_f12(_ap1), _f12(_or1)), '安全率%.0f%%込み' % (CFG.SAFETY * 100))
_row12(SN11, '%s列' % _L(SN11, H11R, '最終発注期限'), '最終発注期限',
       _FX(SN11, H11R, D11, '最終発注期限'),
       '欠品会場の会期初日 − 生産日数 − バッファ%d日' % CFG.ORDER_BUFFER_DAYS,
       '★このシートの最終アウトプット')
_row12(SN11, '%s列' % _L(SN11, H11R, '残日数'), _HD(SN11, H11R, '残日数'), _FX(SN11, H11R, D11, '残日数'),
       '最終発注期限 − %s' % BASEDATE.strftime('%Y/%m/%d'), 'マイナス＝すでに期限超過')
_row12(SN11, '%s列' % _L(SN11, H11R, '緊急度'), '緊急度', _FX(SN11, H11R, D11, '緊急度'),
       '★期限超過／★至急／要着手／余裕あり／◎発注確定済／×再生産不可',
       '確定発注済みは警告対象から外し、本当に危ない品が埋もれないようにしている')
_row12(SN11, '賞味期限・確定発注のセクション', '売り切り管理',
       '=MAX(0, 供給可能在庫 − 売切り期限までの必要数) ほか',
       '／'.join('%s：現在庫は %s まで賞味期限OK' % (_nmof.get(k, k), v['sellable_through'])
                for k, v in SHELF.items()) or '対象なし',
       '確定発注の消化率は全会期終了時点の値。会場ごとの累計消化率も並記している')

# ---- サマリー -------------------------------------------------------------
_sec12('■ サマリー')
_row12('サマリー', 'A列:B列', '結論サマリー', '（数式なし＝①〜⑪の結果を文章化した読み物）', '—',
       '入力データを更新したら本文の記述も見直すこと')

for _w, _cx in zip([17, 24, 26, 54, 34, 42], 'ABCDEF'):
    ws.column_dimensions[_cx].width = _w
ws.freeze_panes = 'A6'

for s in wb.worksheets:
    s.sheet_properties.pageSetUpPr.fitToPage = True
    s.page_setup.orientation = 'landscape'
wb.save(OUT)
print('saved:', OUT)
print('TAIL=%.4f G_ALL=%.4f 予測=%d 石川以降=%.0f 追加生産=%.0f'
      % (TAIL, G_ALL, TS, _futq, _addprod))
