#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""各会場の「レジ通過単品売上」xlsx と実在庫xlsx を読み、レポート生成用のJSONを出力する。

会場ファイルは会場ごとに列レイアウトが違うため、列位置は決め打ちせずヘッダ文字列から
自動判定する。新しい会場のファイルが来ても、だいたいそのまま通るはず。

usage:
  python extract_data.py --out DIR \
      --venue 東京=path.xlsx --venue 名古屋=path.xlsx ... \
      --ledger path.xlsx

出力（--out ディレクトリ配下）:
  sales.json  会場別 商品別の日次販売・在庫
  att.json    会場別 日次入場者数
  stock.json  商品別 残在庫想定（在庫フロー台帳の最終「実在庫」列）
"""
import argparse
import datetime
import json
import os
import re
import sys

import openpyxl

SHEET = 'レジ通過単品売上'


def _num(x):
    return x if isinstance(x, (int, float)) else 0


def _find_header(ws):
    """商品名/JAN のあるヘッダ行を探す。返り値は (ヘッダ行, 日付行)。"""
    for r in range(1, min(ws.max_row, 40) + 1):
        if ws.cell(r, 4).value == '商品名' and ws.cell(r, 5).value == 'JAN':
            return r, r + 1
    raise RuntimeError('%s: ヘッダ行（商品名/JAN）が見つからない' % ws.title)


def _find_col(ws, row, label, start=1, end=None):
    end = end or ws.max_column
    for c in range(start, end + 1):
        if ws.cell(row, c).value == label:
            return c
    return None


def _day_columns(ws, hr):
    """ヘッダ行に 1,2,3... と連番が入っている列＝会期の日次列。

    東京の「内覧会」列のように連番でない前置列があるため、連番で拾うのが確実。
    """
    cols = []
    expect = 1
    for c in range(1, ws.max_column + 1):
        v = ws.cell(hr, c).value
        if isinstance(v, int) and v == expect:
            cols.append(c)
            expect += 1
    if not cols:
        raise RuntimeError('%s: 日次列（連番ヘッダ）が見つからない' % ws.title)
    return cols


def read_venue(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    hr, dr = _find_header(ws)
    days = _day_columns(ws, hr)
    c0, n = days[0], len(days)

    price_col = _find_col(ws, dr, '(税込)') or 7
    total_col = _find_col(ws, hr, '販売数')
    amt_col = total_col + 1 if total_col else None
    avail_col = _find_col(ws, hr, '販売可能数')
    stock_col = _find_col(ws, hr, '理論上')
    pre_col = _find_col(ws, hr, '内覧会')       # 東京のみ。会期実績からは除外する
    deliv_col = None
    for c in range(1, ws.max_column + 1):
        if ws.cell(hr, c).value == '納品' and ws.cell(dr, c).value == '合計数':
            deliv_col = c
            break
    missing = [k for k, v in [('販売数', total_col), ('販売可能数', avail_col),
                              ('理論上在庫', stock_col)] if v is None]
    if missing:
        raise RuntimeError('%s: 列が見つからない → %s' % (path, missing))

    # 入場者数の行（ラベル列は会場によって10〜12列目）
    att = [0] * n
    for r in range(1, hr):
        for c in range(1, min(ws.max_column, 15) + 1):
            if ws.cell(r, c).value == '入場者数':
                att = [_num(ws.cell(r, c0 + i).value) for i in range(n)]
                break
        if any(att):
            break

    dates = []
    for i in range(n):
        v = ws.cell(dr, c0 + i).value
        dates.append(v.date().isoformat() if isinstance(v, datetime.datetime) else None)

    rows = {}
    for r in range(dr + 1, ws.max_row + 1):
        try:
            jan = int(ws.cell(r, 5).value)
        except (TypeError, ValueError):
            continue
        rows[jan] = dict(
            name=str(ws.cell(r, 4).value).replace('　', ' ').strip(),
            cat=ws.cell(r, 3).value,
            price=_num(ws.cell(r, price_col).value),
            daily=[_num(ws.cell(r, c0 + i).value) for i in range(n)],
            pre=_num(ws.cell(r, pre_col).value) if pre_col else 0,
            total=round(_num(ws.cell(r, total_col).value)),
            amt=_num(ws.cell(r, amt_col).value) if amt_col else 0,
            deliv=round(_num(ws.cell(r, deliv_col).value)) if deliv_col else 0,
            avail=round(_num(ws.cell(r, avail_col).value)),
            stock=round(_num(ws.cell(r, stock_col).value)),
        )
    return dict(dates=dates, rows=rows), att


def check_venue(name, data):
    """日次合計＝販売数、販売可能数−販売数＝理論在庫 の整合を確認する。"""
    warn = []
    for j, x in data['rows'].items():
        if abs(sum(x['daily']) + x['pre'] - x['total']) > 0.5:
            warn.append('%s JAN%s 日次合計≠販売数' % (name, j))
        if x['avail'] and abs(x['avail'] - x['total'] - x['stock']) > 1.5:
            warn.append('%s JAN%s 販売可能数−販売数≠理論在庫' % (name, j))
    return warn


def read_stock(path, return_sheet=None):
    """実在庫ファイル。理論在庫（保管場所別）＋直近会場の返送良品を商品別に合算する。

    返送分の検品結果シートは会場ごとに複数あるので、直近1会場分だけを足す。
    どのシートを使ったかは必ず表示するので、想定と違えば --return-sheet で明示する。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    theo, ret = {}, {}
    for sn in wb.sheetnames:
        ws = wb[sn]
        if sn == '元データ' or (ws.cell(1, 1).value == 'JANコード(EAN/UPC)'):
            for r in range(2, ws.max_row + 1):
                try:
                    j = int(ws.cell(r, 1).value)
                except (TypeError, ValueError):
                    continue
                theo.setdefault(j, {})[ws.cell(r, 4).value] = _num(ws.cell(r, 6).value)

    cands = [sn for sn in wb.sheetnames if '検品結果' in sn and '返送' in sn]
    if return_sheet:
        pick = return_sheet
    elif cands:
        def _key(sn):
            m = re.match(r'(\d+)\.(\d+)', sn)
            return (int(m.group(1)), int(m.group(2))) if m else (0, 0)
        pick = max(cands, key=_key)     # 先頭の「月.日」が最大＝直近の返送分
    else:
        pick = None
    if pick:
        print('  返送良品シート: %s（候補 %s）' % (pick, ' / '.join(cands)))
        ws = wb[pick]
        for r in range(1, ws.max_row + 1):
            if ws.cell(r, 1).value == '商品ID':
                for rr in range(r + 1, ws.max_row + 1):
                    try:
                        j = int(ws.cell(rr, 1).value)
                    except (TypeError, ValueError):
                        continue
                    ret[j] = _num(ws.cell(rr, 3).value)      # 良品
                break
    out = {}
    for j in set(list(theo) + list(ret)):
        cx = theo.get(j, {}).get('CX01', 0)
        db = theo.get(j, {}).get('DB00', 0)
        rg = ret.get(j, 0)
        out[j] = dict(cx=round(cx), db=round(db), ret=round(rg), real=round(cx + db + rg))
    return out


LEDGER_SHEET_KEY = '検品差異'


def read_ledger(path, sheet=None):
    """在庫フロー台帳（【清算(卸値67％)】…_検品差異）を読む。

    会場ごとに「追加発注 → 理論在庫 → 販売数 → 理論在庫 → レッグス検品数量 → 実在庫」が
    横に並んだ台帳。**最終列（最後の会場の「実在庫」）が現時点の残在庫想定**であり、
    これが発注判断の唯一の正解。同じファイルの「元データ」「在庫合算」シートは
    池袋終了時点など古いスナップショットなので使ってはいけない。
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    sn = sheet or next((x for x in wb.sheetnames if LEDGER_SHEET_KEY in x), None)
    if sn is None:
        raise RuntimeError('%s: 台帳シート（%s を含む名前）が見つからない' % (path, LEDGER_SHEET_KEY))
    ws = wb[sn]
    hr = next((r for r in range(1, 40) if ws.cell(r, 1).value == '品目コード'), None)
    if hr is None:
        raise RuntimeError('%s/%s: ヘッダ行（品目コード）が見つからない' % (path, sn))
    stage = hr - 1                     # 「池袋開催中」「梅田まで(SET)」等の段階ラベル行

    def cols(label):
        return [c for c in range(1, ws.max_column + 1) if ws.cell(hr, c).value == label]

    c_add = cols('追加発注')           # 会場ごとの追加発注
    c_init = (cols('初回発注') or [None])[0]
    c_setn = (cols('発注総数(SET数)') or [None])[0]
    c_totn = (cols('発注総数') or [None])[0]
    c_real = [c for c in range(1, ws.max_column + 1)
              if str(ws.cell(hr, c).value or '').startswith('実在庫')]
    if not c_real:
        raise RuntimeError('%s/%s: 実在庫列が見つからない' % (path, sn))
    last = c_real[-1]
    stages = [str(ws.cell(stage, c).value or '') for c in c_real]
    print('  台帳シート: %s' % sn)
    print('  実在庫の段階: %s' % ' → '.join(stages))
    print('  ★採用する残在庫列 = %s「%s」（%s）'
          % (gcl_(last), ws.cell(hr, last).value, stages[-1]))

    # 台帳は列によって単位が混在する（SET商品の発注数・理論在庫はバラ、実在庫はSET）。
    # 単位が確実に「販売単位」で揃っているのは実在庫列だけなので、在庫はそこだけを使う。
    # 発注総数は参考値として持つが、SET商品では単位が信頼できないため判定には使わない。
    out = {}
    for r in range(hr + 1, ws.max_row + 1):
        try:
            j = int(ws.cell(r, 2).value)
        except (TypeError, ValueError):
            continue
        g = lambda c: ws.cell(r, c).value if isinstance(ws.cell(r, c).value, (int, float)) else 0
        setn, totn = (g(c_setn) if c_setn else 0), (g(c_totn) if c_totn else 0)
        size = round(totn / setn) if setn and totn and round(totn / setn) >= 1 else 1
        ordered = (g(c_init) if c_init else 0) + sum(g(c) for c in c_add)
        out[j] = dict(
            stock=round(g(last)),                       # ★採用する残在庫想定（販売単位）
            stock_stage=stages[-1],
            stages=stages,
            hist=[round(g(c)) for c in c_real],         # 会場ごとの実在庫推移（販売単位）
            set_size=size,                              # 1＝単品、2以上＝SET商品
            ordered_raw=round(ordered),                 # 参考：台帳上の累計発注（単位混在）
        )
    return out


def gcl_(c):
    from openpyxl.utils import get_column_letter
    return get_column_letter(c)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--out', required=True)
    ap.add_argument('--venue', action='append', default=[], metavar='NAME=PATH')
    ap.add_argument('--ledger', required=True,
                    help='在庫フロー台帳のxlsx（【清算…_検品差異】シートを含むもの）')
    ap.add_argument('--ledger-sheet', default=None, help='台帳シート名を明示する場合')
    a = ap.parse_args()
    os.makedirs(a.out, exist_ok=True)

    sales, att, warns = {}, {}, []
    for spec in a.venue:
        name, path = spec.split('=', 1)
        d, at = read_venue(path)
        warns += check_venue(name, d)
        sales[name] = dict(dates=d['dates'], rows={str(k): v for k, v in d['rows'].items()})
        att[name] = at
        print('%-6s 日数%3d 商品%4d %s〜%s 入場計%7s'
              % (name, len(d['dates']), len(d['rows']), d['dates'][0], d['dates'][-1],
                 format(int(sum(at)), ',')))
    json.dump(sales, open(os.path.join(a.out, 'sales.json'), 'w'), ensure_ascii=False)
    json.dump(att, open(os.path.join(a.out, 'att.json'), 'w'))

    led = read_ledger(a.ledger, a.ledger_sheet)
    json.dump({str(k): v for k, v in led.items()},
              open(os.path.join(a.out, 'stock.json'), 'w'))
    print('残在庫  商品%4d  残在庫合計%s  ／ SET商品%d件（発注数はバラ単位のため在庫判定には使わない）'
          % (len(led), format(sum(v['stock'] for v in led.values()), ','),
             sum(1 for v in led.values() if v['set_size'] > 1)))
    bad = [j for j, v in led.items() if v['stock'] < 0]
    if bad:
        warns.append('台帳の残在庫が負の品目 %d件: %s' % (len(bad), bad[:5]))

    if warns:
        print('\n[要確認] 元データの整合性:', file=sys.stderr)
        for w in warns[:20]:
            print('  ' + w, file=sys.stderr)
        print('  ...計%d件' % len(warns), file=sys.stderr)


if __name__ == '__main__':
    main()
